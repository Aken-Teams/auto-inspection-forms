"""Debug PARSE_ONLY failures."""
import sys, os, io
sys.path.insert(0, os.path.dirname(__file__))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from parsers.generic_parser import GenericParser
from services.export_service import _annotate_sheet
import logging
logging.basicConfig(level=logging.INFO, format="  %(message)s")

AU_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "AU未建立规格点检表", "AU")

files = [
    ("F-RD09AY", "F-RD09AY-SMD AU 首件检查记录表.xlsx"),
    ("F-RD09B10", "F-RD09B10_纯水电阻率检查记录表_AU.xlsx"),
    ("F-RD09EA", "F-RD09EA_AU_新弘田清洗机检查记录表.xlsx"),
    ("F-RD2123", "F-RD2123_防潮柜检查记录表.xlsx"),
]

parser = GenericParser()
for form_code, fname in files:
    filepath = os.path.join(AU_DIR, fname)
    if not os.path.exists(filepath):
        print(f"SKIP: {fname}")
        continue

    wb = load_workbook(filepath, data_only=True)
    sn = None
    for s in wb.sheetnames:
        if "汇总" not in s and (wb[s].max_row or 0) >= 3:
            sn = s
            break
    if not sn:
        wb.close()
        continue

    ws = wb[sn]
    parsed = parser._parse_impl(ws, sn)
    wb.close()

    meta = parsed["meta"]
    rm = meta["row_map"]
    print(f"\n=== {form_code} sheet={sn} ===")
    print(f"  rows={len(parsed['rows'])}, row_map={len(rm)}, judgment_col={meta.get('judgment_col')}")
    if rm:
        print(f"  First row_map: row={rm[0].get('row')}, cells_keys={list(rm[0].get('cells',{}).keys())[:5]}")
        all_rows = set()
        for r in rm:
            if r.get("row"):
                all_rows.add(r["row"])
            for cp in r.get("cells", {}).values():
                if isinstance(cp, list) and len(cp) >= 2:
                    all_rows.add(cp[0])
        if all_rows:
            print(f"  Physical rows: min={min(all_rows)}, max={max(all_rows)}")
            print(f"  header_label_row would be: {max(1, min(all_rows) - 1)}")

    # Test annotation
    wb2 = load_workbook(filepath)
    ws2 = wb2[sn]
    judged_rows = []
    for i, row in enumerate(parsed["rows"]):
        values = {}
        for k, v in row.get("values", {}).items():
            values[k] = {"raw": v, "judgment": "OK", "spec": {"spec_type": "skip"}}
        judged_rows.append({"values": values, "row_judgment": "OK"})

    _annotate_sheet(ws2, {"has_spec": True, "judged_rows": judged_rows, "meta": meta})

    # Search for 判定 header
    found_header = False
    for c in range(max(1, (ws2.max_column or 1) - 5), (ws2.max_column or 1) + 5):
        for r in range(1, 20):
            try:
                v = ws2.cell(row=r, column=c).value
                if v == "判定":
                    print(f"  Found '判定' at row={r}, col={c}")
                    found_header = True
                    # Check first few judgment cells
                    for rm_entry in rm[:3]:
                        check_row = rm_entry.get("row")
                        if check_row:
                            cell_val = ws2.cell(row=check_row, column=c).value
                            print(f"    row={check_row} value={cell_val}")
            except AttributeError:
                pass
    if not found_header:
        print(f"  '判定' NOT FOUND anywhere!")
        print(f"  ws2.max_column={ws2.max_column}")

    wb2.close()
