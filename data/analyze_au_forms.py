#!/usr/bin/env python3
"""Analyze AU inspection form Excel files to understand their structure."""

import os
import re
import sys
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Force UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

BASE_DIR = r"D:\github\auto-inspection-forms\data\AU未建立规格点检表\AU"

# Files to skip (already have built-in parsers)
SKIP_PREFIXES = ["F-QA1021", "F-RD09AA", "F-RD09AB", "F-RD09AJ", "F-RD09AK"]

# Files to analyze
TARGET_FILES = [
    "F-RD09AC_Clip Bond检查记录表.xlsx",
    "F-RD09AL_SMD(Clip）切弯脚外观抽验记录表.xlsx",
    "F-RD09AN_TMTT印字影像准确度检查记录表.xlsx",
    "F-RD09AX-声达超声波清洗机检查记录表.xlsx",
    "F-RD09AY-SMD AU 首件检查记录表.xlsx",
    "F-RD09B10_纯水电阻率检查记录表_AU.xlsx",
    "F-RD09BU-Clip Bond 外观检查记录表表.xlsx",
    "F-RD09BW_SMD AU TMTT站Vision开机首件检查记录表.xlsx",
    "F-RD09CS_Clip Bond出炉外观检查表.xlsx",
    "F-RD09EA_AU_新弘田清洗机检查记录表.xlsx",
    "F-RD09F1_温度检查记录表.xlsx",
    "F-RD09F1_湿度检查记录表.xlsx",
    "F-RD09FZ_SMD-C AU TMTT站外观检查记录表.xlsx",
    "F-RD09GA_SMD-C AU TMTT站开机首件检查记录表.xlsx",
    "F-RD09GB_SMD-C AU  TMTT封合拉力测试记录表.xlsx",
    "F-RD09Q1_锡膏放置冰箱温度检查记录表.xlsx",
    "F-RD09X7-机台调整后复机点检表.xlsx",
    "F-RD0976_S焊_清洗液添排液检查记录表.xlsx",
    "F-RD1024-SMD AU Line外观检查记录表.xlsx",
    "F-RD2123_防潮柜检查记录表.xlsx",
    "F-RD2140_烤箱作业记录表.xlsx",
]

# Equipment ID patterns
EQUIP_PATTERNS = [
    re.compile(r'[A-Z]{1,2}\d{4}-\d{3,4}'),     # e.g. W1234-0001
    re.compile(r'RD-\w{2,4}-\d{2}'),               # e.g. RD-XX-01
    re.compile(r'[A-Z]{2,3}-\d{2,4}'),             # e.g. AU-01
    re.compile(r'\d{4}-\d{4}'),                     # e.g. 1234-0001
]


def find_header_row(ws, max_scan=20):
    """Find the first row with >= 3 non-empty cells (header row)."""
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        non_empty = 0
        for col_idx in range(1, min(ws.max_column + 1, 30)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip():
                non_empty += 1
        if non_empty >= 3:
            return row_idx
    return None


def get_headers(ws, header_row, max_cols=25):
    """Get header values from the header row."""
    headers = []
    for col_idx in range(1, min(ws.max_column + 1, max_cols + 1)):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is not None:
            headers.append((get_column_letter(col_idx), str(val).strip().replace('\n', ' ')))
    return headers


def find_equip_ids(ws, max_scan_rows=10, max_scan_cols=20):
    """Search for equipment-ID-like patterns in first few rows."""
    found = []
    for row_idx in range(1, min(max_scan_rows + 1, ws.max_row + 1)):
        for col_idx in range(1, min(max_scan_cols + 1, ws.max_column + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            s = str(val).strip()
            for pat in EQUIP_PATTERNS:
                m = pat.search(s)
                if m:
                    loc = f"{get_column_letter(col_idx)}{row_idx}"
                    found.append((loc, s[:80]))
                    break
    return found


def check_judgment_column(ws, header_row):
    """Check if there's a 判定 column."""
    if header_row is None:
        return None
    for col_idx in range(1, min(ws.max_column + 1, 40)):
        val = ws.cell(row=header_row, column=col_idx).value
        if val and '判定' in str(val):
            return (get_column_letter(col_idx), str(val).strip())
    # Also check header_row + 1 (sub-headers)
    if header_row + 1 <= ws.max_row:
        for col_idx in range(1, min(ws.max_column + 1, 40)):
            val = ws.cell(row=header_row + 1, column=col_idx).value
            if val and '判定' in str(val):
                return (get_column_letter(col_idx), str(val).strip())
    return None


def get_sample_data(ws, header_row, num_rows=3):
    """Get a few rows of sample data after the header."""
    if header_row is None:
        return []
    samples = []
    start = header_row + 1
    # Check if there's a sub-header row
    for row_idx in range(start, min(start + num_rows + 2, ws.max_row + 1)):
        row_data = {}
        has_data = False
        for col_idx in range(1, min(ws.max_column + 1, 15)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                col_letter = get_column_letter(col_idx)
                row_data[col_letter] = str(val).strip()[:40]
                has_data = True
        if has_data:
            samples.append((row_idx, row_data))
        if len(samples) >= num_rows:
            break
    return samples


def get_merged_cells_in_header(ws, header_row, depth=3):
    """Get merged cell ranges in the header area."""
    merged = []
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= header_row + depth and mc.min_row >= max(1, header_row - 1):
            merged.append(str(mc))
    return merged


def scan_first_rows(ws, n=8):
    """Print the raw content of the first N rows for context."""
    rows_data = []
    for row_idx in range(1, min(n + 1, ws.max_row + 1)):
        cells = []
        for col_idx in range(1, min(ws.max_column + 1, 20)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                cells.append(f"{get_column_letter(col_idx)}:{str(val).strip()[:35]}")
        if cells:
            rows_data.append((row_idx, cells))
    return rows_data


def analyze_file(filepath, filename):
    """Analyze a single Excel file."""
    print(f"\n{'='*90}")
    print(f"FILE: {filename}")
    print(f"{'='*90}")

    try:
        wb = load_workbook(filepath, data_only=True, read_only=False)
    except Exception as e:
        print(f"  ERROR opening file: {e}")
        return

    sheet_names = wb.sheetnames
    print(f"  Sheets ({len(sheet_names)}): {sheet_names}")

    has_summary = any('汇总' in s for s in sheet_names)
    print(f"  Has 汇总 (summary) sheet: {has_summary}")

    for sname in sheet_names:
        is_summary = '汇总' in sname
        print(f"\n  --- Sheet: '{sname}' {'[SUMMARY]' if is_summary else ''} ---")

        ws = wb[sname]
        print(f"  Dimensions: {ws.dimensions} | max_row={ws.max_row}, max_col={ws.max_column}")
        print(f"  Merged cells count: {len(ws.merged_cells.ranges)}")

        if is_summary:
            # Still show basic info for summary sheets
            print(f"  (Summary sheet - showing first few rows only)")
            raw = scan_first_rows(ws, 5)
            for ridx, cells in raw:
                print(f"    Row {ridx}: {', '.join(cells[:10])}")
            continue

        # Raw first rows
        print(f"  First 8 rows (raw):")
        raw = scan_first_rows(ws, 8)
        for ridx, cells in raw:
            print(f"    Row {ridx}: {', '.join(cells[:12])}")

        # Header detection
        header_row = find_header_row(ws)
        print(f"  Detected header row: {header_row}")

        if header_row:
            headers = get_headers(ws, header_row)
            print(f"  Headers ({len(headers)} cols): ", end="")
            for col, val in headers[:20]:
                print(f"[{col}]{val}", end=" | ")
            print()

            # Check for sub-header (row below header)
            sub_row = header_row + 1
            if sub_row <= ws.max_row:
                sub_headers = get_headers(ws, sub_row)
                if sub_headers:
                    non_empty_sub = [(c, v) for c, v in sub_headers if v.strip()]
                    if len(non_empty_sub) >= 2:
                        print(f"  Sub-header row {sub_row}: ", end="")
                        for col, val in non_empty_sub[:20]:
                            print(f"[{col}]{val}", end=" | ")
                        print()

            # Merged cells in header area
            merged = get_merged_cells_in_header(ws, header_row)
            if merged:
                print(f"  Merged in header area: {merged[:10]}")

            # Judgment column
            judgment = check_judgment_column(ws, header_row)
            if judgment:
                print(f"  判定 column found: col {judgment[0]} = '{judgment[1]}'")
            else:
                print(f"  判定 column: NOT FOUND")

            # Sample data
            samples = get_sample_data(ws, header_row)
            if samples:
                print(f"  Sample data rows:")
                for ridx, data in samples:
                    items = [f"{k}:{v}" for k, v in list(data.items())[:10]]
                    print(f"    Row {ridx}: {', '.join(items)}")

        # Equipment IDs
        equip = find_equip_ids(ws)
        if equip:
            print(f"  Equipment ID patterns found:")
            for loc, val in equip:
                print(f"    {loc}: {val}")

        # Check sheet name for equipment pattern
        for pat in EQUIP_PATTERNS:
            m = pat.search(sname)
            if m:
                print(f"  Equipment ID in sheet name: {m.group()}")

    wb.close()


def main():
    print("AU Inspection Form Structure Analysis")
    print(f"Directory: {BASE_DIR}")
    print(f"Files to analyze: {len(TARGET_FILES)}")

    for filename in TARGET_FILES:
        filepath = os.path.join(BASE_DIR, filename)
        if not os.path.exists(filepath):
            print(f"\n{'='*90}")
            print(f"FILE: {filename}")
            print(f"  WARNING: File not found!")
            continue
        analyze_file(filepath, filename)

    print(f"\n\n{'='*90}")
    print("ANALYSIS COMPLETE")
    print(f"{'='*90}")


if __name__ == "__main__":
    main()
