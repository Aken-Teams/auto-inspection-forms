"""Spec management API endpoints."""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from pydantic import BaseModel
from typing import Optional
import os
import re
import uuid
import shutil
from collections import Counter
from openpyxl import load_workbook

from database import get_db
from models import FormType, FormSpec, SpecItem, UploadRecord, SpecVersion, InspectionResult
from config import SPEC_DIR
from services.spec_service import import_specs_from_excel, init_form_types
from services.spec_file_service import compute_file_hash, find_duplicate_across_all
from parsers.identifier import identify_form_type, _FORM_CODE_RE
from config import UPLOAD_DIR

router = APIRouter(prefix="/api/specs", tags=["specs"])


class SpecItemUpdate(BaseModel):
    item_name: str
    spec_type: str
    min_value: Optional[float] = None
    max_value: Optional[float] = None
    expected_text: Optional[str] = None
    threshold_value: Optional[float] = None
    threshold_operator: Optional[str] = None
    group_name: Optional[str] = None
    sub_group: Optional[str] = None


class FormSpecUpdate(BaseModel):
    equipment_name: Optional[str] = None
    items: Optional[list[SpecItemUpdate]] = None


class FormSpecPatch(BaseModel):
    equipment_name: Optional[str] = None


class FormSpecCreate(BaseModel):
    equipment_id: str
    equipment_name: str


class FormTypeCreate(BaseModel):
    form_code: str
    form_name: str
    file_pattern: Optional[str] = None
    description: Optional[str] = None


class FormTypePatch(BaseModel):
    form_name: Optional[str] = None
    file_pattern: Optional[str] = None
    description: Optional[str] = None


@router.get("/form-types")
def list_form_types(db: Session = Depends(get_db)):
    """List all form types."""
    types = db.query(FormType).all()
    return [
        {
            "id": ft.id,
            "form_code": ft.form_code,
            "form_name": ft.form_name,
            "description": ft.description,
            "file_pattern": ft.file_pattern,
            "is_builtin": ft.is_builtin,
            "spec_count": db.query(FormSpec).filter(FormSpec.form_type_id == ft.id).count(),
        }
        for ft in types
    ]


@router.post("/form-types")
def create_form_type(data: FormTypeCreate, db: Session = Depends(get_db)):
    """Create a new custom form type."""
    existing = db.query(FormType).filter(FormType.form_code == data.form_code).first()
    if existing:
        raise HTTPException(409, f"Form type {data.form_code} already exists")

    ft = FormType(
        form_code=data.form_code,
        form_name=data.form_name,
        file_pattern=data.file_pattern,
        description=data.description,
        is_builtin=False,
    )
    db.add(ft)
    db.commit()
    return {"success": True, "id": ft.id, "form_code": ft.form_code}


@router.patch("/form-types/{form_code}")
def patch_form_type(form_code: str, data: FormTypePatch, db: Session = Depends(get_db)):
    """Update a form type (name, file_pattern, description)."""
    ft = db.query(FormType).filter(FormType.form_code == form_code).first()
    if not ft:
        raise HTTPException(404, f"Form type {form_code} not found")
    if data.form_name is not None:
        ft.form_name = data.form_name
    if data.file_pattern is not None:
        ft.file_pattern = data.file_pattern
    if data.description is not None:
        ft.description = data.description
    db.commit()
    return {"success": True}


@router.delete("/form-types/{form_code}")
def delete_form_type(form_code: str, db: Session = Depends(get_db)):
    """Delete a form type and all its specs."""
    ft = db.query(FormType).filter(FormType.form_code == form_code).first()
    if not ft:
        raise HTTPException(404, f"Form type {form_code} not found")
    # Nullify FK in upload_records to avoid constraint violation
    db.query(UploadRecord).filter(UploadRecord.form_type_id == ft.id).update(
        {"form_type_id": None}, synchronize_session=False
    )
    # Clear FKs referencing form_specs, then delete specs
    spec_ids = [s.id for s in db.query(FormSpec).filter(FormSpec.form_type_id == ft.id).all()]
    if spec_ids:
        db.query(InspectionResult).filter(InspectionResult.form_spec_id.in_(spec_ids)).update(
            {"form_spec_id": None}, synchronize_session=False
        )
        db.query(SpecVersion).filter(SpecVersion.form_spec_id.in_(spec_ids)).delete(synchronize_session=False)
        db.query(SpecItem).filter(SpecItem.form_spec_id.in_(spec_ids)).delete(synchronize_session=False)
        db.query(FormSpec).filter(FormSpec.id.in_(spec_ids)).delete(synchronize_session=False)
    db.delete(ft)
    db.commit()
    # Clean up stored spec files
    spec_dir = os.path.join(SPEC_DIR, form_code)
    if os.path.isdir(spec_dir):
        shutil.rmtree(spec_dir, ignore_errors=True)
    return {"success": True}


@router.get("/form-types/{form_code}/specs")
def list_specs(form_code: str, include_items: bool = True, db: Session = Depends(get_db)):
    """List all specs for a form type."""
    form_type = db.query(FormType).filter(FormType.form_code == form_code).first()
    if not form_type:
        raise HTTPException(404, f"Form type {form_code} not found")

    specs = db.query(FormSpec).filter(
        FormSpec.form_type_id == form_type.id
    ).options(joinedload(FormSpec.items)).all()

    if not include_items:
        return [
            {
                "id": s.id,
                "equipment_id": s.equipment_id,
                "equipment_name": s.equipment_name,
                "extra_info": s.extra_info,
                "item_count": len(s.items),
            }
            for s in specs
        ]

    return [
        {
            "id": s.id,
            "equipment_id": s.equipment_id,
            "equipment_name": s.equipment_name,
            "extra_info": s.extra_info,
            "items": [
                {
                    "id": item.id,
                    "item_name": item.item_name,
                    "spec_type": item.spec_type,
                    "min_value": float(item.min_value) if item.min_value else None,
                    "max_value": float(item.max_value) if item.max_value else None,
                    "expected_text": item.expected_text,
                    "threshold_value": float(item.threshold_value) if item.threshold_value else None,
                    "threshold_operator": item.threshold_operator,
                    "display_order": item.display_order,
                    "group_name": item.group_name,
                    "sub_group": item.sub_group,
                }
                for item in sorted(s.items, key=lambda x: x.display_order)
            ],
        }
        for s in specs
    ]


@router.put("/specs/{spec_id}")
def update_spec(spec_id: int, data: FormSpecUpdate, db: Session = Depends(get_db)):
    """Update a spec and its items."""
    spec = db.query(FormSpec).get(spec_id)
    if not spec:
        raise HTTPException(404, "Spec not found")

    if data.equipment_name:
        spec.equipment_name = data.equipment_name

    if data.items is not None:
        # Clear existing items and replace
        db.query(SpecItem).filter(SpecItem.form_spec_id == spec_id).delete()
        for i, item_data in enumerate(data.items):
            item = SpecItem(
                form_spec_id=spec_id,
                item_name=item_data.item_name,
                spec_type=item_data.spec_type,
                min_value=item_data.min_value,
                max_value=item_data.max_value,
                expected_text=item_data.expected_text,
                threshold_value=item_data.threshold_value,
                threshold_operator=item_data.threshold_operator,
                display_order=i,
                group_name=item_data.group_name,
                sub_group=item_data.sub_group,
            )
            db.add(item)

    db.commit()
    return {"success": True}


@router.delete("/specs/{spec_id}")
def delete_spec(spec_id: int, db: Session = Depends(get_db)):
    """Delete a spec group and all its items."""
    spec = db.query(FormSpec).get(spec_id)
    if not spec:
        raise HTTPException(404, "Spec not found")
    db.query(SpecItem).filter(SpecItem.form_spec_id == spec_id).delete()
    db.delete(spec)
    db.commit()
    return {"success": True}


@router.patch("/specs/{spec_id}")
def patch_spec(spec_id: int, data: FormSpecPatch, db: Session = Depends(get_db)):
    """Rename a spec group (update equipment_name)."""
    spec = db.query(FormSpec).get(spec_id)
    if not spec:
        raise HTTPException(404, "Spec not found")
    if data.equipment_name is not None:
        spec.equipment_name = data.equipment_name
    db.commit()
    return {"success": True}


@router.post("/form-types/{form_code}/specs")
def create_spec(form_code: str, data: FormSpecCreate, db: Session = Depends(get_db)):
    """Create a new empty spec group."""
    form_type = db.query(FormType).filter(FormType.form_code == form_code).first()
    if not form_type:
        raise HTTPException(404, f"Form type {form_code} not found")

    existing = db.query(FormSpec).filter(
        FormSpec.form_type_id == form_type.id,
        FormSpec.equipment_id == data.equipment_id,
    ).first()
    if existing:
        raise HTTPException(409, f"Spec for equipment {data.equipment_id} already exists")

    spec = FormSpec(
        form_type_id=form_type.id,
        equipment_id=data.equipment_id,
        equipment_name=data.equipment_name,
    )
    db.add(spec)
    db.commit()
    return {"success": True, "id": spec.id}


@router.post("/import")
async def import_specs(
    form_code: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Import specs from an Excel file's 汇总 sheet (legacy, kept for compatibility)."""
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    filepath = os.path.join(UPLOAD_DIR, f"spec_{uuid.uuid4().hex}_{file.filename}")

    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)

    try:
        # Store file permanently
        from services.spec_file_service import store_spec_file
        file_info = store_spec_file(content, form_code, file.filename)

        result = import_specs_from_excel(
            db, filepath, form_code,
            source_filename=file.filename,
            stored_filepath=file_info["stored_path"],
            file_hash=file_info["file_hash"],
        )
        return result
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)


@router.post("/import/preview")
async def preview_import_endpoint(
    form_code: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Preview what an import would do WITHOUT committing changes.

    Returns validation, parsed specs, and diffs for user review.
    """
    from services.import_preview_service import preview_import

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    filepath = os.path.join(UPLOAD_DIR, f"preview_{uuid.uuid4().hex}_{file.filename}")

    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)

    try:
        result = preview_import(db, filepath, content, form_code, file.filename)
        return result
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)


def _check_content_dup(db: Session, filepath: str, form_code: str):
    """Parse the file's summary sheet and raise 409 if specs are identical to DB."""
    from services.import_preview_service import _preview_builtin, _preview_ai, BUILTIN_PARSERS
    from services.spec_file_service import check_specs_identical

    wb = load_workbook(filepath, data_only=True)
    try:
        if "汇总" not in wb.sheetnames:
            return
        ws = wb["汇总"]
        form_type = db.query(FormType).filter(FormType.form_code == form_code).first()
        if not form_type:
            return

        if form_code in BUILTIN_PARSERS:
            parsed_specs = _preview_builtin(db, ws, form_type)
        else:
            parsed_specs, _ = _preview_ai(ws, form_type)

        if parsed_specs and check_specs_identical(db, form_code, parsed_specs):
            raise HTTPException(409, "此檔案的規格內容與現有資料完全相同，無需重複匯入")
    finally:
        wb.close()


@router.post("/import/confirm")
async def confirm_import_endpoint(
    form_code: str,
    force: bool = False,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Execute import after user has reviewed the preview.

    1. Check for duplicates (blocked unless force=True)
    2. Store file permanently
    3. Create version snapshots for affected specs
    4. Import new specs
    5. Save structural fingerprint
    """
    from services.spec_file_service import store_spec_file, find_duplicate, check_specs_identical
    from services.fingerprint_service import generate_fingerprint, save_fingerprint

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    filepath = os.path.join(UPLOAD_DIR, f"confirm_{uuid.uuid4().hex}_{file.filename}")

    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)

    try:
        # Duplicate checks (skip if force=True)
        if not force:
            file_hash_check = compute_file_hash(content)

            dup_version = find_duplicate(db, form_code, file_hash_check)
            if dup_version:
                raise HTTPException(
                    409,
                    f"此檔案已匯入過 (設備: {dup_version['equipment_id']}, "
                    f"時間: {dup_version['created_at']})"
                )

            dup_file = find_duplicate_across_all(file_hash_check)
            if dup_file:
                raise HTTPException(
                    409,
                    f"此檔案內容與「{dup_file['form_code']}」中的"
                    f"「{dup_file['filename']}」完全相同"
                )

            # Content-level: parse and compare spec items against DB
            _check_content_dup(db, filepath, form_code)

        # Store file permanently
        file_info = store_spec_file(content, form_code, file.filename)

        # Import with version tracking
        result = import_specs_from_excel(
            db, filepath, form_code,
            source_filename=file.filename,
            stored_filepath=file_info["stored_path"],
            file_hash=file_info["file_hash"],
        )

        # Save structural fingerprint for future comparison
        if result.get("success"):
            try:
                from openpyxl import load_workbook as _lw
                wb2 = _lw(filepath, data_only=True)
                if "汇总" in wb2.sheetnames:
                    fp = generate_fingerprint(wb2["汇总"])
                    save_fingerprint(db, form_code, fp)
                    db.commit()
                wb2.close()
            except Exception as e:
                import logging
                logging.getLogger(__name__).warning(f"Failed to save fingerprint: {e}")

        return result
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)


@router.post("/analyze-file")
async def analyze_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Analyze an Excel file structure for creating a new form type.

    Returns detected sheets, headers, content keywords, and suggested specs.
    Also checks if the file matches an existing form type.
    """
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    filepath = os.path.join(UPLOAD_DIR, f"analyze_{uuid.uuid4().hex}_{file.filename}")

    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)

    try:
        wb = load_workbook(filepath, data_only=True)
        sheet_names = wb.sheetnames
        data_sheets = [s for s in sheet_names if s != "汇总"]

        # Extract keywords from first few rows of each sheet
        all_keywords = Counter()
        sheets_info = []

        for sn in data_sheets[:10]:  # Limit to first 10 sheets
            ws = wb[sn]
            # Extract header info
            headers = []
            header_row = None

            for row in range(1, min(20, ws.max_row + 1)):
                non_empty = 0
                for col in range(1, min(30, ws.max_column + 1)):
                    val = ws.cell(row=row, column=col).value
                    if val is not None:
                        non_empty += 1
                if non_empty >= 3:
                    header_row = row
                    break

            if header_row:
                for col in range(1, min(30, ws.max_column + 1)):
                    val = ws.cell(row=header_row, column=col).value
                    if val:
                        label = str(val).replace("\n", " ").strip()
                        if label and len(label) < 50:
                            headers.append(label)

            # Extract content keywords from first 5 rows
            content_words = []
            for row in range(1, min(6, ws.max_row + 1)):
                for col in range(1, min(20, ws.max_column + 1)):
                    val = ws.cell(row=row, column=col).value
                    if val:
                        text = str(val).strip()
                        if 2 <= len(text) <= 30 and not re.match(r'^[\d\.\-\s]+$', text):
                            content_words.append(text)
                            all_keywords[text] += 1

            # Count data rows
            data_rows = 0
            if header_row:
                for row in range(header_row + 1, ws.max_row + 1):
                    has_data = False
                    for col in range(1, min(10, ws.max_column + 1)):
                        if ws.cell(row=row, column=col).value is not None:
                            has_data = True
                            break
                    if has_data:
                        data_rows += 1

            sheets_info.append({
                "name": sn,
                "headers": headers,
                "data_rows": data_rows,
                "sample_keywords": content_words[:10],
            })

        wb.close()

        # Suggest a form name from common keywords
        common_keywords = [kw for kw, count in all_keywords.most_common(10) if count >= 1]

        # Auto-detect identification keywords (appear in most sheets)
        sheet_count = len(data_sheets)
        id_keywords = [kw for kw, count in all_keywords.items()
                       if count >= max(1, sheet_count * 0.5) and 2 <= len(kw) <= 20]

        # Check if this file matches an existing form type (uses full identification cascade)
        sheet_contents = {}
        for si in sheets_info:
            sheet_contents[si["name"]] = " ".join(si.get("sample_keywords", []))
        matched_form_code = identify_form_type(
            file.filename, sheet_names, sheet_contents, db=db
        )

        # Extract form code from filename using regex (raw extraction)
        extracted_form_code = None
        code_match = _FORM_CODE_RE.search(file.filename)
        if code_match:
            extracted_form_code = code_match.group(1).upper()

        # If identify_form_type corrected the code (e.g., F-QA10212 -> F-QA1021),
        # also update extracted_form_code to the corrected version
        if matched_form_code and extracted_form_code:
            if extracted_form_code != matched_form_code and extracted_form_code.startswith(matched_form_code):
                extracted_form_code = matched_form_code

        # Check for file content duplicate across all spec_files
        file_hash = compute_file_hash(content)
        duplicate_spec_file = find_duplicate_across_all(file_hash)

        # Only report matched_form_code if it actually exists in DB
        if matched_form_code:
            existing = db.query(FormType).filter(FormType.form_code == matched_form_code).first()
            if existing:
                # File matches an existing form type - flag it
                pass
            else:
                # The code was identified but not yet in DB - use it as extracted
                if not extracted_form_code:
                    extracted_form_code = matched_form_code
                matched_form_code = None

        return {
            "filename": file.filename,
            "total_sheets": len(data_sheets),
            "has_summary": "汇总" in sheet_names,
            "sheets": sheets_info,
            "common_keywords": common_keywords[:10],
            "suggested_id_keywords": id_keywords[:5],
            "suggested_file_pattern": os.path.splitext(file.filename)[0],
            "extracted_form_code": extracted_form_code,
            "matched_form_code": matched_form_code,
            "duplicate_spec_file": duplicate_spec_file,
        }

    finally:
        if os.path.exists(filepath):
            os.remove(filepath)


@router.post("/create-from-file")
async def create_from_file(
    form_code: str,
    form_name: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Create a new form type + import specs from a sample Excel file.

    Automatically creates:
    1. A FormType with file_pattern derived from filename
    2. If file has 汇总 sheet, auto-imports spec items from it
    """
    # Check duplicate form code
    existing = db.query(FormType).filter(FormType.form_code == form_code).first()
    if existing:
        raise HTTPException(409, f"Form type {form_code} already exists")

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    filepath = os.path.join(UPLOAD_DIR, f"sample_{uuid.uuid4().hex}_{file.filename}")

    content = await file.read()

    # Check for file content duplicate across all spec_files
    file_hash_full = compute_file_hash(content)
    dup = find_duplicate_across_all(file_hash_full)
    if dup:
        raise HTTPException(
            409,
            f"此檔案內容與已存在的「{dup['form_code']}」中的「{dup['filename']}」完全相同，請勿重複上傳"
        )
    with open(filepath, "wb") as f:
        f.write(content)

    try:
        wb = load_workbook(filepath, data_only=True)
        data_sheets = [s for s in wb.sheetnames if s != "汇总"]
        has_summary = "汇总" in wb.sheetnames

        # Use form_code as file_pattern for reliable matching
        file_pattern = re.escape(form_code)

        # Create form type
        ft = FormType(
            form_code=form_code,
            form_name=form_name,
            file_pattern=file_pattern,
            is_builtin=False,
        )
        db.add(ft)
        db.flush()

        specs_created = 0
        items_imported = 0

        # Store file permanently in spec_files/
        import hashlib
        spec_dir = os.path.join(SPEC_DIR, form_code)
        os.makedirs(spec_dir, exist_ok=True)
        file_hash = hashlib.sha256(content).hexdigest()[:12]
        stored_name = f"{file_hash}_{file.filename}"
        stored_path = os.path.join(spec_dir, stored_name)
        with open(stored_path, "wb") as sf:
            sf.write(content)

        # If file has 汇总 sheet, auto-import specs from it (creates proper spec groups + items)
        if has_summary:
            import_result = import_specs_from_excel(
                db, filepath, form_code,
                source_filename=file.filename,
                stored_filepath=stored_path,
                file_hash=file_hash,
            )
            if import_result.get("success"):
                # Count what was created by the import
                specs_created = db.query(FormSpec).filter(FormSpec.form_type_id == ft.id).count()
                items_imported = db.query(SpecItem).join(FormSpec).filter(FormSpec.form_type_id == ft.id).count()

        wb.close()
        db.commit()

        return {
            "success": True,
            "form_code": form_code,
            "specs_created": specs_created,
            "items_imported": items_imported,
            "has_summary": has_summary,
        }

    finally:
        if os.path.exists(filepath):
            os.remove(filepath)


@router.post("/init")
def initialize_form_types(db: Session = Depends(get_db)):
    """Initialize form types in database."""
    init_form_types(db)
    return {"success": True, "message": "Form types initialized"}


# ─── Version History Endpoints ───

@router.get("/specs/{spec_id}/versions")
def get_spec_versions(spec_id: int, db: Session = Depends(get_db)):
    """List all versions for a spec, newest first."""
    from services.spec_version_service import list_versions
    spec = db.query(FormSpec).get(spec_id)
    if not spec:
        raise HTTPException(404, "Spec not found")
    return list_versions(db, spec_id)


@router.get("/specs/{spec_id}/versions/{version_id}")
def get_version_detail_endpoint(spec_id: int, version_id: int, db: Session = Depends(get_db)):
    """Get full detail of a specific version including items snapshot."""
    from services.spec_version_service import get_version_detail
    detail = get_version_detail(db, version_id)
    if not detail or detail["form_spec_id"] != spec_id:
        raise HTTPException(404, "Version not found")
    return detail


@router.post("/specs/{spec_id}/versions/{version_id}/rollback")
def rollback_version_endpoint(spec_id: int, version_id: int, db: Session = Depends(get_db)):
    """Rollback a spec to a specific version."""
    from services.spec_version_service import rollback_to_version
    result = rollback_to_version(db, spec_id, version_id)
    if result.get("error"):
        raise HTTPException(400, result["error"])
    db.commit()
    return result
