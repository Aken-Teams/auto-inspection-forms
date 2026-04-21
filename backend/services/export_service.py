"""Service for exporting inspection results by annotating original Excel files."""
import io
import json
import logging
import os
import zipfile
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from models import InspectionResult, UploadRecord
from config import UPLOAD_DIR

logger = logging.getLogger(__name__)

# Color fills for judgment
FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_NG = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FONT_NG = Font(color="FF0000", bold=True)
FONT_OK = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def _get_judged_data_for_export(db: Session, result: InspectionResult, form_code: str | None) -> dict:
    """Get judged data for export, re-judging with current specs if possible.

    This ensures the export always uses the LATEST specs, even if specs were
    added/updated after the file was uploaded.
    """
    raw_data = result.raw_data if isinstance(result.raw_data, dict) else json.loads(result.raw_data or "{}")

    # Re-judge using current specs if we have raw_data with rows
    if form_code and raw_data.get("rows"):
        from services.judgment import judge_sheet_data
        judged = judge_sheet_data(db, form_code, result.equipment_id, raw_data)
        logger.info(f"Re-judged sheet {result.sheet_name}: has_spec={judged.get('has_spec')}, "
                     f"meta={'yes' if judged.get('meta') else 'no'}, "
                     f"rows={len(judged.get('judged_rows', []))}")
        return judged

    # Fallback to stored judged_data
    judged_data = result.judged_data if isinstance(result.judged_data, dict) else json.loads(result.judged_data or "{}")

    # If stored judged_data has no meta, try to get it from raw_data
    if not judged_data.get("meta") and raw_data.get("meta"):
        judged_data["meta"] = raw_data["meta"]

    return judged_data


def export_upload_results(db: Session, upload_id: int) -> io.BytesIO:
    """Export results by annotating the original uploaded Excel file.

    Opens the original file, re-judges with current specs, and writes
    OK/NG results with color coding.
    """
    upload = db.query(UploadRecord).get(upload_id)
    if not upload:
        raise ValueError(f"Upload {upload_id} not found")

    results = db.query(InspectionResult).filter(
        InspectionResult.upload_id == upload_id
    ).all()

    # Build lookup: sheet_name -> InspectionResult
    result_map = {}
    for r in results:
        result_map[r.sheet_name] = r

    # Open original file
    filepath = os.path.join(UPLOAD_DIR, upload.stored_filename)
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Original file not found: {filepath}")

    form_code = upload.form_type.form_code if upload.form_type else None

    wb = load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        if sheet_name not in result_map:
            continue

        result = result_map[sheet_name]
        ws = wb[sheet_name]
        judged_data = _get_judged_data_for_export(db, result, form_code)
        _annotate_sheet(ws, judged_data)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_result_to_excel(db: Session, result: InspectionResult) -> io.BytesIO:
    """Export a single sheet result by annotating the original file."""
    upload = db.query(UploadRecord).get(result.upload_id)
    if not upload:
        raise ValueError(f"Upload {result.upload_id} not found")

    filepath = os.path.join(UPLOAD_DIR, upload.stored_filename)
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Original file not found: {filepath}")

    form_code = upload.form_type.form_code if upload.form_type else None

    wb = load_workbook(filepath)

    if result.sheet_name in wb.sheetnames:
        ws = wb[result.sheet_name]
        judged_data = _get_judged_data_for_export(db, result, form_code)
        _annotate_sheet(ws, judged_data)

    # Remove other sheets
    for sn in wb.sheetnames:
        if sn != result.sheet_name:
            del wb[sn]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_batch_results(db: Session, upload_ids: list[int]) -> io.BytesIO:
    """Export multiple uploads as a ZIP of annotated Excel files."""
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for upload_id in upload_ids:
            upload = db.query(UploadRecord).get(upload_id)
            if not upload:
                continue

            excel_data = export_upload_results(db, upload_id)
            filename = upload.original_filename.replace(".xlsx", "_判定结果.xlsx")
            zf.writestr(filename, excel_data.read())

    zip_buffer.seek(0)
    return zip_buffer


def _annotate_sheet(ws, judged_data: dict):
    """Annotate an original worksheet with judgment results.

    1. Only color NG cells red (no green for OK — keeps the sheet clean)
    2. Always write OK/NG into a judgment column (create one if needed)
    """
    judged_rows = judged_data.get("judged_rows", [])
    has_spec = judged_data.get("has_spec", False)
    meta = judged_data.get("meta")

    logger.info(f"_annotate_sheet: has_spec={has_spec}, meta={'yes' if meta else 'no'}, "
                f"judged_rows={len(judged_rows)}")

    if not judged_rows or not meta:
        logger.warning(f"_annotate_sheet: skipping - no judged_rows or no meta")
        return

    row_map = meta.get("row_map", [])
    judgment_col = meta.get("judgment_col")

    if len(row_map) != len(judged_rows):
        logger.warning(f"_annotate_sheet: row count mismatch - row_map={len(row_map)}, "
                        f"judged_rows={len(judged_rows)}, annotating min overlap")

    # Step 1: Apply NG color fills FIRST (before any column inserts, so positions are correct)
    if has_spec:
        for i, jr in enumerate(judged_rows):
            if i >= len(row_map):
                break
            cells_map = row_map[i].get("cells", {})
            for key, val_data in jr.get("values", {}).items():
                if not isinstance(val_data, dict):
                    continue
                if val_data.get("judgment") != "NG":
                    continue
                cell_pos = cells_map.get(key)
                if not cell_pos:
                    continue
                excel_row, excel_col = cell_pos
                try:
                    cell = ws.cell(row=excel_row, column=excel_col)
                    cell.fill = FILL_NG
                    cell.font = FONT_NG
                except AttributeError:
                    pass  # Merged cell

    # Step 2: Prepare judgment column
    if not judgment_col and row_map:
        # Find first data row for header placement
        all_physical_rows = set()
        for rm in row_map:
            if rm.get("row"):
                all_physical_rows.add(rm["row"])
            for cell_pos in rm.get("cells", {}).values():
                if isinstance(cell_pos, list) and len(cell_pos) >= 2:
                    all_physical_rows.add(cell_pos[0])
        first_data_row = min(all_physical_rows) if all_physical_rows else 1

        # Find rightmost VISIBLE column with content.
        # Skip hidden columns so judgment isn't placed after a hidden gap.
        actual_rightmost = 0
        for r in range(max(1, first_data_row - 5),
                       min(first_data_row + 10, (ws.max_row or 1) + 1)):
            for c in range(min(ws.max_column or 1, 200), 0, -1):
                col_dim = ws.column_dimensions.get(get_column_letter(c))
                if col_dim and col_dim.hidden:
                    continue  # Skip hidden columns
                try:
                    if ws.cell(row=r, column=c).value is not None:
                        actual_rightmost = max(actual_rightmost, c)
                        break
                except AttributeError:
                    actual_rightmost = max(actual_rightmost, c)
                    break

        judgment_col = (actual_rightmost if actual_rightmost > 0
                        else (ws.max_column or 1)) + 1
        header_label_row = max(1, first_data_row - 1)

        # Force judgment column to be VISIBLE with a proper width
        ws.column_dimensions[get_column_letter(judgment_col)].hidden = False
        ws.column_dimensions[get_column_letter(judgment_col)].width = 10

        try:
            hcell = ws.cell(row=header_label_row, column=judgment_col)
            hcell.value = "判定"
            hcell.font = Font(bold=True)
            hcell.border = THIN_BORDER
            hcell.alignment = CENTER_ALIGN
        except AttributeError:
            pass
        logger.info(f"_annotate_sheet: created judgment column at col={judgment_col} "
                     f"(actual_rightmost={actual_rightmost}, header_row={header_label_row})")
    elif judgment_col:
        # Clear existing judgment column
        for rm in row_map:
            excel_row = rm.get("row")
            if excel_row:
                try:
                    cell = ws.cell(row=excel_row, column=judgment_col)
                    cell.value = None
                    cell.fill = PatternFill()
                except AttributeError:
                    pass

    # Step 3: Write row judgment values
    # For multi-row records (e.g. 上模/下模), merge the judgment cell across
    # all physical rows of that record to match the form's visual style.
    if judgment_col:
        for i, jr in enumerate(judged_rows):
            if i >= len(row_map):
                break
            row_judgment = jr.get("row_judgment", "SKIP")

            # Find ALL physical rows for this record (not just rm["row"])
            rm = row_map[i]
            record_rows = set()
            if rm.get("row"):
                record_rows.add(rm["row"])
            for cell_pos in rm.get("cells", {}).values():
                if isinstance(cell_pos, list) and len(cell_pos) >= 2:
                    record_rows.add(cell_pos[0])
            if not record_rows:
                continue

            min_row = min(record_rows)
            max_row = max(record_rows)

            # Set borders on ALL rows first (before merge, so they persist)
            for r in range(min_row, max_row + 1):
                try:
                    ws.cell(row=r, column=judgment_col).border = THIN_BORDER
                except AttributeError:
                    pass

            # Write value at the top row
            # SKIP = no spec matched → show "—"; OK/NG = actual judgment
            display_value = row_judgment if row_judgment in ("OK", "NG") else "—"
            try:
                judge_cell = ws.cell(row=min_row, column=judgment_col)
                judge_cell.value = display_value
                judge_cell.alignment = CENTER_ALIGN
                if row_judgment == "NG":
                    judge_cell.fill = FILL_NG
                    judge_cell.font = FONT_NG
                elif row_judgment == "OK":
                    judge_cell.font = FONT_OK
            except AttributeError:
                pass

            # Merge vertically if record spans multiple rows
            if max_row > min_row:
                try:
                    ws.merge_cells(start_row=min_row, end_row=max_row,
                                  start_column=judgment_col, end_column=judgment_col)
                except (ValueError, KeyError):
                    pass
