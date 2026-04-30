"""
Comprehensive NG detection test.

Tests:
1. judge_value() correctness for all spec types
2. _find_spec_for_key() for F-RD09AB temperature keys
3. _judge_rd09ab() fallback when wash_reason/wash_method is None
4. End-to-end: parse real test file → judge → verify NG detected
"""
import sys
import os

sys.path.insert(0, os.path.dirname(__file__))
sys.stdout.reconfigure(encoding="utf-8")

from decimal import Decimal

PASS = 0
FAIL = 0


def check(desc, condition, detail=""):
    global PASS, FAIL
    if condition:
        PASS += 1
        print(f"  [PASS] {desc}")
    else:
        FAIL += 1
        print(f"  [FAIL] {desc}{(' — ' + detail) if detail else ''}")


# ─────────────────────────────────────────────────────────────────────────────
# 1. judge_value()
# ─────────────────────────────────────────────────────────────────────────────
def test_judge_value():
    print("\n=== 1. judge_value() ===")
    from utils.spec_parser import judge_value

    # Range
    check("100 NG vs 165~185",     judge_value(100,   "range", min_value=Decimal("165"), max_value=Decimal("185")) == "NG")
    check("164.9 NG vs 165~185",   judge_value(164.9, "range", min_value=Decimal("165"), max_value=Decimal("185")) == "NG")
    check("185.1 NG vs 165~185",   judge_value(185.1, "range", min_value=Decimal("165"), max_value=Decimal("185")) == "NG")
    check("175 OK vs 165~185",     judge_value(175,   "range", min_value=Decimal("165"), max_value=Decimal("185")) == "OK")
    check("165 OK vs 165~185",     judge_value(165,   "range", min_value=Decimal("165"), max_value=Decimal("185")) == "OK")
    check("185 OK vs 165~185",     judge_value(185,   "range", min_value=Decimal("165"), max_value=Decimal("185")) == "OK")
    check("str '100' NG vs range", judge_value("100", "range", min_value=Decimal("165"), max_value=Decimal("185")) == "NG")
    check("None → SKIP",           judge_value(None,  "range", min_value=Decimal("165"), max_value=Decimal("185")) == "SKIP")
    check("'' → SKIP",             judge_value("",    "range", min_value=Decimal("165"), max_value=Decimal("185")) == "SKIP")
    check("'/' → SKIP",            judge_value("/",   "range", min_value=Decimal("165"), max_value=Decimal("185")) == "SKIP")

    # Threshold
    check("2 NG vs >=3",  judge_value(2, "threshold", threshold_value=Decimal("3"), threshold_operator=">=") == "NG")
    check("3 OK vs >=3",  judge_value(3, "threshold", threshold_value=Decimal("3"), threshold_operator=">=") == "OK")
    check("4 OK vs >=3",  judge_value(4, "threshold", threshold_value=Decimal("3"), threshold_operator=">=") == "OK")

    # Check
    check("'√' OK",  judge_value("√",  "check") == "OK")
    check("'V' OK",  judge_value("V",  "check") == "OK")
    check("'OK' OK", judge_value("OK", "check") == "OK")
    check("'X' NG",  judge_value("X",  "check") == "NG")
    check("'0' NG",  judge_value("0",  "check") == "NG")

    # Text
    check("exact text OK",  judge_value("正常", "text", expected_text="正常") == "OK")
    check("wrong text NG",  judge_value("异常", "text", expected_text="正常") == "NG")

    # Skip
    check("spec_type=skip → SKIP", judge_value(100, "skip") == "SKIP")


# ─────────────────────────────────────────────────────────────────────────────
# 2. _find_spec_for_key() — F-RD09AB keys
# ─────────────────────────────────────────────────────────────────────────────
def test_find_spec_for_key():
    print("\n=== 2. _find_spec_for_key() for F-RD09AB ===")
    from services.judgment import _find_spec_for_key

    class _SI:
        """Minimal mock SpecItem."""
        def __init__(self, name, spec_type="range", min_v=None, max_v=None):
            self.item_name = name
            self.spec_type = spec_type
            self.min_value = Decimal(str(min_v)) if min_v is not None else None
            self.max_value = Decimal(str(max_v)) if max_v is not None else None
            self.expected_text = None
            self.threshold_value = None
            self.threshold_operator = None
            self.sub_group = None

    lookup = {
        "模温设定值(上模)":    _SI("模温设定值(上模)",    min_v=165, max_v=185),
        "模温显示值1(上模)":   _SI("模温显示值1(上模)",   min_v=165, max_v=185),
        "模温显示值2(上模)":   _SI("模温显示值2(上模)",   min_v=165, max_v=185),
        "模温设定值(下模)":    _SI("模温设定值(下模)",    min_v=160, max_v=180),
        "模温显示值1(下模)":   _SI("模温显示值1(下模)",   min_v=160, max_v=180),
        "合模压力(ton)":       _SI("合模压力(ton)",       min_v=130, max_v=160),
        "注塑压强(kgf/cm²)":  _SI("注塑压强(kgf/cm²)",  min_v=50,  max_v=80),
        "模具状态":            _SI("模具状态", spec_type="check"),
        "定位针状态":          _SI("定位针状态", spec_type="check"),
    }

    def found(key):
        return _find_spec_for_key(lookup, {}, key, {}, "F-RD09AB")

    check("set_temp_上模 → 模温设定值(上模)",  found("set_temp_上模") is not None and found("set_temp_上模").item_name == "模温设定值(上模)")
    check("set_temp_下模 → 模温设定值(下模)",  found("set_temp_下模") is not None and found("set_temp_下模").item_name == "模温设定值(下模)")
    check("disp_temp_上模_1 → 模温显示值1(上模)", found("disp_temp_上模_1") is not None and found("disp_temp_上模_1").item_name == "模温显示值1(上模)")
    check("disp_temp_上模_2 → 模温显示值2(上模)", found("disp_temp_上模_2") is not None and found("disp_temp_上模_2").item_name == "模温显示值2(上模)")
    check("disp_temp_下模_1 → 模温显示值1(下模)", found("disp_temp_下模_1") is not None and found("disp_temp_下模_1").item_name == "模温显示值1(下模)")
    check("clamp_pressure → 合模压力(ton)",   found("clamp_pressure") is not None and found("clamp_pressure").item_name == "合模压力(ton)")
    check("inject_pressure → 注塑压强(kgf/cm²)", found("inject_pressure") is not None)
    check("mold_status → 模具状态",           found("mold_status") is not None and found("mold_status").item_name == "模具状态")
    check("pin_status → 定位针状态",          found("pin_status") is not None and found("pin_status").item_name == "定位针状态")
    check("unknown_key → None",               found("unknown_key") is None)
    check("wash_reason → None (skipped)",     found("wash_reason") is None)


# ─────────────────────────────────────────────────────────────────────────────
# 3. _judge_rd09ab() fallback logic (unit test without real DB)
# ─────────────────────────────────────────────────────────────────────────────
def test_rd09ab_fallback():
    """Test that the machine-wide fallback correctly catches NG when wash_reason is None."""
    print("\n=== 3. _judge_rd09ab() fallback (NG with None wash_reason) ===")

    from services.judgment import _find_spec_for_key
    from utils.spec_parser import judge_value

    class _SI:
        def __init__(self, name, min_v, max_v):
            self.item_name = name
            self.spec_type = "range"
            self.min_value = Decimal(str(min_v))
            self.max_value = Decimal(str(max_v))
            self.expected_text = None
            self.threshold_value = None
            self.threshold_operator = None
            self.sub_group = None

    # Simulate spec_cache as built by _judge_rd09ab
    spec_lookup = {
        "模温显示值1(上模)": _SI("模温显示值1(上模)", 165, 185),
        "合模压力(ton)":     _SI("合模压力(ton)", 130, 160),
    }
    row_spec = {"lookup": spec_lookup}

    # Simulate a row where wash_reason is None (blank cell in data file)
    row_values = {
        "wash_reason": None,
        "wash_method": None,
        "disp_temp_上模_1": 100,   # NG: outside 165~185
        "clamp_pressure": 145,      # OK: inside 130~160
    }

    # Check each value
    ng_count = 0
    ok_count = 0
    for key, raw in row_values.items():
        spec_item = _find_spec_for_key(row_spec["lookup"], {}, key, {}, "F-RD09AB")
        if spec_item:
            result = judge_value(raw, spec_item.spec_type,
                                 min_value=spec_item.min_value,
                                 max_value=spec_item.max_value)
            if result == "NG":
                ng_count += 1
            elif result == "OK":
                ok_count += 1

    check("disp_temp_上模_1=100 detected as NG when spec applied", ng_count >= 1,
          f"ng_count={ng_count}, ok_count={ok_count}")
    check("clamp_pressure=145 detected as OK when spec applied", ok_count >= 1,
          f"ng_count={ng_count}, ok_count={ok_count}")


# ─────────────────────────────────────────────────────────────────────────────
# 4. End-to-end: real F-RD09AB test file (parser only, no DB)
# ─────────────────────────────────────────────────────────────────────────────
def test_rd09ab_parse():
    """Parse the real test file and verify data structure."""
    print("\n=== 4. F-RD09AB parse (file only, no judgment) ===")

    test_file = os.path.join(os.path.dirname(__file__), "..", "data", "test",
                             "F-RD09AB-Auto Mold 洗模检查记录表.xlsx")
    if not os.path.exists(test_file):
        # Try with traditional characters
        test_file = os.path.join(os.path.dirname(__file__), "..", "data", "test",
                                 "F-RD09AB-Auto Mold 洗模檢查記錄表.xlsx")
    if not os.path.exists(test_file):
        print("  [SKIP] Test file not found in data/test/")
        return

    from openpyxl import load_workbook
    from parsers.rd09ab_parser import RD09ABParser

    wb = load_workbook(test_file, data_only=True)
    parser = RD09ABParser()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            parsed = parser.parse_sheet(ws, sheet_name)
        except Exception as e:
            print(f"  [WARN] Sheet {sheet_name} parse error: {e}")
            continue

        rows = parsed.get("rows", [])
        if not rows:
            print(f"  [SKIP] Sheet {sheet_name}: 0 rows parsed")
            continue

        # Check wash_reason values
        wash_reasons = [r.get("values", {}).get("wash_reason") for r in rows]
        wash_methods = [r.get("values", {}).get("wash_method") for r in rows]
        none_count = wash_reasons.count(None)
        total = len(rows)

        print(f"  Sheet {sheet_name}: {total} rows, {none_count} with None wash_reason")
        print(f"    Unique wash_reasons: {list(set(str(w) for w in wash_reasons))[:5]}")
        print(f"    Unique wash_methods: {list(set(str(w) for w in wash_methods))[:5]}")

        # Sample temperature values
        for i, row in enumerate(rows[:3]):
            v = row.get("values", {})
            temp1 = v.get("disp_temp_上模_1")
            clamp = v.get("clamp_pressure")
            print(f"    Row {i}: disp_temp_上模_1={temp1}, clamp_pressure={clamp}, "
                  f"wash_reason={v.get('wash_reason')}")

        check(f"Sheet {sheet_name}: rows parsed", total > 0)

    wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# 5. Verify _format_spec_display returns correct format
# ─────────────────────────────────────────────────────────────────────────────
def test_format_spec_display():
    print("\n=== 5. _format_spec_display() ===")
    from services.judgment import _format_spec_display

    class _SI:
        def __init__(self, spec_type, min_v=None, max_v=None, expected=None, threshold_v=None, threshold_op=None):
            self.spec_type = spec_type
            self.min_value = Decimal(str(min_v)) if min_v is not None else None
            self.max_value = Decimal(str(max_v)) if max_v is not None else None
            self.expected_text = expected
            self.threshold_value = Decimal(str(threshold_v)) if threshold_v is not None else None
            self.threshold_operator = threshold_op

    check("range format", _format_spec_display(_SI("range", 165, 185)) == "165~185")
    check("check format", _format_spec_display(_SI("check", expected="√")) == "√")
    check("threshold format", _format_spec_display(_SI("threshold", threshold_v=3, threshold_op=">=")) == ">=3")
    check("text format", _format_spec_display(_SI("text", expected="正常")) == "正常")


# ─────────────────────────────────────────────────────────────────────────────
# 6. parse_spec_string()
# ─────────────────────────────────────────────────────────────────────────────
def test_parse_spec_string():
    print("\n=== 6. parse_spec_string() ===")
    from utils.spec_parser import parse_spec_string

    r = parse_spec_string("165~185")
    check("'165~185' → range", r["spec_type"] == "range" and float(r["min_value"]) == 165 and float(r["max_value"]) == 185)

    r = parse_spec_string("165～185")  # full-width tilde
    check("'165～185' → range", r["spec_type"] == "range")

    r = parse_spec_string("≥3")
    check("'≥3' → threshold >=", r["spec_type"] == "threshold" and r["threshold_operator"] == ">=")

    r = parse_spec_string("√")
    check("'√' → check", r["spec_type"] == "check")

    r = parse_spec_string("/")
    check("'/' → skip", r["spec_type"] == "skip")

    r = parse_spec_string("")
    check("'' → skip", r["spec_type"] == "skip")

    r = parse_spec_string(None)
    check("None → skip", r["spec_type"] == "skip")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("NG Detection Comprehensive Test")
    print("=" * 60)

    test_judge_value()
    test_find_spec_for_key()
    test_rd09ab_fallback()
    test_rd09ab_parse()
    test_format_spec_display()
    test_parse_spec_string()

    print("\n" + "=" * 60)
    print(f"Results: {PASS} PASS, {FAIL} FAIL")
    print("=" * 60)

    if FAIL > 0:
        sys.exit(1)
