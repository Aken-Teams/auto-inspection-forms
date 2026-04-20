"""Service for managing specs - import from Excel and CRUD operations."""
import re
from decimal import Decimal
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from models import FormType, FormSpec, SpecItem
from utils.spec_parser import parse_spec_string


def init_form_types(db: Session):
    """Initialize form types in DB if not exist."""
    form_types = [
        {
            "form_code": "F-QA1021",
            "form_name": "离子消散设备点检记录表",
            "identifier_keywords": ["离子消散", "QA1021"],
        },
        {
            "form_code": "F-RD09AA",
            "form_name": "Auto Mold 机台检查记录表",
            "identifier_keywords": ["机台检查", "RD09AA", "Auto Mold"],
        },
        {
            "form_code": "F-RD09AB",
            "form_name": "Auto Mold 洗模检查记录表",
            "identifier_keywords": ["洗模检查", "RD09AB"],
        },
        {
            "form_code": "F-RD09AJ",
            "form_name": "RO 焊接炉检查记录表",
            "identifier_keywords": ["焊接炉", "RD09AJ"],
        },
        {
            "form_code": "F-RD09AK",
            "form_name": "SMD(Clip）切弯脚尺寸检查记录表",
            "identifier_keywords": ["切弯脚", "RD09AK"],
        },
    ]

    for ft in form_types:
        existing = db.query(FormType).filter(FormType.form_code == ft["form_code"]).first()
        if not existing:
            db.add(FormType(**ft))
    db.commit()


def import_specs_from_excel(db: Session, filepath: str, form_code: str):
    """Import specs from the 汇总 sheet of an Excel file."""
    wb = load_workbook(filepath, data_only=True)

    if "汇总" not in wb.sheetnames:
        wb.close()
        return {"error": "No 汇总 sheet found"}

    ws = wb["汇总"]
    form_type = db.query(FormType).filter(FormType.form_code == form_code).first()
    if not form_type:
        wb.close()
        return {"error": f"Form type {form_code} not found"}

    if form_code == "F-QA1021":
        _import_qa1021_specs(db, ws, form_type)
    elif form_code == "F-RD09AA":
        _import_rd09aa_specs(db, ws, form_type)
    elif form_code == "F-RD09AB":
        _import_rd09ab_specs(db, ws, form_type)
    elif form_code == "F-RD09AJ":
        _import_rd09aj_specs(db, ws, form_type)
    elif form_code == "F-RD09AK":
        _import_rd09ak_specs(db, ws, form_type)

    wb.close()
    db.commit()
    return {"success": True}


def _get_or_create_spec(db: Session, form_type_id: int, equipment_id: str,
                        equipment_name: str = None, extra_info: dict = None) -> FormSpec:
    """Get or create a FormSpec record."""
    spec = db.query(FormSpec).filter(
        FormSpec.form_type_id == form_type_id,
        FormSpec.equipment_id == equipment_id,
    ).first()
    if not spec:
        spec = FormSpec(
            form_type_id=form_type_id,
            equipment_id=equipment_id,
            equipment_name=equipment_name or equipment_id,
            extra_info=extra_info,
        )
        db.add(spec)
        db.flush()
    else:
        # Clear existing items for re-import
        db.query(SpecItem).filter(SpecItem.form_spec_id == spec.id).delete()
    return spec


def _add_spec_item(db: Session, form_spec_id: int, item_name: str, spec_str: str,
                   display_order: int = 0, group_name: str = None, sub_group: str = None):
    """Parse a spec string and add as SpecItem."""
    parsed = parse_spec_string(spec_str)
    item = SpecItem(
        form_spec_id=form_spec_id,
        item_name=item_name,
        spec_type=parsed["spec_type"],
        min_value=parsed.get("min_value"),
        max_value=parsed.get("max_value"),
        expected_text=parsed.get("expected_text"),
        threshold_value=parsed.get("threshold_value"),
        threshold_operator=parsed.get("threshold_operator"),
        display_order=display_order,
        group_name=group_name,
        sub_group=sub_group,
    )
    db.add(item)


def _cell_val(ws, row, col):
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val == "":
            return None
    return val


# ─── F-QA1021 ───
def _import_qa1021_specs(db: Session, ws, form_type):
    """Import specs for 离子消散设备点检记录表.

    汇总 structure:
    Row 2: 设备编号 | √离子风扇
    Row 3: 电源开关 | 风扇飘带 | 空气滤网 | 风扇角度 | 风扇风速
    Row 4: 开启 | 飘动 | 无破损/堵塞 | 红线区域内 | 调节钮处于规定位置
    Row 5: RD-LZ-XX | √ | √ | √ | √ | √
    """
    # This form has a universal spec for all RD-LZ-XX equipment
    spec = _get_or_create_spec(db, form_type.id, "RD-LZ-XX", "通用离子消散设备")

    items = [
        ("电源开关", "√", "离子风扇"),
        ("风扇飘带", "√", "离子风扇"),
        ("空气滤网", "√", "离子风扇"),
        ("风扇角度", "√", "离子风扇"),
        ("风扇风速", "√", "离子风扇"),
    ]

    for i, (name, spec_str, group) in enumerate(items):
        _add_spec_item(db, spec.id, name, spec_str, i, group)


# ─── F-RD09AA ───
def _import_rd09aa_specs(db: Session, ws, form_type):
    """Import specs for Auto Mold 机台检查记录表.

    汇总 has blocks per machine, each block:
    Row N: 机台：WPRN-0001
    Row N+1: headers
    Row N+2: 区分/设定值/显示值
    Row N+3: 1,2,3,4
    Row N+4: data row with specs (上模)
    Row N+5: data row with specs (下模)
    """
    row = 1
    while row <= ws.max_row:
        val = _cell_val(ws, row, 1)
        if val and "机台" in str(val):
            # Extract machine ID
            match = re.search(r"(WP\w+-\d+)", str(val))
            if match:
                machine_id = match.group(1)

                # Find the spec data rows (skip headers, find the row with product type)
                spec_row = row + 4  # typically 4 rows down
                product = _cell_val(ws, spec_row, 2)  # B column
                mold_no = _cell_val(ws, spec_row, 3)  # C column

                extra = {"product_type": str(product) if product else "", "mold_no": str(mold_no) if mold_no else ""}
                spec = _get_or_create_spec(db, form_type.id, machine_id, f"机台{machine_id}", extra)

                order = 0
                # D: 合模压力, E: 注塑压强, F: 固化时间, G: 注塑时间, H: 预热台温度
                params = [
                    (4, "合模压力(ton)"),
                    (5, "注塑压强(kgf/cm²)"),
                    (6, "固化时间(sec)"),
                    (7, "注塑时间(sec)"),
                    (8, "预热台温度(℃)"),
                ]
                for col, name in params:
                    spec_str = _cell_val(ws, spec_row, col)
                    if spec_str:
                        _add_spec_item(db, spec.id, name, str(spec_str), order, "参数")
                        order += 1

                # Temperature specs: K: 设定值, L-O: 显示值1-4
                for offset, pos in enumerate(["上模", "下模"]):
                    r = spec_row + offset
                    # 设定值
                    set_val = _cell_val(ws, r, 11)  # K column
                    if set_val:
                        _add_spec_item(db, spec.id, f"模温设定值({pos})", str(set_val), order, "模温", pos)
                        order += 1
                    # 显示值 1-4
                    for i, col in enumerate([12, 13, 14, 15]):  # L, M, N, O
                        disp_val = _cell_val(ws, r, col)
                        if disp_val:
                            _add_spec_item(db, spec.id, f"模温显示值{i+1}({pos})", str(disp_val), order, "模温", pos)
                            order += 1
        row += 1


# ─── F-RD09AB ───
def _import_rd09ab_specs(db: Session, ws, form_type):
    """Import specs for Auto Mold 洗模检查记录表."""
    row = 1
    while row <= ws.max_row:
        # Find machine header
        val_a = _cell_val(ws, row, 1)
        val_b = _cell_val(ws, row, 2)

        if val_a and "机台" in str(val_a) and val_b:
            machine_id = str(val_b).strip()

            # Parse spec rows for this machine (multiple wash types: A-X, A-Y, A-Z)
            header_row = row + 1  # headers
            data_start = row + 4  # first data row

            r = data_start
            while r <= ws.max_row:
                reason = _cell_val(ws, r, 2)  # B: 洗模原因
                if not reason:
                    break

                method = _cell_val(ws, r, 3)  # C: 洗模方式
                wash_key = f"{machine_id}_{reason}_{method}"

                extra = {"wash_reason": str(reason), "wash_method": str(method)}
                spec = _get_or_create_spec(db, form_type.id, wash_key, f"{machine_id} {reason}-{method}", extra)

                order = 0
                # D: 模数, E: 固化时间, F: 模具号, G: 合模压力, H: 注塑压强
                params = [
                    (7, "合模压力(ton)"),
                    (8, "注塑压强(kgf/cm²)"),
                ]
                for col, name in params:
                    spec_str = _cell_val(ws, r, col)
                    if spec_str:
                        _add_spec_item(db, spec.id, name, str(spec_str), order, "参数")
                        order += 1

                # I: 上/下模温
                for offset, pos in enumerate(["上模", "下模"]):
                    cur_r = r + offset
                    # J: 设定值, K-N: 显示值1-4
                    set_val = _cell_val(ws, cur_r, 10)  # J
                    if set_val:
                        _add_spec_item(db, spec.id, f"模温设定值({pos})", str(set_val), order, "模温", pos)
                        order += 1
                    for i, col in enumerate([11, 12, 13, 14]):  # K, L, M, N
                        disp_val = _cell_val(ws, cur_r, col)
                        if disp_val:
                            _add_spec_item(db, spec.id, f"模温显示值{i+1}({pos})", str(disp_val), order, "模温", pos)
                            order += 1

                # O-S: 外观确认 (1st-5th)
                for i, col in enumerate([15, 16, 17, 18, 19]):
                    app_val = _cell_val(ws, r, col)
                    if app_val:
                        _add_spec_item(db, spec.id, f"外观确认{i+1}", str(app_val), order, "外观")
                        order += 1

                # T: 模具状态, U: 定位针状态
                for col, name in [(20, "模具状态"), (21, "定位针状态")]:
                    status_val = _cell_val(ws, r, col)
                    if status_val:
                        _add_spec_item(db, spec.id, name, str(status_val), order, "状态")
                        order += 1

                r += 2  # skip 下模 row, move to next wash type

            row = r
        else:
            row += 1


# ─── F-RD09AJ ───
def _import_rd09aj_specs(db: Session, ws, form_type):
    """Import specs for RO焊接炉检查记录表.

    汇总 has blocks per furnace group:
    Row N: 焊接炉编号 | WCBA-XXXX
    Row N+1: headers (温度设定SV, 实际温度, 氮气, 冷却水, 判定)
    Row N+2: sub-headers (温区, gas names)
    Row N+3: zone numbers 1-8
    Row N+4: spec values
    """
    row = 1
    while row <= ws.max_row:
        val = _cell_val(ws, row, 1)
        if val and "焊接炉编号" in str(val):
            furnace_id = str(_cell_val(ws, row, 2) or "").strip()
            if not furnace_id:
                row += 1
                continue

            spec = _get_or_create_spec(db, form_type.id, furnace_id, f"焊接炉{furnace_id}")

            # Spec values are in row+4 (4 rows down from header)
            spec_row = row + 4
            order = 0

            # B-I: SV temp zones 1-8
            for i, col in enumerate(range(2, 10)):  # B=2 to I=9
                spec_str = _cell_val(ws, spec_row, col)
                if spec_str:
                    _add_spec_item(db, spec.id, f"温区{i+1}设定SV(℃)", str(spec_str), order, "温度设定", f"温区{i+1}")
                    order += 1

            # J-Q: PV actual temp zones 1-8
            for i, col in enumerate(range(10, 18)):  # J=10 to Q=17
                spec_str = _cell_val(ws, spec_row, col)
                if spec_str:
                    _add_spec_item(db, spec.id, f"温区{i+1}实际PV(℃)", str(spec_str), order, "实际温度", f"温区{i+1}")
                    order += 1

            # R-W: Gas values (6 columns)
            gas_names_row = row + 2  # sub-header with gas names
            for col in range(18, 24):  # R=18 to W=23
                gas_name = _cell_val(ws, gas_names_row, col)
                spec_str = _cell_val(ws, spec_row, col)
                if spec_str:
                    label = str(gas_name) if gas_name else f"氮气{col-17}"
                    _add_spec_item(db, spec.id, label, str(spec_str), order, "氮气")
                    order += 1

            # X: 冷却水流量
            cooling_spec = _cell_val(ws, spec_row, 24)  # X=24
            if cooling_spec:
                _add_spec_item(db, spec.id, "冷却水流量LPM", str(cooling_spec), order, "冷却水")
                order += 1

            row = spec_row + 1
        else:
            row += 1


# ─── F-RD09AK ───
def _import_rd09ak_specs(db: Session, ws, form_type):
    """Import specs for SMD(Clip）切弯脚尺寸检查记录表.

    汇总 has blocks per package/machine:
    Row N: Package：SMA-C
    Row N+1: 机台编号：WTFB-0004
    Row N+2: headers (部位, 测量值, 判定)
    Row N+3: measurement column numbers
    Row N+4: part A with specs
    Row N+5-8: G1-G4 with specs
    """
    row = 1
    while row <= ws.max_row:
        val = _cell_val(ws, row, 1)
        if val and "Package" in str(val):
            package_match = re.search(r"Package[：:]\s*(\S+)", str(val))
            package = package_match.group(1) if package_match else ""

            machine_row = row + 1
            machine_val = _cell_val(ws, machine_row, 1)
            machine_match = re.search(r"(WTFB-\d+)", str(machine_val) if machine_val else "")
            machine_id = machine_match.group(1) if machine_match else ""

            if not machine_id:
                row += 1
                continue

            spec_key = f"{machine_id}_{package}"
            extra = {"package": package, "machine_id": machine_id}
            spec = _get_or_create_spec(db, form_type.id, spec_key, f"{machine_id} {package}", extra)

            # Find measurement count from number row
            num_row = row + 3
            meas_count = 0
            for col in range(3, 30):
                val = _cell_val(ws, num_row, col)
                if val and str(val).strip().isdigit():
                    meas_count = int(str(val).strip())
                else:
                    break

            # Read spec rows for parts A, G1-G4
            order = 0
            data_start = row + 4
            for part_row in range(data_start, data_start + 5):
                if part_row > ws.max_row:
                    break
                part_name = _cell_val(ws, part_row, 2)  # B column
                if not part_name:
                    continue

                # First measurement column spec (same for all columns)
                spec_str = _cell_val(ws, part_row, 3)  # C column = first measurement
                if spec_str:
                    for meas_num in range(1, meas_count + 1):
                        _add_spec_item(
                            db, spec.id,
                            f"meas_{meas_num}",
                            str(spec_str),
                            order, "测量", str(part_name),
                        )
                        order += 1

            row = data_start + 5
        else:
            row += 1
