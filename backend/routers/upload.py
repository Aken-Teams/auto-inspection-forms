"""Upload API endpoints."""
import os
import re
import uuid
import json
import logging
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from database import get_db
from models import UploadRecord, InspectionResult, FormType
from parsers.identifier import identify_form_type, get_form_type_from_db, extract_equipment_id_from_sheet
from parsers.qa1021_parser import QA1021Parser
from parsers.rd09aa_parser import RD09AAParser
from parsers.rd09ab_parser import RD09ABParser
from parsers.rd09aj_parser import RD09AJParser
from parsers.rd09ak_parser import RD09AKParser
from parsers.generic_parser import GenericParser
from services.judgment import judge_sheet_data
from services.ai_service import identify_form_type_ai, extract_form_name_ai, is_ai_available
from config import UPLOAD_DIR

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/upload", tags=["upload"])

PARSERS = {
    "F-QA1021": QA1021Parser(),
    "F-RD09AA": RD09AAParser(),
    "F-RD09AB": RD09ABParser(),
    "F-RD09AJ": RD09AJParser(),
    "F-RD09AK": RD09AKParser(),
}
GENERIC_PARSER = GenericParser()


@router.post("")
async def upload_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Upload a single Excel file for inspection."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only Excel files (.xlsx) are supported")

    # Save file
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    stored_name = f"{uuid.uuid4().hex}_{file.filename}"
    filepath = os.path.join(UPLOAD_DIR, stored_name)

    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)

    # Create upload record
    batch_id = uuid.uuid4().hex
    upload = UploadRecord(
        original_filename=file.filename,
        stored_filename=stored_name,
        batch_id=batch_id,
        status="processing",
    )
    db.add(upload)
    db.flush()

    try:
        results = _process_file(db, upload, filepath, file.filename)
        upload.status = "completed"
        db.commit()

        return {
            "upload_id": upload.id,
            "filename": file.filename,
            "form_type": upload.form_type.form_code if upload.form_type else None,
            "form_name": upload.form_type.form_name if upload.form_type else "未识别",
            "total_sheets": upload.total_sheets,
            "processed_sheets": upload.processed_sheets,
            "results": results,
        }
    except Exception as e:
        upload.status = "error"
        upload.error_message = str(e)
        db.commit()
        raise HTTPException(500, f"Processing error: {str(e)}")


@router.post("/batch")
async def upload_batch(files: list[UploadFile] = File(...), db: Session = Depends(get_db)):
    """Upload multiple Excel files for batch inspection."""
    batch_id = uuid.uuid4().hex
    results = []
    for file in files:
        if not file.filename.endswith((".xlsx", ".xls")):
            results.append({"filename": file.filename, "error": "Not an Excel file"})
            continue

        os.makedirs(UPLOAD_DIR, exist_ok=True)
        stored_name = f"{uuid.uuid4().hex}_{file.filename}"
        filepath = os.path.join(UPLOAD_DIR, stored_name)

        content = await file.read()
        with open(filepath, "wb") as f:
            f.write(content)

        upload = UploadRecord(
            original_filename=file.filename,
            stored_filename=stored_name,
            batch_id=batch_id,
            status="processing",
        )
        db.add(upload)
        db.flush()

        try:
            file_results = _process_file(db, upload, filepath, file.filename)
            upload.status = "completed"
            results.append({
                "upload_id": upload.id,
                "filename": file.filename,
                "form_type": upload.form_type.form_code if upload.form_type else None,
                "form_name": upload.form_type.form_name if upload.form_type else "未识别",
                "total_sheets": upload.total_sheets,
                "processed_sheets": upload.processed_sheets,
                "results_summary": {
                    "ok": sum(1 for r in file_results if r["overall_result"] == "OK"),
                    "ng": sum(1 for r in file_results if r["overall_result"] == "NG"),
                    "no_spec": sum(1 for r in file_results if r["overall_result"] == "NO_SPEC"),
                },
            })
        except Exception as e:
            upload.status = "error"
            upload.error_message = str(e)
            results.append({"filename": file.filename, "error": str(e)})

    db.commit()
    return {"batch_id": batch_id, "total_files": len(files), "results": results}


def _process_file(db: Session, upload: UploadRecord, filepath: str, filename: str) -> list:
    """Process an uploaded Excel file."""
    wb = load_workbook(filepath, data_only=True)
    sheet_names = wb.sheetnames

    # Identify form type
    # Get content from first non-汇总 sheet for identification
    sheet_contents = {}
    for sn in sheet_names:
        if sn == "汇总":
            continue
        ws = wb[sn]
        text_parts = []
        for row in range(1, min(10, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                val = ws.cell(row=row, column=col).value
                if val:
                    text_parts.append(str(val))
        sheet_contents[sn] = " ".join(text_parts)

    form_code = identify_form_type(filename, sheet_names, sheet_contents, db=db)

    if form_code:
        form_type = get_form_type_from_db(db, form_code)
        if not form_type:
            # Auto-create form type for newly identified form codes
            form_type = _auto_create_form_type(db, form_code, filename, sheet_contents)
        if form_type:
            upload.form_type_id = form_type.id

    # Process each sheet (skip 汇总)
    data_sheets = [s for s in sheet_names if s != "汇总"]
    upload.total_sheets = len(data_sheets)

    parser = PARSERS.get(form_code, GENERIC_PARSER)
    results = []

    for sheet_name in data_sheets:
        ws = wb[sheet_name]
        try:
            parsed = parser.parse_sheet(ws, sheet_name)
            equipment_id = (extract_equipment_id_from_sheet(sheet_name, form_code) if form_code else None) or sheet_name

            # Judge against specs
            if form_code:
                judgment = judge_sheet_data(db, form_code, equipment_id, parsed)
            else:
                judgment = {
                    "has_spec": False,
                    "form_spec_id": None,
                    "overall_result": "NO_SPEC",
                    "judged_rows": [],
                    "summary": {"total": 0, "ok": 0, "ng": 0, "skip": 0},
                }

            # Store result
            result = InspectionResult(
                upload_id=upload.id,
                sheet_name=sheet_name,
                equipment_id=equipment_id,
                form_spec_id=judgment.get("form_spec_id"),
                has_spec=judgment.get("has_spec", False),
                overall_result=judgment.get("overall_result", "NO_SPEC"),
                raw_data=parsed,
                judged_data=judgment,
                inspection_date=parsed.get("inspection_date", ""),
            )
            db.add(result)
            upload.processed_sheets += 1

            results.append({
                "sheet_name": sheet_name,
                "equipment_id": equipment_id,
                "has_spec": judgment.get("has_spec", False),
                "overall_result": judgment.get("overall_result", "NO_SPEC"),
                "summary": judgment.get("summary", {}),
            })
        except Exception as e:
            result = InspectionResult(
                upload_id=upload.id,
                sheet_name=sheet_name,
                equipment_id=sheet_name,
                overall_result="ERROR",
                raw_data={"error": str(e)},
                judged_data={"error": str(e)},
            )
            db.add(result)
            results.append({
                "sheet_name": sheet_name,
                "overall_result": "ERROR",
                "error": str(e),
            })

    wb.close()
    db.flush()
    return results


def _auto_create_form_type(db: Session, form_code: str, filename: str, sheet_contents: dict) -> FormType | None:
    """Auto-create a new FormType when an unknown form code is identified."""
    # Check if already exists (race condition guard)
    existing = db.query(FormType).filter(FormType.form_code == form_code).first()
    if existing:
        return existing

    try:
        # Try to get a meaningful name via AI
        form_name = None
        if is_ai_available() and sheet_contents:
            sample = "\n".join(f"[{k}] {v}" for k, v in list(sheet_contents.items())[:2])
            form_name = extract_form_name_ai(filename, sample)

        if not form_name:
            # Derive name from filename: remove form code, extension, and UUID prefix
            name = filename.replace(".xlsx", "").replace(".xls", "")
            # Remove form code from name
            name = name.replace(form_code, "").strip(" -_")
            form_name = name if name else form_code

        # Build a file_pattern from the form code
        file_pattern = re.escape(form_code)

        form_type = FormType(
            form_code=form_code,
            form_name=form_name,
            file_pattern=file_pattern,
            is_builtin=False,
        )
        db.add(form_type)
        db.flush()
        logger.info(f"Auto-created form type: {form_code} - {form_name}")
        return form_type
    except Exception as e:
        logger.error(f"Failed to auto-create form type {form_code}: {e}")
        db.rollback()
        # Try to find existing after IntegrityError (concurrent creation)
        existing = db.query(FormType).filter(FormType.form_code == form_code).first()
        return existing
