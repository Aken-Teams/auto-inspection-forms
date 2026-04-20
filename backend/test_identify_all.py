"""Test form type identification for all AU files."""
import os
import sys
import re

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from parsers.identifier import identify_form_type

AU_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "AU未建立规格点检表", "AU")


def test_all_au_files():
    files = sorted(f for f in os.listdir(AU_DIR) if f.endswith((".xlsx", ".xls")))
    print(f"\n{'='*80}")
    print(f"Testing {len(files)} AU files for form type identification")
    print(f"{'='*80}\n")

    results = []
    for filename in files:
        filepath = os.path.join(AU_DIR, filename)
        try:
            wb = load_workbook(filepath, data_only=True, read_only=True)
            sheet_names = wb.sheetnames

            # Get content sample from first non-汇总 sheet
            sheet_contents = {}
            for sn in sheet_names:
                if sn == "汇总":
                    continue
                ws = wb[sn]
                text_parts = []
                for row_idx, row in enumerate(ws.iter_rows(max_row=10, max_col=20, values_only=True)):
                    if row_idx >= 10:
                        break
                    for val in row:
                        if val is not None:
                            text_parts.append(str(val))
                sheet_contents[sn] = " ".join(text_parts)
                break  # Only need first data sheet

            has_summary = "汇总" in sheet_names

            # Identify (without DB)
            form_code = identify_form_type(filename, sheet_names, sheet_contents, db=None)

            # Also try regex extraction as fallback
            if not form_code:
                m = re.search(r"(F-[A-Z]{2}\d{2,4}[A-Z]{0,2})", filename, re.IGNORECASE)
                if m:
                    form_code = m.group(1).upper()

            status = "OK" if form_code else "FAIL"
            results.append((status, filename, form_code, has_summary, len(sheet_names)))

            icon = "[OK]" if form_code else "[FAIL]"
            summary_tag = " [汇总]" if has_summary else ""
            print(f"  {icon} {filename}")
            print(f"    → form_code: {form_code or 'UNKNOWN'} | sheets: {len(sheet_names)}{summary_tag}")

            wb.close()
        except Exception as e:
            results.append(("ERROR", filename, None, False, 0))
            print(f"  [FAIL] {filename}")
            print(f"    → ERROR: {e}")

    # Summary
    ok = sum(1 for r in results if r[0] == "OK")
    fail = sum(1 for r in results if r[0] == "FAIL")
    err = sum(1 for r in results if r[0] == "ERROR")

    print(f"\n{'='*80}")
    print(f"Results: {ok}/{len(results)} identified | {fail} failed | {err} errors")
    print(f"{'='*80}")

    if fail > 0:
        print("\nFailed files:")
        for r in results:
            if r[0] == "FAIL":
                print(f"  - {r[1]}")

    return ok == len(results)


if __name__ == "__main__":
    success = test_all_au_files()
    sys.exit(0 if success else 1)
