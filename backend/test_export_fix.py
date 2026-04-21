"""Test the export_service._annotate_sheet fix against real Excel files.

This script:
1. Parses each form type to get row_map + judged data
2. Calls _annotate_sheet on a copy of the workbook
3. Verifies the judgment column was actually written (not silently skipped)
4. Saves the annotated file for manual inspection
"""
import os
import sys
import io

sys.path.insert(0, os.path.dirname(__file__))

from openpyxl import load_workbook
from copy import copy

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "test")

FILES = {
    "QA1021": os.path.join(DATA_DIR, "F-QA1021_离子消散设备点检记录表.xlsx"),
    "RD09AA": os.path.join(DATA_DIR, "F-RD09AA-Auto Mold 机台检查记录表.xlsx"),
    "RD09AB": os.path.join(DATA_DIR, "F-RD09AB-Auto Mold 洗模检查记录表.xlsx"),
    "RD09AJ": os.path.join(DATA_DIR, "F-RD09AJ-RO 焊接炉检查记录表.xlsx"),
    "RD09AK": os.path.join(DATA_DIR, "F-RD09AK_SMD(Clip）切弯脚尺寸检查记录表.xlsx"),
}


def _col_letter(col):
    result = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _run_parser(form_code, ws, sheet_name):
    if form_code == "QA1021":
        from parsers.qa1021_parser import QA1021Parser
        return QA1021Parser()._parse_impl(ws, sheet_name)
    elif form_code == "RD09AA":
        from parsers.rd09aa_parser import RD09AAParser
        return RD09AAParser()._parse_impl(ws, sheet_name)
    elif form_code == "RD09AB":
        from parsers.rd09ab_parser import RD09ABParser
        return RD09ABParser()._parse_impl(ws, sheet_name)
    elif form_code == "RD09AJ":
        from parsers.rd09aj_parser import RD09AJParser
        return RD09AJParser()._parse_impl(ws, sheet_name)
    elif form_code == "RD09AK":
        from parsers.rd09ak_parser import RD09AKParser
        return RD09AKParser()._parse_impl(ws, sheet_name)
    return None


def _build_fake_judged_data(parsed_data):
    """Build fake judged_data with alternating OK/NG for testing."""
    rows = parsed_data.get("rows", [])
    meta = parsed_data.get("meta", {})
    judged_rows = []
    for i, row in enumerate(rows):
        values = {}
        for key, val in row.get("values", {}).items():
            raw = val if not isinstance(val, dict) else val.get("raw", val)
            values[key] = {
                "raw": raw,
                "judgment": "NG" if i % 5 == 0 else "OK",
                "spec": {"spec_type": "skip"}
            }
        row_judgment = "NG" if i % 5 == 0 else "OK"
        judged_rows.append({
            "values": values,
            "row_judgment": row_judgment,
        })
    return {
        "has_spec": True,
        "judged_rows": judged_rows,
        "meta": meta,
    }


def test_form(form_code, filepath):
    print(f"\n{'='*70}")
    print(f"  Testing {form_code}")
    print(f"{'='*70}")

    if not os.path.exists(filepath):
        print(f"  SKIP: file not found")
        return False

    # Parse
    wb_read = load_workbook(filepath, data_only=True)
    sheet_name = None
    for sn in wb_read.sheetnames:
        if "汇总" not in sn and "summary" not in sn.lower():
            sheet_name = sn
            break
    if not sheet_name:
        print(f"  SKIP: no data sheet found")
        wb_read.close()
        return False

    ws_read = wb_read[sheet_name]
    parsed = _run_parser(form_code, ws_read, sheet_name)
    wb_read.close()

    if not parsed or not parsed.get("meta"):
        print(f"  SKIP: parser returned no meta")
        return False

    meta = parsed["meta"]
    row_map = meta.get("row_map", [])
    judgment_col_from_parser = meta.get("judgment_col")
    print(f"  Parser: judgment_col={judgment_col_from_parser}, row_map_entries={len(row_map)}")

    # Build fake judged data
    judged_data = _build_fake_judged_data(parsed)

    # Open writable copy and annotate
    wb = load_workbook(filepath)
    ws = wb[sheet_name]

    # Import and call _annotate_sheet
    import logging
    logging.basicConfig(level=logging.INFO, format="  %(message)s")
    from services.export_service import _annotate_sheet
    _annotate_sheet(ws, judged_data)

    # Verify: find cells with "OK" or "NG" values
    ok_count = 0
    ng_count = 0
    judgment_col_found = None

    # Find the actual judgment column by scanning for "判定" header
    check_col = None
    if judgment_col_from_parser:
        check_col = judgment_col_from_parser
    else:
        # Search for "判定" in the header area
        for c in range(1, (ws.max_column or 1) + 1):
            for r in range(1, 15):
                try:
                    if ws.cell(row=r, column=c).value == "判定":
                        check_col = c
                        break
                except AttributeError:
                    pass
                if check_col:
                    break
        if not check_col:
            print(f"  ERROR: Could not find '判定' header in sheet")
            wb.close()
            return False

    # Check each row_map entry for judgment values
    # For multi-row records, the value is at min_row (top of merge), not rm["row"]
    dash_count = 0
    for rm in row_map:
        # Find the actual min_row for this record (top of merged range)
        record_rows = set()
        if rm.get("row"):
            record_rows.add(rm["row"])
        for cell_pos in rm.get("cells", {}).values():
            if isinstance(cell_pos, list) and len(cell_pos) >= 2:
                record_rows.add(cell_pos[0])
        if not record_rows:
            continue
        check_row = min(record_rows)
        try:
            cell = ws.cell(row=check_row, column=check_col)
            val = cell.value
            if val == "OK":
                ok_count += 1
            elif val == "NG":
                ng_count += 1
            elif val == "—":
                dash_count += 1
        except AttributeError:
            pass

    total = ok_count + ng_count + dash_count
    print(f"\n  Results at col {check_col} ({_col_letter(check_col)}):")
    print(f"    OK cells: {ok_count}")
    print(f"    NG cells: {ng_count}")
    print(f"    — (no spec) cells: {dash_count}")
    print(f"    Total judgment cells: {total}")

    # Check header
    header_found = False
    for r in range(1, 10):
        try:
            val = ws.cell(row=r, column=check_col).value
            if val == "判定":
                header_found = True
                print(f"    '判定' header at row {r}")
                break
        except AttributeError:
            pass

    # Check border on first judgment cell
    border_ok = False
    for rm in row_map:
        excel_row = rm.get("row")
        if not excel_row:
            continue
        try:
            cell = ws.cell(row=excel_row, column=check_col)
            if cell.value in ("OK", "NG"):
                has_border = (cell.border and cell.border.left and
                             cell.border.left.style is not None)
                border_ok = has_border
                print(f"    Border on first judgment cell: {'YES' if has_border else 'NO'}")
                break
        except AttributeError:
            pass

    # Save output file
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, f"{form_code}_test_export.xlsx")
    wb.save(output_path)
    wb.close()
    print(f"    Saved to: {output_path}")

    # Verdict
    success = total > 0 and header_found
    if success:
        print(f"\n  >>> PASS: {total} judgment values written, header found")
    else:
        problems = []
        if total == 0:
            problems.append("NO judgment values written")
        if not header_found:
            problems.append("No '判定' header found")
        print(f"\n  >>> FAIL: {', '.join(problems)}")

    return success


if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    print("Testing export_service._annotate_sheet fix")
    print(f"Output dir: {OUTPUT_DIR}")

    results = {}
    for form_code, filepath in FILES.items():
        results[form_code] = test_form(form_code, filepath)

    print(f"\n{'='*70}")
    print("  SUMMARY")
    print(f"{'='*70}")
    for form_code, passed in results.items():
        status = "PASS" if passed else "FAIL"
        print(f"  {form_code}: {status}")

    all_passed = all(results.values())
    print(f"\n  Overall: {'ALL PASSED' if all_passed else 'SOME FAILED'}")
