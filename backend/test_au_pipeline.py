"""Comprehensive test for the AU file pipeline.

Tests 3 stages for each file:
1. GenericParser parsing (data extraction)
2. header_spec_extractor (spec extraction)
3. Export annotation (_annotate_sheet)

Also tests the original 5 built-in types for regression.
"""
import io
import os
import sys
import re
import json

sys.path.insert(0, os.path.dirname(__file__))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

AU_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "AU未建立规格点检表", "AU")
DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")

# Original 5 built-in files
BUILTIN_FILES = {
    "F-QA1021": os.path.join(DATA_DIR, "F-QA1021_离子消散设备点检记录表.xlsx"),
    "F-RD09AA": os.path.join(DATA_DIR, "F-RD09AA-Auto Mold 机台检查记录表.xlsx"),
    "F-RD09AB": os.path.join(DATA_DIR, "F-RD09AB-Auto Mold 洗模检查记录表.xlsx"),
    "F-RD09AJ": os.path.join(DATA_DIR, "F-RD09AJ-RO 焊接炉检查记录表.xlsx"),
    "F-RD09AK": os.path.join(DATA_DIR, "F-RD09AK_SMD(Clip）切弯脚尺寸检查记录表.xlsx"),
}

# Map filename -> form_code
_FORM_CODE_RE = re.compile(r"(F-[A-Z]{2}\d{2,4}[A-Z0-9]{0,3})")


def extract_form_code(filename):
    m = _FORM_CODE_RE.search(filename)
    return m.group(1) if m else None


def test_generic_parser(filepath, form_code):
    """Test GenericParser on a file. Returns (sheet_count, total_rows, total_headers)."""
    from parsers.generic_parser import GenericParser
    parser = GenericParser()

    wb = load_workbook(filepath, data_only=True)
    results = []
    for sn in wb.sheetnames:
        if "汇总" in sn:
            continue
        ws = wb[sn]
        if (ws.max_row or 0) < 2:
            continue
        parsed = parser._parse_impl(ws, sn)
        if parsed:
            results.append({
                "sheet": sn,
                "equipment_id": parsed.get("equipment_id"),
                "headers": len(parsed.get("headers", [])),
                "rows": len(parsed.get("rows", [])),
                "has_meta": bool(parsed.get("meta")),
                "judgment_col": parsed.get("meta", {}).get("judgment_col"),
            })
    wb.close()
    return results


def test_header_extractor(filepath, form_code):
    """Test header_spec_extractor on a file. Returns (item_count, items_with_spec)."""
    from services.header_spec_extractor import extract_specs_from_headers

    wb = load_workbook(filepath, data_only=True)
    specs = extract_specs_from_headers(wb, form_code, "")
    wb.close()

    if not specs:
        return 0, 0, []

    total_items = 0
    items_with_spec = 0
    item_names = []
    for eq in specs:
        for item in eq.get("items", []):
            total_items += 1
            item_names.append(item["item_name"])
            parsed = item.get("parsed_spec", {})
            if parsed.get("spec_type") != "skip":
                items_with_spec += 1

    return total_items, items_with_spec, item_names


def test_export_annotation(filepath, form_code):
    """Test _annotate_sheet on a file. Returns (judgment_cells_written, total_rows)."""
    from parsers.generic_parser import GenericParser
    from services.export_service import _annotate_sheet

    parser = GenericParser()
    wb_read = load_workbook(filepath, data_only=True)

    # Find first data sheet
    sheet_name = None
    for sn in wb_read.sheetnames:
        if "汇总" not in sn:
            ws = wb_read[sn]
            if (ws.max_row or 0) >= 3:
                sheet_name = sn
                break
    if not sheet_name:
        wb_read.close()
        return 0, 0

    parsed = parser._parse_impl(wb_read[sheet_name], sheet_name)
    wb_read.close()

    if not parsed or not parsed.get("meta") or not parsed.get("rows"):
        return 0, 0

    # Build fake judged data
    rows = parsed["rows"]
    meta = parsed["meta"]
    judged_rows = []
    for i, row in enumerate(rows):
        values = {}
        for key, val in row.get("values", {}).items():
            values[key] = {
                "raw": val,
                "judgment": "NG" if i % 5 == 0 else "OK",
                "spec": {"spec_type": "skip"},
            }
        judged_rows.append({
            "values": values,
            "row_judgment": "NG" if i % 5 == 0 else "OK",
        })

    judged_data = {
        "has_spec": True,
        "judged_rows": judged_rows,
        "meta": meta,
    }

    # Open writable copy and annotate
    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    _annotate_sheet(ws, judged_data)

    # Count judgment cells
    row_map = meta.get("row_map", [])
    ok_count = ng_count = dash_count = 0

    # Find judgment column
    check_col = meta.get("judgment_col")
    if not check_col:
        for c in range(1, (ws.max_column or 1) + 1):
            for r in range(1, 15):
                try:
                    if ws.cell(row=r, column=c).value == "判定":
                        check_col = c
                        break
                except AttributeError:
                    pass
            if check_col:
                break

    if check_col:
        for rm in row_map:
            record_rows = set()
            if rm.get("row"):
                record_rows.add(rm["row"])
            for cell_pos in rm.get("cells", {}).values():
                if isinstance(cell_pos, list) and len(cell_pos) >= 2:
                    record_rows.add(cell_pos[0])
            if not record_rows:
                continue
            check_row = min(record_rows)
            try:
                val = ws.cell(row=check_row, column=check_col).value
                if val == "OK":
                    ok_count += 1
                elif val == "NG":
                    ng_count += 1
                elif val == "—":
                    dash_count += 1
            except AttributeError:
                pass

    wb.close()
    return ok_count + ng_count + dash_count, len(row_map)


def test_builtin_parser(filepath, form_code):
    """Test built-in parser for regression."""
    from parsers.qa1021_parser import QA1021Parser
    from parsers.rd09aa_parser import RD09AAParser
    from parsers.rd09ab_parser import RD09ABParser
    from parsers.rd09aj_parser import RD09AJParser
    from parsers.rd09ak_parser import RD09AKParser

    parsers = {
        "F-QA1021": QA1021Parser(),
        "F-RD09AA": RD09AAParser(),
        "F-RD09AB": RD09ABParser(),
        "F-RD09AJ": RD09AJParser(),
        "F-RD09AK": RD09AKParser(),
    }

    parser = parsers.get(form_code)
    if not parser:
        return []

    wb = load_workbook(filepath, data_only=True)
    results = []
    for sn in wb.sheetnames:
        if "汇总" in sn:
            continue
        ws = wb[sn]
        if (ws.max_row or 0) < 2:
            continue
        try:
            parsed = parser._parse_impl(ws, sn)
            if parsed:
                results.append({
                    "sheet": sn,
                    "equipment_id": parsed.get("equipment_id"),
                    "headers": len(parsed.get("headers", [])),
                    "rows": len(parsed.get("rows", [])),
                    "has_meta": bool(parsed.get("meta")),
                })
        except Exception as e:
            results.append({"sheet": sn, "error": str(e)})
    wb.close()
    return results


def main():
    import logging
    logging.basicConfig(level=logging.WARNING, format="  %(message)s")

    # ─── Part 1: Test original 5 built-in types ───
    print("=" * 80)
    print("  PART 1: Built-in Form Types (Regression Test)")
    print("=" * 80)

    builtin_results = {}
    for form_code, filepath in BUILTIN_FILES.items():
        if not os.path.exists(filepath):
            print(f"  {form_code}: SKIP (file not found)")
            builtin_results[form_code] = "SKIP"
            continue

        sheets = test_builtin_parser(filepath, form_code)
        export_ok, export_total = test_export_annotation(filepath, form_code)

        total_rows = sum(s.get("rows", 0) for s in sheets if "error" not in s)
        errors = [s for s in sheets if "error" in s]

        status = "PASS" if total_rows > 0 and not errors and export_ok > 0 else "FAIL"
        builtin_results[form_code] = status

        print(f"\n  {form_code}: {status}")
        print(f"    Sheets: {len(sheets)}, Total rows: {total_rows}")
        print(f"    Export: {export_ok}/{export_total} judgment cells written")
        if errors:
            for e in errors:
                print(f"    ERROR in {e['sheet']}: {e['error']}")

    # ─── Part 2: Test AU files with GenericParser ───
    print(f"\n{'=' * 80}")
    print("  PART 2: AU Files (GenericParser + Header Extractor + Export)")
    print("=" * 80)

    if not os.path.isdir(AU_DIR):
        print(f"  SKIP: AU directory not found: {AU_DIR}")
        return

    au_files = sorted([f for f in os.listdir(AU_DIR) if f.endswith(".xlsx") and not f.startswith("~")])
    au_results = {}

    for filename in au_files:
        filepath = os.path.join(AU_DIR, filename)
        form_code = extract_form_code(filename)
        if not form_code:
            print(f"\n  {filename}: SKIP (no form code)")
            continue

        # Skip built-in types (already tested above)
        if form_code in BUILTIN_FILES:
            continue

        print(f"\n  {form_code} ({filename})")

        # Stage 1: GenericParser
        try:
            gp_results = test_generic_parser(filepath, form_code)
            gp_sheets = len(gp_results)
            gp_rows = sum(r["rows"] for r in gp_results)
            gp_headers = max((r["headers"] for r in gp_results), default=0)
            gp_judgment = any(r["judgment_col"] for r in gp_results)
            print(f"    GenericParser: {gp_sheets} sheets, {gp_rows} rows, {gp_headers} headers"
                  f"{', has 判定 col' if gp_judgment else ''}")
        except Exception as e:
            gp_rows = 0
            gp_headers = 0
            print(f"    GenericParser: ERROR - {e}")

        # Stage 2: Header Extractor
        try:
            he_items, he_with_spec, he_names = test_header_extractor(filepath, form_code)
            print(f"    HeaderExtractor: {he_items} items ({he_with_spec} with spec)")
            if he_names:
                preview = ", ".join(he_names[:5])
                if len(he_names) > 5:
                    preview += f"... +{len(he_names)-5} more"
                print(f"      Items: {preview}")
        except Exception as e:
            he_items = 0
            print(f"    HeaderExtractor: ERROR - {e}")

        # Stage 3: Export annotation
        try:
            ex_cells, ex_total = test_export_annotation(filepath, form_code)
            print(f"    Export: {ex_cells}/{ex_total} judgment cells")
        except Exception as e:
            ex_cells = 0
            ex_total = 0
            print(f"    Export: ERROR - {e}")

        # Verdict
        if gp_rows == 0 and he_items == 0:
            status = "EMPTY"
        elif gp_rows > 0 and ex_cells > 0:
            status = "FULL"  # All 3 stages work
        elif gp_rows > 0:
            status = "PARSE_ONLY"  # Parse works, export may have issues
        elif he_items > 0:
            status = "SPEC_ONLY"  # Only spec extraction works
        else:
            status = "FAIL"

        au_results[form_code] = status
        print(f"    Status: {status}")

    # ─── Summary ───
    print(f"\n{'=' * 80}")
    print("  SUMMARY")
    print("=" * 80)

    print("\n  Built-in types:")
    for code, status in builtin_results.items():
        marker = "✓" if status == "PASS" else "✗"
        print(f"    {marker} {code}: {status}")

    print("\n  AU types:")
    status_counts = {"FULL": 0, "PARSE_ONLY": 0, "SPEC_ONLY": 0, "EMPTY": 0, "FAIL": 0}
    for code, status in sorted(au_results.items()):
        marker = "✓" if status == "FULL" else "△" if status in ("PARSE_ONLY", "SPEC_ONLY") else "✗"
        print(f"    {marker} {code}: {status}")
        status_counts[status] = status_counts.get(status, 0) + 1

    print(f"\n  Totals: {status_counts}")
    builtin_ok = sum(1 for s in builtin_results.values() if s == "PASS")
    print(f"  Built-in: {builtin_ok}/{len(builtin_results)} PASS")
    print(f"  AU FULL: {status_counts.get('FULL', 0)}/{len(au_results)}")


if __name__ == "__main__":
    main()
