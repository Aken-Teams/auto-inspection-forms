"""AI-powered generic parser for 汇总 (summary) sheets.

Used for non-built-in form types where no hardcoded parser exists.
Leverages DeepSeek AI to understand the summary sheet structure and
extract equipment IDs and spec items.
"""
import json
import re
import logging
from openpyxl.worksheet.worksheet import Worksheet
from services.ai_service import _call_deepseek, _extract_json, is_ai_available
from utils.spec_parser import parse_spec_string

logger = logging.getLogger(__name__)


def extract_summary_content(ws: Worksheet, max_rows: int = 50, max_cols: int = 30) -> str:
    """Extract text content from a summary sheet for AI analysis.

    Returns a formatted string preserving row/column structure.
    """
    effective_rows = min(max_rows, ws.max_row or 1)
    effective_cols = min(max_cols, ws.max_column or 1)

    lines = []
    for row in range(1, effective_rows + 1):
        cells = []
        for col in range(1, effective_cols + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                cells.append(str(val).strip()[:80])
            else:
                cells.append("")
        line = " | ".join(cells)
        # Skip fully empty rows
        if line.replace(" | ", "").replace(" ", ""):
            lines.append(f"Row {row}: {line}")
    return "\n".join(lines)


def validate_summary_structure(ws: Worksheet) -> dict:
    """Validate basic structure of a summary sheet before parsing.

    Returns:
        {"valid": bool, "has_data": bool, "row_count": int, "col_count": int,
         "warnings": [...], "detected_headers": [...]}
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    warnings = []

    if max_row < 2:
        warnings.append("Summary sheet has fewer than 2 rows")
    if max_col < 2:
        warnings.append("Summary sheet has fewer than 2 columns")

    # Check if sheet has actual data
    has_data = False
    data_cells = 0
    for row in range(1, min(10, max_row + 1)):
        for col in range(1, min(20, max_col + 1)):
            if ws.cell(row=row, column=col).value is not None:
                data_cells += 1
                has_data = True

    if data_cells < 5:
        warnings.append("Summary sheet appears to have very little data")

    # Detect header row (first row with 3+ non-empty cells)
    detected_headers = []
    for row in range(1, min(15, max_row + 1)):
        non_empty = 0
        headers = []
        for col in range(1, min(30, max_col + 1)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                non_empty += 1
                headers.append(str(val).strip()[:50])
        if non_empty >= 3:
            detected_headers = headers
            break

    return {
        "valid": has_data and max_row >= 2 and max_col >= 2 and len(warnings) == 0,
        "has_data": has_data,
        "row_count": max_row,
        "col_count": max_col,
        "warnings": warnings,
        "detected_headers": detected_headers,
    }


def ai_parse_summary_sheet(ws: Worksheet, form_code: str, form_name: str) -> dict | None:
    """Use DeepSeek AI to parse a summary sheet and extract spec data.

    Returns:
        {
            "equipment_specs": [
                {
                    "equipment_id": "WPRN-0001",
                    "equipment_name": "Machine WPRN-0001",
                    "items": [
                        {
                            "item_name": "Temperature",
                            "spec_value": "125~145",
                            "group_name": "Parameters",
                            "sub_group": null,
                            "display_order": 0
                        }
                    ]
                }
            ],
            "confidence": 0.85,
            "analysis_notes": "..."
        }
    """
    if not is_ai_available():
        return None

    content = extract_summary_content(ws)
    if not content.strip():
        return None

    prompt = _build_parse_prompt(content, form_code, form_name)

    result = _call_deepseek(
        [{"role": "user", "content": prompt}],
        temperature=0.1,
        max_tokens=8000,
    )

    if not result:
        logger.warning(f"AI returned empty response for {form_code}")
        return None

    parsed = _extract_json(result)
    if not parsed or "equipment_specs" not in parsed:
        logger.warning(f"Failed to extract structured data from AI response for {form_code}")
        return None

    # Post-process: run each spec_value through parse_spec_string
    for eq in parsed.get("equipment_specs", []):
        for item in eq.get("items", []):
            spec_str = str(item.get("spec_value", ""))
            spec_parsed = parse_spec_string(spec_str)
            item["parsed_spec"] = spec_parsed

    logger.info(
        f"AI parsed {form_code}: {len(parsed['equipment_specs'])} equipment(s), "
        f"confidence={parsed.get('confidence', 'N/A')}"
    )
    return parsed


def ai_parse_data_sheet(ws: Worksheet, form_code: str, form_name: str) -> dict | None:
    """Use DeepSeek AI to parse a DATA sheet (not summary) and extract spec items.

    This is the last-resort fallback when:
    1. No 汇总 sheet exists
    2. Rule-based header extraction also failed

    Returns same format as ai_parse_summary_sheet.
    """
    if not is_ai_available():
        return None

    content = extract_summary_content(ws)
    if not content.strip():
        return None

    prompt = _build_data_sheet_prompt(content, form_code, form_name)

    result = _call_deepseek(
        [{"role": "user", "content": prompt}],
        temperature=0.1,
        max_tokens=8000,
    )

    if not result:
        logger.warning(f"AI returned empty response for data sheet of {form_code}")
        return None

    parsed = _extract_json(result)
    if not parsed or "equipment_specs" not in parsed:
        logger.warning(f"Failed to extract structured data from AI for data sheet of {form_code}")
        return None

    # Post-process: run each spec_value through parse_spec_string
    for eq in parsed.get("equipment_specs", []):
        for item in eq.get("items", []):
            spec_str = str(item.get("spec_value", ""))
            spec_parsed = parse_spec_string(spec_str)
            item["parsed_spec"] = spec_parsed

    logger.info(
        f"AI parsed data sheet for {form_code}: "
        f"{len(parsed['equipment_specs'])} equipment(s), "
        f"confidence={parsed.get('confidence', 'N/A')}"
    )
    return parsed


def _build_data_sheet_prompt(sheet_content: str, form_code: str, form_name: str) -> str:
    """Build AI prompt for parsing a raw data sheet (no summary sheet available)."""
    truncated = sheet_content[:6000]

    return f"""你是一个工业检验表格分析专家。以下是一张检验记录的**原始数据工作表**（不是汇总表），请从中提取检验项目定义。

表单信息:
- 表单编号: {form_code}
- 表单名称: {form_name}

工作表内容（每行格式为 Row N: cell1 | cell2 | ...）:
{truncated}

这是一张**原始数据表**，检验项目可能出现在:
1. **列标题**（表头行）: 检验项目名称排列在某一行中，每列一个项目，下方是逐次检验数据
2. **行标签**（纵向排列）: 检验项目名称排列在某一列中，每行一个项目，右方是逐次检验数据
3. **日历/网格格式**: 表头区域包含「XXX规格：value」形式的规格定义
4. **混合格式**: 多层表头、分区编号等复杂结构

请从表头/标签中识别出**检验项目名称**和**规格值**（如果有的话）。

需要**跳过**以下类型的栏位（它们是元数据，不是检验项目）:
- 日期、时间、班别、班次
- 签名、确认、判定、备注
- 序号、工单号、批号、料号
- 机台号、设备编号、柜号

请以JSON格式返回（不要添加其他说明文字）:
{{
  "equipment_specs": [
    {{
      "equipment_id": "UNIVERSAL",
      "equipment_name": "{form_name or form_code} 通用规格",
      "items": [
        {{
          "item_name": "检验项目名称",
          "spec_value": "规格值（如 125~145, ≥3, √ 等，若无规格则留空字符串）",
          "group_name": "分组名（如有，否则null）",
          "sub_group": null,
          "display_order": 0
        }}
      ]
    }}
  ],
  "confidence": 0.7,
  "analysis_notes": "简要说明你识别到的表格结构和检验项目"
}}

注意:
1. spec_value 保持原始格式，不要转换
2. 如果表头中嵌入了规格范围（如「顶针高度（300-750）um」），提取为 spec_value="300~750"
3. 如果表头中嵌入了阈值（如「缺胶≤0.07%」），提取为 spec_value="≤0.07%"
4. 如果看到「XXX规格：value」形式（如「温度规格：20-26℃」），提取项目名=XXX，spec_value=value
5. 没有明确规格的项目，spec_value 设为空���符串
6. confidence 填写你对结果的信心度（0-1）"""


def _build_parse_prompt(sheet_content: str, form_code: str, form_name: str) -> str:
    """Build the AI prompt for summary sheet parsing."""
    # Limit content to prevent token overflow
    truncated = sheet_content[:6000]

    return f"""你是一个工业检验表格分析专家。请分析以下Excel「汇总」工作表的内容，提取所有设备的检验规格数据。

表单信息:
- 表单编号: {form_code}
- 表单名称: {form_name}

汇总工作表内容（每行格式为 Row N: cell1 | cell2 | ...）:
{truncated}

请分析汇总表的结构，识别:
1. 每个设备/机台的编号（equipment_id）和名称
2. 每个设备的检验项目名称（item_name）
3. 每个检验项目的规格值（spec_value）- 可能是范围如 "125~145"，阈值如 "≥3"，打勾如 "√"，文字如 "OK" 等
4. 检验项目的分组（group_name）和子分组（sub_group）（如果有的话）

汇总表常见的结构模式:
- 模式A: 每个设备占一个区块，设备编号在区块标题行，下方是该设备的规格项目
- 模式B: 表格形式，每行一个设备，每列一个检验项目
- 模式C: 混合模式，有分组标题行，下面是该组的项目

请以JSON格式返回结果（不要添加其他说明文字）:
{{
  "equipment_specs": [
    {{
      "equipment_id": "设备编号（从内容提取）",
      "equipment_name": "设备名称或编号",
      "items": [
        {{
          "item_name": "检验项目名称",
          "spec_value": "原始规格值字符串（如 125~145, ≥3, √, OK 等）",
          "group_name": "分组名（如有，否则null）",
          "sub_group": "子分组（如有，否则null）",
          "display_order": 0
        }}
      ]
    }}
  ],
  "confidence": 0.85,
  "analysis_notes": "简要说明你识别到的汇总表结构"
}}

注意:
1. spec_value 保持Excel中的原始格式，不要转换（系统会自动解析）
2. 如果同一检验项目在不同部位有不同规格（如上模/下模），用 sub_group 区分
3. equipment_id 通常是机台编号如 WPRN-0001, WCBA-0001, RD-LZ-14, WTFB-0004 等
4. 如果汇总表是通用规格（所有设备共用一套），equipment_id 使用 "UNIVERSAL"
5. 跳过空行、注释行和无效数据
6. confidence 填写你对解析结果的信心度（0-1），0.5以下表示不太确定"""
