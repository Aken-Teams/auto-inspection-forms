"""
Core flow integration test:
  1. Test 5 built-in types (upload → spec match → OK/NG judgment)
  2. Check which AU files have 汇总 sheets
  3. For files with 汇总: import specs → upload → verify OK/NG
"""
import sys
import os
import json
import shutil
import tempfile

sys.stdout.reconfigure(encoding='utf-8')

import httpx

BASE = "http://localhost:8000"
AU_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "AU未建立规格点检表", "AU")

client = httpx.Client(base_url=BASE, timeout=30)


def upload_file(filepath):
    """Upload a file and return the response."""
    fname = os.path.basename(filepath)
    # Copy to temp to avoid encoding issues
    tmp = os.path.join(tempfile.gettempdir(), fname)
    shutil.copy2(filepath, tmp)
    with open(tmp, "rb") as f:
        resp = client.post("/api/upload", files={"file": (fname, f, "application/octet.stream")})
    os.remove(tmp)
    return resp


def get_upload_detail(upload_id):
    """Get detailed results for an upload."""
    resp = client.get(f"/api/results/{upload_id}")
    return resp.json()


def check_sheets_for_summary(filepath):
    """Check if a file has a 汇总 sheet."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        has_summary = "汇总" in wb.sheetnames
        sheet_names = wb.sheetnames
        data_sheets = [s for s in sheet_names if s != "汇总"]
        wb.close()
        return has_summary, data_sheets
    except Exception as e:
        return False, []


def import_specs(form_code, filepath):
    """Import specs from a file's 汇总 sheet."""
    fname = os.path.basename(filepath)
    tmp = os.path.join(tempfile.gettempdir(), fname)
    shutil.copy2(filepath, tmp)
    with open(tmp, "rb") as f:
        resp = client.post(
            f"/api/specs/import?form_code={form_code}",
            files={"file": (fname, f, "application/octet.stream")},
        )
    os.remove(tmp)
    return resp


def print_separator():
    print("=" * 80)


def test_builtin_types():
    """Test 1: Upload 5 built-in type files and verify spec matching + judgment."""
    print_separator()
    print("TEST 1: Testing 5 built-in form types")
    print_separator()

    builtin_files = {
        "F-QA1021": "F-QA1021_离子消散设备点检记录表.xlsx",
        "F-RD09AA": "F-RD09AA-Auto Mold 机台检查记录表.xlsx",
        "F-RD09AB": "F-RD09AB-Auto Mold 洗模检查记录表.xlsx",
        "F-RD09AJ": "F-RD09AJ-RO 焊接炉检查记录表.xlsx",
        "F-RD09AK": "F-RD09AK_SMD(Clip）切弯脚尺寸检查记录表.xlsx",
    }

    results = {}
    for form_code, filename in builtin_files.items():
        filepath = os.path.join(AU_DIR, filename)
        if not os.path.exists(filepath):
            print(f"  [SKIP] {form_code}: File not found: {filename}")
            continue

        print(f"\n  Testing {form_code} ({filename})...")

        # Check current specs (include_items=false for count only)
        specs_resp = client.get(f"/api/specs/form-types/{form_code}/specs", params={"include_items": "false"})
        specs = specs_resp.json()
        print(f"    Specs in DB: {len(specs)} groups")
        for s in specs[:3]:
            print(f"      - {s['equipment_id']}: {s.get('item_count', '?')} items")

        # Upload file
        resp = upload_file(filepath)
        if resp.status_code != 200:
            print(f"    [FAIL] Upload failed: {resp.status_code} {resp.text[:200]}")
            results[form_code] = "UPLOAD_FAIL"
            continue

        data = resp.json()
        upload_id = data.get("id") or data.get("upload_id")
        if not upload_id and isinstance(data, list):
            upload_id = data[0].get("id")
        if not upload_id and "results" in data:
            upload_id = data["results"][0]["id"] if data["results"] else None

        print(f"    Upload response keys: {list(data.keys()) if isinstance(data, dict) else 'list'}")
        print(f"    Upload ID: {upload_id}")

        if not upload_id:
            print(f"    [FAIL] No upload_id in response")
            print(f"    Response: {json.dumps(data, ensure_ascii=False)[:500]}")
            results[form_code] = "NO_UPLOAD_ID"
            continue

        # Get details
        detail = get_upload_detail(upload_id)
        print(f"    Form matched: {detail.get('form_code', 'N/A')}")
        print(f"    Total sheets: {len(detail.get('sheets', []))}")

        ok_count = 0
        ng_count = 0
        no_spec = 0
        for sheet in detail.get("sheets", []):
            result = sheet["overall_result"]
            if result == "OK":
                ok_count += 1
            elif result == "NG":
                ng_count += 1
            elif result == "NO_SPEC":
                no_spec += 1

        print(f"    Results: OK={ok_count}, NG={ng_count}, NO_SPEC={no_spec}")

        matched_code = detail.get("form_code")
        if matched_code != form_code:
            print(f"    [FAIL] Wrong form type! Expected {form_code}, got {matched_code}")
            results[form_code] = "WRONG_TYPE"
        elif no_spec > 0 and ok_count == 0 and ng_count == 0:
            print(f"    [WARN] All sheets NO_SPEC - specs may not be configured")
            results[form_code] = "NO_SPEC"
        elif ok_count + ng_count > 0:
            print(f"    [OK] Judgment working! {ok_count} OK, {ng_count} NG")
            results[form_code] = "PASS"
        else:
            print(f"    [WARN] No sheets processed")
            results[form_code] = "NO_SHEETS"

        # Show sample judgment details for first OK or NG sheet
        for sheet in detail.get("sheets", []):
            if sheet["overall_result"] in ("OK", "NG"):
                jd = sheet.get("judged_data")
                if jd and isinstance(jd, list):
                    print(f"    Sample judgment ({sheet['sheet_name']}):")
                    for item in jd[:3]:
                        if isinstance(item, dict):
                            name = item.get("item_name", item.get("name", "?"))
                            result = item.get("result", "?")
                            actual = item.get("actual_value", item.get("value", "?"))
                            print(f"      {name}: {actual} -> {result}")
                break

    print(f"\n  Summary:")
    for fc, r in results.items():
        status = "[OK]" if r == "PASS" else f"[{r}]"
        print(f"    {fc}: {status}")

    return results


def test_au_files_summary_check():
    """Test 2: Check which AU files have 汇总 sheets."""
    print_separator()
    print("TEST 2: Checking which AU files have 汇总 sheets")
    print_separator()

    with_summary = []
    without_summary = []

    files = sorted(os.listdir(AU_DIR))
    for fname in files:
        filepath = os.path.join(AU_DIR, fname)
        has_summary, data_sheets = check_sheets_for_summary(filepath)
        if has_summary:
            with_summary.append((fname, data_sheets))
            print(f"  [HAS] {fname} ({len(data_sheets)} data sheets)")
        else:
            without_summary.append((fname, data_sheets))
            print(f"  [NO]  {fname} ({len(data_sheets)} sheets)")

    print(f"\n  With 汇总: {len(with_summary)}")
    print(f"  Without 汇总: {len(without_summary)}")

    return with_summary, without_summary


def test_non_builtin_with_summary(with_summary_files):
    """Test 3: For non-built-in files with 汇总, import specs and test judgment."""
    print_separator()
    print("TEST 3: Testing non-built-in files WITH 汇总 sheets")
    print_separator()

    builtin_codes = {"F-QA1021", "F-RD09AA", "F-RD09AB", "F-RD09AJ", "F-RD09AK"}

    # Extract form code from filename
    import re
    form_code_re = re.compile(r"(F-[A-Z]{2}\d{2,4}[A-Z0-9]{0,3})(?=[_\-\s.\u4e00-\u9fff]|$)", re.IGNORECASE)

    results = {}
    for fname, data_sheets in with_summary_files:
        match = form_code_re.search(fname)
        if not match:
            print(f"\n  [SKIP] Cannot extract form code from: {fname}")
            continue

        form_code = match.group(1).upper()
        if form_code in builtin_codes:
            continue  # Already tested

        filepath = os.path.join(AU_DIR, fname)
        print(f"\n  Testing {form_code} ({fname})...")
        print(f"    Data sheets: {data_sheets[:5]}{'...' if len(data_sheets) > 5 else ''}")

        # Ensure form type exists
        ft_resp = client.get("/api/specs/form-types")
        existing_codes = [ft["form_code"] for ft in ft_resp.json()]
        if form_code not in existing_codes:
            print(f"    Creating form type {form_code}...")
            create_resp = client.post("/api/specs/form-types", json={
                "form_code": form_code,
                "form_name": fname.replace(".xlsx", ""),
                "file_pattern": re.escape(form_code),
            })
            if create_resp.status_code not in (200, 201):
                # Maybe already exists
                print(f"    Create response: {create_resp.status_code}")

        # Import specs from 汇总
        print(f"    Importing specs from 汇总...")
        import_resp = import_specs(form_code, filepath)
        if import_resp.status_code == 200:
            import_data = import_resp.json()
            print(f"    Import result: {json.dumps(import_data, ensure_ascii=False)[:200]}")
        else:
            print(f"    [WARN] Import status: {import_resp.status_code} {import_resp.text[:200]}")

        # Check specs were created
        specs_resp = client.get(f"/api/specs/form-types/{form_code}/specs", params={"include_items": "false"})
        specs = specs_resp.json()
        total_items = sum(s.get("item_count", 0) for s in specs)
        print(f"    Specs after import: {len(specs)} groups, {total_items} total items")

        if total_items == 0:
            print(f"    [SKIP] No spec items imported (汇总 parser may not support this format)")
            results[form_code] = "NO_ITEMS_IMPORTED"
            continue

        # Upload the file and check judgment
        resp = upload_file(filepath)
        if resp.status_code != 200:
            print(f"    [FAIL] Upload failed: {resp.status_code}")
            results[form_code] = "UPLOAD_FAIL"
            continue

        data = resp.json()
        upload_id = data.get("id") or data.get("upload_id")
        if not upload_id:
            print(f"    Response: {json.dumps(data, ensure_ascii=False)[:300]}")
            results[form_code] = "NO_UPLOAD_ID"
            continue

        detail = get_upload_detail(upload_id)
        matched = detail.get("form_code")
        sheets = detail.get("sheets", [])

        ok = sum(1 for s in sheets if s["overall_result"] == "OK")
        ng = sum(1 for s in sheets if s["overall_result"] == "NG")
        ns = sum(1 for s in sheets if s["overall_result"] == "NO_SPEC")

        print(f"    Matched form: {matched}")
        print(f"    Results: OK={ok}, NG={ng}, NO_SPEC={ns}, Total={len(sheets)}")

        if matched != form_code:
            print(f"    [FAIL] Wrong form type!")
            results[form_code] = "WRONG_TYPE"
        elif ok + ng > 0:
            print(f"    [OK] Judgment working!")
            results[form_code] = "PASS"
            # Show a sample
            for s in sheets:
                if s["overall_result"] in ("OK", "NG") and s.get("judged_data"):
                    jd = s["judged_data"]
                    if isinstance(jd, list) and len(jd) > 0:
                        print(f"    Sample ({s['sheet_name']} = {s['overall_result']}):")
                        for item in jd[:2]:
                            if isinstance(item, dict):
                                print(f"      {item.get('item_name','?')}: actual={item.get('actual_value','?')} -> {item.get('result','?')}")
                    break
        elif ns == len(sheets):
            print(f"    [WARN] All NO_SPEC despite importing")
            results[form_code] = "ALL_NO_SPEC"
        else:
            results[form_code] = "UNKNOWN"

    print(f"\n  Summary:")
    for fc, r in results.items():
        status = "[OK]" if r == "PASS" else f"[{r}]"
        print(f"    {fc}: {status}")

    return results


def test_identification_only(without_summary_files):
    """Test 4: For files without 汇总, just check identification works."""
    print_separator()
    print("TEST 4: Testing file identification for files WITHOUT 汇总")
    print_separator()

    import re
    form_code_re = re.compile(r"(F-[A-Z]{2}\d{2,4}[A-Z0-9]{0,3})(?=[_\-\s.\u4e00-\u9fff]|$)", re.IGNORECASE)
    builtin_codes = {"F-QA1021", "F-RD09AA", "F-RD09AB", "F-RD09AJ", "F-RD09AK"}

    results = {}
    for fname, data_sheets in without_summary_files:
        match = form_code_re.search(fname)
        if not match:
            continue
        form_code = match.group(1).upper()
        if form_code in builtin_codes:
            continue

        filepath = os.path.join(AU_DIR, fname)
        resp = upload_file(filepath)
        if resp.status_code != 200:
            print(f"  [FAIL] {form_code} ({fname}): Upload failed {resp.status_code}")
            results[form_code] = "UPLOAD_FAIL"
            continue

        data = resp.json()
        upload_id = data.get("id") or data.get("upload_id")
        if not upload_id:
            results[form_code] = "NO_ID"
            continue

        detail = get_upload_detail(upload_id)
        matched = detail.get("form_code")
        sheets = detail.get("sheets", [])
        ns = sum(1 for s in sheets if s["overall_result"] == "NO_SPEC")

        if matched == form_code:
            print(f"  [OK] {form_code}: Identified correctly, {len(sheets)} sheets ({ns} NO_SPEC)")
            results[form_code] = "IDENTIFIED"
        else:
            print(f"  [FAIL] {form_code}: Expected {form_code}, got {matched}")
            results[form_code] = "WRONG_TYPE"

    return results


if __name__ == "__main__":
    print("Core Flow Integration Test")
    print(f"AU directory: {AU_DIR}")
    print(f"Files found: {len(os.listdir(AU_DIR))}")

    # Test 1: Built-in types
    builtin_results = test_builtin_types()

    # Test 2: Check for 汇总
    with_summary, without_summary = test_au_files_summary_check()

    # Test 3: Non-built-in with 汇总
    summary_results = test_non_builtin_with_summary(with_summary)

    # Test 4: Identification for files without 汇总
    ident_results = test_identification_only(without_summary)

    # Final summary
    print_separator()
    print("FINAL SUMMARY")
    print_separator()

    all_pass = True
    for label, results in [("Built-in", builtin_results), ("With 汇总", summary_results), ("ID only", ident_results)]:
        print(f"\n  {label}:")
        for fc, r in results.items():
            ok = r in ("PASS", "IDENTIFIED", "NO_ITEMS_IMPORTED")
            if not ok:
                all_pass = False
            print(f"    {fc}: {r}")

    print(f"\n  Overall: {'ALL PASS' if all_pass else 'SOME FAILURES - NEEDS FIX'}")
