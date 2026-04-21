"""Diagnostic script to analyze Excel column structure for judgment column insertion."""
import os
import sys
sys.path.insert(0, os.path.dirname(__file__))

from openpyxl import load_workbook

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")

FILES = {
    "QA1021": os.path.join(DATA_DIR, "F-QA1021_离子消散设备点检记录表.xlsx"),
    "RD09AA": os.path.join(DATA_DIR, "F-RD09AA-Auto Mold 机台检查记录表.xlsx"),
    "RD09AB": os.path.join(DATA_DIR, "F-RD09AB-Auto Mold 洗模检查记录表.xlsx"),
    "RD09AJ": os.path.join(DATA_DIR, "F-RD09AJ-RO 焊接炉检查记录表.xlsx"),
    "RD09AK": os.path.join(DATA_DIR, "F-RD09AK_SMD(Clip）切弯脚尺寸检查记录表.xlsx"),
}


def analyze_sheet(form_code, wb, sheet_name):
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"  {form_code} / {sheet_name}")
    print(f"  ws.max_column={ws.max_column}, ws.max_row={ws.max_row}")
    print(f"{'='*80}")

    # 1. Show merged cell ranges (sorted by min_col desc, show last 15)
    merged = sorted(ws.merged_cells.ranges, key=lambda m: m.min_col)
    print(f"\n  Merged ranges ({len(merged)} total):")
    for mr in merged:
        print(f"    rows {mr.min_row}-{mr.max_row}, cols {mr.min_col}-{mr.max_col} "
              f"({_col_letter(mr.min_col)}-{_col_letter(mr.max_col)})")

    # 2. Show header rows content (rows 1-8, all columns with content)
    print(f"\n  Header area (rows 1-8):")
    for row in range(1, 9):
        cells_with_content = []
        for col in range(1, (ws.max_column or 1) + 1):
            try:
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    cells_with_content.append(f"col{col}({_col_letter(col)})={repr(val)[:30]}")
            except AttributeError:
                cells_with_content.append(f"col{col}({_col_letter(col)})=<MERGED>")
        if cells_with_content:
            print(f"    Row {row}: {', '.join(cells_with_content)}")

    # 3. Show first 2 data rows content (rows 9-12)
    print(f"\n  Data area (rows 9-12):")
    for row in range(9, 13):
        cells_with_content = []
        for col in range(1, (ws.max_column or 1) + 1):
            try:
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    cells_with_content.append(f"col{col}({_col_letter(col)})={repr(val)[:25]}")
            except AttributeError:
                cells_with_content.append(f"col{col}({_col_letter(col)})=<MERGED>")
        if cells_with_content:
            print(f"    Row {row}: {', '.join(cells_with_content)}")

    # 4. Run parser to get row_map and compute max_data_col
    print(f"\n  Parser analysis:")
    try:
        parsed = _run_parser(form_code, ws, sheet_name)
        if parsed and parsed.get("meta"):
            row_map = parsed["meta"].get("row_map", [])
            judgment_col = parsed["meta"].get("judgment_col")
            print(f"    judgment_col from parser: {judgment_col}")
            print(f"    row_map entries: {len(row_map)}")

            # Compute max_data_col
            max_data_col = 0
            all_physical_rows = set()
            for rm in row_map:
                if rm.get("row"):
                    all_physical_rows.add(rm["row"])
                for cell_pos in rm.get("cells", {}).values():
                    if isinstance(cell_pos, list) and len(cell_pos) >= 2:
                        max_data_col = max(max_data_col, cell_pos[1])
                        all_physical_rows.add(cell_pos[0])

            insert_col = max_data_col + 1
            first_data_row = min(all_physical_rows) if all_physical_rows else 1
            last_data_row = max(all_physical_rows) if all_physical_rows else 1

            print(f"    max_data_col: {max_data_col} ({_col_letter(max_data_col)})")
            print(f"    insert_col: {insert_col} ({_col_letter(insert_col)})")
            print(f"    physical rows: {first_data_row}-{last_data_row} ({len(all_physical_rows)} rows)")
            if row_map:
                print(f"    first row_map entry row: {row_map[0].get('row')}")
                print(f"    first row_map cells: {list(row_map[0].get('cells', {}).keys())}")

            # Check what's at insert_col
            print(f"\n  Content at insert_col ({insert_col}/{_col_letter(insert_col)}):")
            for r in range(max(1, first_data_row - 5), min(first_data_row + 5, ws.max_row + 1)):
                try:
                    val = ws.cell(row=r, column=insert_col).value
                    cell_obj = ws.cell(row=r, column=insert_col)
                    cell_type = type(cell_obj).__name__
                    if val is not None or cell_type != "Cell":
                        print(f"    Row {r}: value={repr(val)}, type={cell_type}")
                except AttributeError:
                    print(f"    Row {r}: <MergedCell>")

            # Check which merged ranges overlap with insert_col
            print(f"\n  Merged ranges overlapping insert_col ({insert_col}):")
            for mr in merged:
                if mr.min_col <= insert_col <= mr.max_col:
                    print(f"    rows {mr.min_row}-{mr.max_row}, cols {mr.min_col}-{mr.max_col} "
                          f"(min_col {'<' if mr.min_col < insert_col else '>='} insert_col) "
                          f"{'*** SPANS ACROSS - would be SKIPPED by current code ***' if mr.min_col < insert_col else ''}")
        else:
            print(f"    No meta in parsed result")
    except Exception as e:
        print(f"    Parser error: {e}")


def _run_parser(form_code, ws, sheet_name):
    """Run the appropriate parser."""
    if form_code == "QA1021":
        from parsers.qa1021_parser import QA1021Parser
        return QA1021Parser()._parse_impl(ws, sheet_name)
    elif form_code == "RD09AA":
        from parsers.rd09aa_parser import RD09AAParser
        return RD09AAParser()._parse_impl(ws, sheet_name)
    elif form_code == "RD09AB":
        from parsers.rd09ab_parser import RD09ABParser
        return RD09ABParser()._parse_impl(ws, sheet_name)
    elif form_code == "RD09AJ":
        from parsers.rd09aj_parser import RD09AJParser
        return RD09AJParser()._parse_impl(ws, sheet_name)
    elif form_code == "RD09AK":
        from parsers.rd09ak_parser import RD09AKParser
        return RD09AKParser()._parse_impl(ws, sheet_name)
    return None


def _col_letter(col):
    """Convert column number to letter."""
    result = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result


if __name__ == "__main__":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    for form_code, filepath in FILES.items():
        if not os.path.exists(filepath):
            print(f"\n  SKIP {form_code}: file not found at {filepath}")
            continue
        wb = load_workbook(filepath, data_only=True)
        # Analyze first non-summary sheet
        for sn in wb.sheetnames:
            if "汇总" in sn or "summary" in sn.lower():
                continue
            analyze_sheet(form_code, wb, sn)
            break  # Just first data sheet
        wb.close()
