"""Service for generating and comparing structural fingerprints of Excel summary sheets."""
import hashlib
import json
import logging

from sqlalchemy.orm import Session
from openpyxl.worksheet.worksheet import Worksheet
from models import FormType

logger = logging.getLogger(__name__)


def generate_fingerprint(ws: Worksheet, max_rows: int = 20) -> dict:
    """Generate a structural fingerprint from a summary sheet.

    Captures column headers, row labels, and layout pattern to identify
    whether two files share the same format.
    """
    max_col = min(ws.max_column or 1, 30)
    max_row_limit = min(ws.max_row or 1, max_rows)

    # Find the first dense row (likely header)
    headers = []
    header_row = None
    for row in range(1, max_row_limit + 1):
        non_empty = 0
        row_vals = []
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                non_empty += 1
                row_vals.append(str(val).strip()[:50])
            else:
                row_vals.append("")
        if non_empty >= 3 and not header_row:
            header_row = row
            headers = [v for v in row_vals if v]

    # Extract first-column labels (row labels)
    row_labels = []
    for row in range(1, max_row_limit + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            label = str(val).strip()[:50]
            if label:
                row_labels.append(label)

    # Build a layout pattern: which cells have data in the first N rows
    layout_cells = []
    for row in range(1, max_row_limit + 1):
        for col in range(1, max_col + 1):
            if ws.cell(row=row, column=col).value is not None:
                layout_cells.append(f"{row},{col}")

    layout_str = "|".join(layout_cells)
    layout_hash = hashlib.sha256(layout_str.encode()).hexdigest()[:16]

    return {
        "headers": headers[:20],
        "row_labels": row_labels[:20],
        "col_count": max_col,
        "row_count": ws.max_row or 0,
        "layout_hash": layout_hash,
    }


def compare_fingerprints(fp1: dict, fp2: dict) -> dict:
    """Compare two fingerprints and return similarity analysis.

    Returns:
        {"match": bool, "similarity": float, "header_match": float, "warnings": [...]}
    """
    if not fp1 or not fp2:
        return {"match": False, "similarity": 0.0, "header_match": 0.0, "warnings": ["Missing fingerprint data"]}

    warnings = []

    # Header similarity (Jaccard)
    h1 = set(fp1.get("headers", []))
    h2 = set(fp2.get("headers", []))
    if h1 and h2:
        header_match = len(h1 & h2) / len(h1 | h2)
    elif not h1 and not h2:
        header_match = 1.0
    else:
        header_match = 0.0

    if header_match < 0.5:
        warnings.append("Column headers significantly different")

    # Layout hash match
    layout_match = 1.0 if fp1.get("layout_hash") == fp2.get("layout_hash") else 0.3

    # Column count similarity
    c1 = fp1.get("col_count", 0)
    c2 = fp2.get("col_count", 0)
    if c1 and c2:
        col_match = min(c1, c2) / max(c1, c2)
    else:
        col_match = 0.0

    if abs(c1 - c2) > 5:
        warnings.append(f"Column count differs: {c1} vs {c2}")

    # Weighted overall similarity
    similarity = header_match * 0.5 + layout_match * 0.3 + col_match * 0.2

    return {
        "match": similarity >= 0.6,
        "similarity": round(similarity, 2),
        "header_match": round(header_match, 2),
        "warnings": warnings,
    }


def save_fingerprint(db: Session, form_code: str, fingerprint: dict):
    """Save a structural fingerprint to the FormType record."""
    ft = db.query(FormType).filter(FormType.form_code == form_code).first()
    if ft:
        ft.structural_fingerprint = fingerprint
        db.flush()
        logger.info(f"Saved fingerprint for {form_code}")
