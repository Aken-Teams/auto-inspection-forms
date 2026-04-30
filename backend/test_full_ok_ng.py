"""
全面 OK/NG 判定測試

覆蓋：
  A. judge_value() 邊界值（所有規格類型）
  B. _find_spec_for_key() 全 5 個表單 key 對應
  C. _judge_rd09ab() 3 層備援 × NG/OK 各場景
  D. 全 pipeline：解析真實檔案 → 注入 NG 值 → 判定 → 驗證
  E. 匯出標注：NG 儲存格紅色、判定欄寫入 OK/NG/—
  F. GenericParser col_N 模糊比對
"""
import sys, os, copy
sys.path.insert(0, os.path.dirname(__file__))
sys.stdout.reconfigure(encoding="utf-8")

from decimal import Decimal

PASS_CNT = 0
FAIL_CNT = 0
SECTION = ""


def section(title):
    global SECTION
    SECTION = title
    print(f"\n{'=' * 60}")
    print(f"  {title}")
    print('=' * 60)


def chk(desc, ok, detail=""):
    global PASS_CNT, FAIL_CNT
    if ok:
        PASS_CNT += 1
        print(f"  [PASS] {desc}")
    else:
        FAIL_CNT += 1
        print(f"  [FAIL] {desc}{(' — ' + detail) if detail else ''}")


# ═══════════════════════════════════════════════════════════
# 輔助：建立 Mock SpecItem
# ═══════════════════════════════════════════════════════════
class SI:
    """Mock SpecItem — 只包含 judgment.py 需要的欄位"""
    def __init__(self, name, spec_type, *, min_v=None, max_v=None,
                 expected=None, threshold_v=None, threshold_op=None, sub_group=None):
        self.item_name     = name
        self.spec_type     = spec_type
        self.min_value     = Decimal(str(min_v)) if min_v is not None else None
        self.max_value     = Decimal(str(max_v)) if max_v is not None else None
        self.expected_text = expected
        self.threshold_value    = Decimal(str(threshold_v)) if threshold_v is not None else None
        self.threshold_operator = threshold_op
        self.sub_group     = sub_group


def _jv(raw, spec_type, *, min_v=None, max_v=None, expected=None,
        threshold_v=None, threshold_op=None):
    from utils.spec_parser import judge_value
    return judge_value(
        raw, spec_type,
        min_value=Decimal(str(min_v)) if min_v is not None else None,
        max_value=Decimal(str(max_v)) if max_v is not None else None,
        expected_text=expected,
        threshold_value=Decimal(str(threshold_v)) if threshold_v is not None else None,
        threshold_operator=threshold_op,
    )


# ═══════════════════════════════════════════════════════════
# A. judge_value() 邊界值
# ═══════════════════════════════════════════════════════════
def test_A_judge_value():
    section("A. judge_value() — 邊界值，所有規格類型")

    # ── Range ──────────────────────────────────
    print("  [Range 165~185]")
    chk("min-1 (164.9) → NG",  _jv(164.9, "range", min_v=165, max_v=185) == "NG")
    chk("min   (165)   → OK",  _jv(165,   "range", min_v=165, max_v=185) == "OK")
    chk("mid   (175)   → OK",  _jv(175,   "range", min_v=165, max_v=185) == "OK")
    chk("max   (185)   → OK",  _jv(185,   "range", min_v=165, max_v=185) == "OK")
    chk("max+1 (185.1) → NG",  _jv(185.1, "range", min_v=165, max_v=185) == "NG")
    chk("far below (100) → NG",_jv(100,   "range", min_v=165, max_v=185) == "NG")
    chk("far above (250) → NG",_jv(250,   "range", min_v=165, max_v=185) == "NG")
    chk("str '100' → NG",      _jv("100", "range", min_v=165, max_v=185) == "NG")
    chk("str '175.0' → OK",    _jv("175.0","range",min_v=165, max_v=185) == "OK")
    chk("int 0 → NG",          _jv(0,     "range", min_v=165, max_v=185) == "NG")

    print("  [Range SKIP 情況]")
    chk("None → SKIP",         _jv(None,  "range", min_v=165, max_v=185) == "SKIP")
    chk("'' → SKIP",           _jv("",    "range", min_v=165, max_v=185) == "SKIP")
    chk("'-' → SKIP",          _jv("-",   "range", min_v=165, max_v=185) == "SKIP")
    chk("'/' → SKIP",          _jv("/",   "range", min_v=165, max_v=185) == "SKIP")

    # ── Threshold ──────────────────────────────
    print("  [Threshold >=3]")
    chk("2 → NG (>=3)",  _jv(2, "threshold", threshold_v=3, threshold_op=">=") == "NG")
    chk("3 → OK (>=3)",  _jv(3, "threshold", threshold_v=3, threshold_op=">=") == "OK")
    chk("9 → OK (>=3)",  _jv(9, "threshold", threshold_v=3, threshold_op=">=") == "OK")
    chk("None → SKIP",   _jv(None,"threshold",threshold_v=3,threshold_op=">=") == "SKIP")

    print("  [Threshold <=10]")
    chk("11 → NG (<=10)", _jv(11,"threshold",threshold_v=10,threshold_op="<=") == "NG")
    chk("10 → OK (<=10)", _jv(10,"threshold",threshold_v=10,threshold_op="<=") == "OK")
    chk("5  → OK (<=10)", _jv(5, "threshold",threshold_v=10,threshold_op="<=") == "OK")

    # ── Check ──────────────────────────────────
    print("  [Check — 所有合法打勾符號]")
    for v in ["√", "✓", "✔", "V", "v", "○", "O", "o",
              "OK", "ok", "Ok", "Y", "y", "YES", "yes",
              "合格", "正常", "良", "良好", "PASS", "pass",
              "TRUE", "True", "true", "1", "是"]:
        chk(f"'{v}' → OK", _jv(v, "check") == "OK")
    for v in ["X", "x", "N", "NG", "0", "否", "不合格", "FAIL", "False"]:
        chk(f"'{v}' → NG", _jv(v, "check") == "NG")
    chk("None → SKIP (check)", _jv(None, "check") == "SKIP")

    # ── Text ───────────────────────────────────
    print("  [Text]")
    chk("exact match → OK",    _jv("正常", "text", expected="正常") == "OK")
    chk("mismatch → NG",       _jv("異常", "text", expected="正常") == "NG")
    chk("case-insensitive OK", _jv("OK",   "text", expected="ok")   == "OK")
    chk("None → SKIP (text)",  _jv(None,   "text", expected="正常") == "SKIP")

    # ── Skip ───────────────────────────────────
    print("  [Skip]")
    chk("spec_type=skip 任何值 → SKIP", _jv(100, "skip") == "SKIP")
    chk("spec_type=skip None → SKIP",   _jv(None,"skip") == "SKIP")


# ═══════════════════════════════════════════════════════════
# B. _find_spec_for_key() — 全 5 個表單
# ═══════════════════════════════════════════════════════════
def test_B_find_spec_for_key():
    section("B. _find_spec_for_key() — 全 5 個表單類型 key 對應")
    from services.judgment import _find_spec_for_key

    def found(lookup, key, form_code, row_data=None, lookup_sg=None, col_label=None):
        return _find_spec_for_key(
            lookup, lookup_sg or {}, key, row_data or {},
            form_code, col_label,
        )

    # ── F-RD09AB / F-RD09AA 模溫 ──
    print("  [F-RD09AB — 模溫 / 壓力]")
    ab_lk = {
        "模温设定值(上模)":  SI("模温设定值(上模)",  "range", min_v=165, max_v=185),
        "模温设定值(下模)":  SI("模温设定值(下模)",  "range", min_v=160, max_v=180),
        "模温显示值1(上模)": SI("模温显示值1(上模)", "range", min_v=165, max_v=185),
        "模温显示值2(上模)": SI("模温显示值2(上模)", "range", min_v=165, max_v=185),
        "模温显示值3(上模)": SI("模温显示值3(上模)", "range", min_v=165, max_v=185),
        "模温显示值4(上模)": SI("模温显示值4(上模)", "range", min_v=165, max_v=185),
        "模温显示值1(下模)": SI("模温显示值1(下模)", "range", min_v=160, max_v=180),
        "合模压力(ton)":     SI("合模压力(ton)",     "range", min_v=125, max_v=145),
        "注塑压强(kgf/cm²)":SI("注塑压强(kgf/cm²)","range", min_v=50,  max_v=80),
        "固化时间(sec)":     SI("固化时间(sec)",     "range", min_v=60,  max_v=90),
        "注塑时间(sec)":     SI("注塑时间(sec)",     "range", min_v=5,   max_v=15),
        "预热台温度(℃)":    SI("预热台温度(℃)",    "range", min_v=40,  max_v=80),
        "模具状态":          SI("模具状态",          "check"),
        "定位针状态":        SI("定位针状态",        "check"),
    }
    for key, expected_name in [
        ("set_temp_上模",     "模温设定值(上模)"),
        ("set_temp_下模",     "模温设定值(下模)"),
        ("disp_temp_上模_1",  "模温显示值1(上模)"),
        ("disp_temp_上模_2",  "模温显示值2(上模)"),
        ("disp_temp_上模_3",  "模温显示值3(上模)"),
        ("disp_temp_上模_4",  "模温显示值4(上模)"),
        ("disp_temp_下模_1",  "模温显示值1(下模)"),
        ("clamp_pressure",    "合模压力(ton)"),
        ("inject_pressure",   "注塑压强(kgf/cm²)"),
        ("cure_time",         "固化时间(sec)"),
        ("inject_time",       "注塑时间(sec)"),
        ("preheat_temp",      "预热台温度(℃)"),
        ("mold_status",       "模具状态"),
        ("pin_status",        "定位针状态"),
    ]:
        r = found(ab_lk, key, "F-RD09AB")
        chk(f"  {key} → {expected_name}", r is not None and r.item_name == expected_name,
            f"got {r.item_name if r else None}")

    for bad_key in ["wash_reason", "wash_method", "mold_no", "mold_count", "signer"]:
        chk(f"  {bad_key} → None (不判定)", found(ab_lk, bad_key, "F-RD09AB") is None)

    # ── F-RD09AJ 溫區 ──
    print("  [F-RD09AJ — 溫區 SV/PV / 氣體 / 冷卻水]")
    aj_lk = {}
    for i in range(1, 9):
        aj_lk[f"温区{i}设定SV(℃)"] = SI(f"温区{i}设定SV(℃)", "range", min_v=200, max_v=230)
        aj_lk[f"温区{i}实际PV(℃)"] = SI(f"温区{i}实际PV(℃)", "range", min_v=195, max_v=235)
    aj_lk["N2"]             = SI("N2",             "threshold", threshold_v=95, threshold_op=">=")
    aj_lk["N2(PRE)"]        = SI("N2(PRE)",        "threshold", threshold_v=95, threshold_op=">=")
    aj_lk["冷却水流量LPM"]  = SI("冷却水流量LPM",  "threshold", threshold_v=3,  threshold_op=">=")

    for i in range(1, 9):
        r = found(aj_lk, f"sv_{i}", "F-RD09AJ")
        chk(f"  sv_{i} → 温区{i}设定SV(℃)", r is not None and r.item_name == f"温区{i}设定SV(℃)")
        r = found(aj_lk, f"pv_{i}", "F-RD09AJ")
        chk(f"  pv_{i} → 温区{i}实际PV(℃)", r is not None and r.item_name == f"温区{i}实际PV(℃)")
    chk("  gas_N2 → N2",           (r := found(aj_lk, "gas_N2", "F-RD09AJ"))  is not None and r.item_name == "N2")
    chk("  gas_N2(PRE) → N2(PRE)", (r := found(aj_lk, "gas_N2(PRE)", "F-RD09AJ")) is not None and r.item_name == "N2(PRE)")
    chk("  cooling_water → 冷却水流量LPM",
        (r := found(aj_lk, "cooling_water", "F-RD09AJ")) is not None and r.item_name == "冷却水流量LPM")

    # ── F-RD09AK 尺寸 + sub_group ──
    print("  [F-RD09AK — meas_* + sub_group]")
    ak_lk = {
        "meas_1": SI("meas_1", "range", min_v=0.5, max_v=0.7, sub_group="A"),
        "meas_2": SI("meas_2", "range", min_v=0.5, max_v=0.7, sub_group="A"),
        "meas_1_G1": SI("meas_1_G1", "range", min_v=1.0, max_v=1.2, sub_group="G1"),
    }
    ak_lk_sg = {
        ("meas_1", "A"):  ak_lk["meas_1"],
        ("meas_2", "A"):  ak_lk["meas_2"],
        ("meas_1", "G1"): ak_lk["meas_1_G1"],
    }
    row_A  = {"values": {"part": "A"}}
    row_G1 = {"values": {"part": "G1"}}
    chk("  meas_1 part=A  → A spec",
        (r := found(ak_lk, "meas_1", "F-RD09AK", row_A, ak_lk_sg)) is not None
        and float(r.min_value) == 0.5)
    chk("  meas_1 part=G1 → G1 spec",
        (r := found(ak_lk, "meas_1", "F-RD09AK", row_G1, ak_lk_sg)) is not None
        and float(r.min_value) == 1.0)
    chk("  meas_2 part=A  → A spec",
        found(ak_lk, "meas_2", "F-RD09AK", row_A, ak_lk_sg) is not None)

    # ── F-QA1021 勾選項 ──
    print("  [F-QA1021 — 勾選 check items]")
    qa_lk = {k: SI(k, "check") for k in
             ["电源开关", "风扇飘带", "空气滤网", "风扇角度", "风扇风速"]}
    for key in ["电源开关", "风扇飘带", "空气滤网", "风扇角度", "风扇风速"]:
        chk(f"  {key} → check spec", found(qa_lk, key, "F-QA1021") is not None)

    # ── GenericParser col_N 模糊比對 ──
    print("  [GenericParser — col_N 4 層模糊比對]")
    gen_lk = {
        "顶针高度(um)":  SI("顶针高度(um)",  "range", min_v=300, max_v=750),
        "温度":          SI("温度",           "range", min_v=20,  max_v=30),
        "压力":          SI("压力",           "range", min_v=1,   max_v=5),
    }
    col_map = {
        "col_0": "顶针高度（um）",   # 全形括號 → 需標準化
        "col_1": "温度(℃)",          # 含單位 → strip 後比對
        "col_2": "壓力(kPa)",        # 子字串 '壓力' vs '压力'（不同字，測無對應）
        "col_3": "顶针高度",          # 去掉括號後精確比對
    }
    # col_0: 全形 → 標準化後完全比對
    r0 = found(gen_lk, "col_0", "F-RD09F1", {}, {}, col_map)
    chk("  col_0 '顶针高度（um）'（全形括號）→ 找到", r0 is not None, f"got {r0}")
    # col_1: strip units → '温度(℃)' 去掉(℃) = '温度'
    r1 = found(gen_lk, "col_1", "F-RD09F1", {}, {}, col_map)
    chk("  col_1 '温度(℃)' → strip unit → '温度' 比對", r1 is not None, f"got {r1}")
    # col_3: 子字串完全比對
    r3 = found(gen_lk, "col_3", "F-RD09F1", {}, {}, col_map)
    chk("  col_3 '顶针高度' → 子字串比對", r3 is not None, f"got {r3}")


# ═══════════════════════════════════════════════════════════
# C. _judge_rd09ab() — 3 層備援 × NG/OK 場景
# ═══════════════════════════════════════════════════════════
def test_C_rd09ab_fallback():
    section("C. _judge_rd09ab() — 3 層備援 × NG/OK 判定")
    from services.judgment import _find_spec_for_key
    from utils.spec_parser import judge_value

    TEMP_SPEC = SI("模温显示值1(上模)", "range", min_v=165, max_v=185)
    CLAMP_SPEC = SI("合模压力(ton)", "range", min_v=125, max_v=145)
    lookup_AX = {"模温显示值1(上模)": TEMP_SPEC, "合模压力(ton)": CLAMP_SPEC}
    lookup_AY = {"模温显示值1(上模)": SI("模温显示值1(上模)", "range", min_v=165, max_v=185),
                 "合模压力(ton)": SI("合模压力(ton)", "range", min_v=125, max_v=145)}

    # Simulate spec_cache keyed by equipment_id (as stored in DB)
    spec_cache = {
        "WPRN-0001_A_X": {"lookup": lookup_AX},
        "WPRN-0001_A_Y": {"lookup": lookup_AY},
    }

    def judge_row(wash_reason, wash_method, temp_val, clamp_val):
        """Simulate what _judge_rd09ab does for one row."""
        equipment_id = "WPRN-0001"
        spec_key = f"{equipment_id}_{wash_reason}_{wash_method}"
        row_spec = spec_cache.get(spec_key)

        # Fallback 1
        if not row_spec and wash_reason:
            for k, c in spec_cache.items():
                if k.startswith(f"{equipment_id}_{wash_reason}"):
                    row_spec = c
                    break

        # Fallback 2 (新增修復)
        if not row_spec:
            row_spec = next(iter(spec_cache.values()), None)

        if not row_spec:
            return "NO_SPEC", "NO_SPEC"

        def _j(key, val):
            si = _find_spec_for_key(row_spec["lookup"], {}, key, {}, "F-RD09AB")
            if not si:
                return "SKIP"
            return judge_value(val, si.spec_type, min_value=si.min_value, max_value=si.max_value)

        return _j("disp_temp_上模_1", temp_val), _j("clamp_pressure", clamp_val)

    print("  [場景 1：wash_reason=A, wash_method=X → 精確比對]")
    t, c = judge_row("A", "X", 100, 135)
    chk("  temp=100 → NG（精確比對）", t == "NG", f"got {t}")
    chk("  clamp=135 → OK（精確比對）", c == "OK", f"got {c}")
    t, c = judge_row("A", "X", 175, 160)
    chk("  temp=175 → OK", t == "OK", f"got {t}")
    chk("  clamp=160 → NG（超過 145）", c == "NG", f"got {c}")

    print("  [場景 2：wash_reason=A, wash_method=UNKNOWN → Fallback 1]")
    t, c = judge_row("A", "UNKNOWN_METHOD", 100, 135)
    chk("  temp=100 → NG（Fallback 1）", t == "NG", f"got {t}")
    chk("  clamp=135 → OK（Fallback 1）", c == "OK", f"got {c}")

    print("  [場景 3：wash_reason=None → Fallback 2（合併格空值，主要 bug 場景）]")
    t, c = judge_row(None, None, 100, 135)
    chk("  temp=100 → NG（Fallback 2，修復後）", t == "NG", f"got {t}")
    chk("  clamp=135 → OK（Fallback 2）", c == "OK", f"got {c}")

    t, c = judge_row(None, None, 175, 120)
    chk("  temp=175 → OK（Fallback 2）", t == "OK", f"got {t}")
    chk("  clamp=120 → NG（低於 125）", c == "NG", f"got {c}")

    print("  [邊界值]")
    for temp, expected, desc in [
        (164.9, "NG",  "164.9 低於下限 165"),
        (165,   "OK",  "165 = 下限"),
        (185,   "OK",  "185 = 上限"),
        (185.1, "NG",  "185.1 超過上限 185"),
    ]:
        t, _ = judge_row(None, None, temp, 135)
        chk(f"  temp={temp} → {expected}（{desc}）", t == expected, f"got {t}")

    print("  [SKIP 值]")
    for bad_val in [None, "", "/", "-"]:
        t, _ = judge_row(None, None, bad_val, 135)
        chk(f"  temp={repr(bad_val)} → SKIP", t == "SKIP", f"got {t}")


# ═══════════════════════════════════════════════════════════
# D. 全 pipeline：解析真實檔案 → 注入 NG → 判定 → 驗證
# ═══════════════════════════════════════════════════════════
TEST_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "test")


def _run_judgment_with_mock_spec(parsed_data, spec_lookup, form_code,
                                 ng_overrides: dict) -> dict:
    """
    直接呼叫判定邏輯（不走 DB），使用 mock spec_lookup。
    ng_overrides: {key: ng_value} — 將第一行的這些欄位值設為已知 NG 值。
    """
    from services.judgment import _find_spec_for_key, _compute_row_judgment, _format_spec_display
    from utils.spec_parser import judge_value

    col_label_map = {}
    for h in parsed_data.get("headers", []):
        if h.get("key", "").startswith("col_") and h.get("label"):
            col_label_map[h["key"]] = h["label"]

    # 注入 NG 值到第 0 行
    rows = copy.deepcopy(parsed_data.get("rows", []))
    if rows and ng_overrides:
        for k, v in ng_overrides.items():
            rows[0].get("values", {})[k] = v

    judged_rows = []
    ok_cnt = ng_cnt = skip_cnt = 0
    for row_data in rows:
        judged_values = {}
        for key, raw in row_data.get("values", {}).items():
            si = _find_spec_for_key(
                spec_lookup, {}, key, row_data, form_code, col_label_map)
            if si:
                r = judge_value(raw, si.spec_type,
                                min_value=si.min_value, max_value=si.max_value,
                                expected_text=si.expected_text,
                                threshold_value=si.threshold_value,
                                threshold_operator=si.threshold_operator)
                if r == "OK":   ok_cnt += 1
                elif r == "NG": ng_cnt += 1
                else:           skip_cnt += 1
            else:
                r = "SKIP"
                skip_cnt += 1
            judged_values[key] = {"raw": raw, "judgment": r, "spec": ""}
        judged_rows.append({"row_judgment": _compute_row_judgment(judged_values),
                            "values": judged_values})

    has_ng = any(jr["row_judgment"] == "NG" for jr in judged_rows)
    return {
        "ok": ok_cnt, "ng": ng_cnt, "skip": skip_cnt,
        "has_ng": has_ng,
        "row_0_judgments": judged_rows[0]["values"] if judged_rows else {},
    }


def test_D_pipeline():
    section("D. 全 pipeline：解析真實檔案 → 注入 NG → 判定 → 驗證")
    from openpyxl import load_workbook

    # ── F-QA1021 ──
    print("  [F-QA1021]")
    fp = os.path.join(TEST_DIR, "F-QA1021_离子消散设备点检记录表.xlsx")
    if os.path.exists(fp):
        from parsers.qa1021_parser import QA1021Parser
        wb = load_workbook(fp, data_only=True)
        p = QA1021Parser()
        parsed = p.parse_sheet(wb.worksheets[0], wb.sheetnames[0])
        wb.close()
        qa_lk = {k: SI(k, "check") for k in
                 ["电源开关", "风扇飘带", "空气滤网", "风扇角度", "风扇风速"]}
        # 注入 NG：将第一行 电源开关 設為 "X"
        r = _run_judgment_with_mock_spec(parsed, qa_lk, "F-QA1021",
                                        {"电源开关": "X"})
        chk("  注入 电源开关=X → has_ng=True",  r["has_ng"])
        chk("  电源开关='X' 判為 NG",
            r["row_0_judgments"].get("电源开关", {}).get("judgment") == "NG")
        # 注入 OK
        r2 = _run_judgment_with_mock_spec(parsed, qa_lk, "F-QA1021",
                                         {"电源开关": "√"})
        e = r2["row_0_judgments"].get("电源开关", {}).get("judgment")
        chk("  电源开关='√' 判為 OK", e == "OK", f"got {e}")
    else:
        print("  [SKIP] 檔案不存在")

    # ── F-RD09AA ──
    print("  [F-RD09AA]")
    fp = os.path.join(TEST_DIR, "F-RD09AA-Auto Mold 机台检查记录表.xlsx")
    if os.path.exists(fp):
        from parsers.rd09aa_parser import RD09AAParser
        wb = load_workbook(fp, data_only=True)
        p = RD09AAParser()
        data_sheets = [s for s in wb.sheetnames if s != "汇总"]
        if data_sheets:
            parsed = p.parse_sheet(wb[data_sheets[0]], data_sheets[0])
            wb.close()
            aa_lk = {
                "模温显示值1(上模)": SI("模温显示值1(上模)", "range", min_v=165, max_v=185),
                "合模压力(ton)":     SI("合模压力(ton)",     "range", min_v=125, max_v=145),
            }
            # 注入 NG
            r = _run_judgment_with_mock_spec(parsed, aa_lk, "F-RD09AA",
                                            {"disp_temp_上模_1": 100,
                                             "clamp_pressure": 200})
            chk("  注入 temp=100, clamp=200 → has_ng=True", r["has_ng"])
            t_j = r["row_0_judgments"].get("disp_temp_上模_1", {}).get("judgment")
            c_j = r["row_0_judgments"].get("clamp_pressure", {}).get("judgment")
            chk("  disp_temp_上模_1=100 → NG", t_j == "NG", f"got {t_j}")
            chk("  clamp_pressure=200  → NG", c_j == "NG", f"got {c_j}")
            # 注入 OK
            r2 = _run_judgment_with_mock_spec(parsed, aa_lk, "F-RD09AA",
                                             {"disp_temp_上模_1": 175,
                                              "clamp_pressure": 135})
            t_j2 = r2["row_0_judgments"].get("disp_temp_上模_1", {}).get("judgment")
            chk("  disp_temp_上模_1=175 → OK", t_j2 == "OK", f"got {t_j2}")
        else:
            wb.close(); print("  [SKIP] 無 data sheet")
    else:
        print("  [SKIP] 檔案不存在")

    # ── F-RD09AB ──
    print("  [F-RD09AB]")
    fp = os.path.join(TEST_DIR, "F-RD09AB-Auto Mold 洗模检查记录表.xlsx")
    if os.path.exists(fp):
        from parsers.rd09ab_parser import RD09ABParser
        wb = load_workbook(fp, data_only=True)
        p = RD09ABParser()
        data_sheets = [s for s in wb.sheetnames if s != "汇总"]
        if data_sheets:
            parsed = p.parse_sheet(wb[data_sheets[0]], data_sheets[0])
            wb.close()
            rows = parsed.get("rows", [])
            none_wr = sum(1 for r in rows if r["values"].get("wash_reason") is None)
            print(f"    共 {len(rows)} 行，{none_wr} 行 wash_reason=None（合併格場景）")
            ab_lk = {
                "模温显示值1(上模)": SI("模温显示值1(上模)", "range", min_v=165, max_v=185),
                "合模压力(ton)":     SI("合模压力(ton)",     "range", min_v=125, max_v=145),
            }
            # None wash_reason 的行注入 NG（這是原本 bug 場景）
            r = _run_judgment_with_mock_spec(parsed, ab_lk, "F-RD09AB",
                                            {"disp_temp_上模_1": 100})
            chk("  注入 temp=100（row 0）→ has_ng=True", r["has_ng"])
            t_j = r["row_0_judgments"].get("disp_temp_上模_1", {}).get("judgment")
            chk("  disp_temp_上模_1=100 → NG（無論 wash_reason 是否 None）",
                t_j == "NG", f"got {t_j}")
        else:
            wb.close(); print("  [SKIP] 無 data sheet")
    else:
        print("  [SKIP] 檔案不存在")

    # ── F-RD09AJ ──
    print("  [F-RD09AJ]")
    fp = os.path.join(TEST_DIR, "F-RD09AJ-RO 焊接炉检查记录表.xlsx")
    if os.path.exists(fp):
        from parsers.rd09aj_parser import RD09AJParser
        wb = load_workbook(fp, data_only=True)
        p = RD09AJParser()
        data_sheets = [s for s in wb.sheetnames if s != "汇总"]
        if data_sheets:
            parsed = p.parse_sheet(wb[data_sheets[0]], data_sheets[0])
            wb.close()
            aj_lk = {}
            for i in range(1, 9):
                aj_lk[f"温区{i}设定SV(℃)"] = SI(f"温区{i}设定SV(℃)", "range", min_v=200, max_v=230)
                aj_lk[f"温区{i}实际PV(℃)"] = SI(f"温区{i}实际PV(℃)", "range", min_v=195, max_v=235)
            # 注入 sv_1=100（遠低於 200）
            r = _run_judgment_with_mock_spec(parsed, aj_lk, "F-RD09AJ",
                                            {"sv_1": 100, "pv_1": 100})
            chk("  注入 sv_1=100, pv_1=100 → has_ng=True", r["has_ng"])
            sv_j = r["row_0_judgments"].get("sv_1", {}).get("judgment")
            chk("  sv_1=100 → NG", sv_j == "NG", f"got {sv_j}")
        else:
            wb.close(); print("  [SKIP] 無 data sheet")
    else:
        print("  [SKIP] 檔案不存在")

    # ── F-RD09AK ──
    print("  [F-RD09AK]")
    fp = os.path.join(TEST_DIR, "F-RD09AK_SMD(Clip）切弯脚尺寸检查记录表.xlsx")
    if os.path.exists(fp):
        from parsers.rd09ak_parser import RD09AKParser
        wb = load_workbook(fp, data_only=True)
        p = RD09AKParser()
        data_sheets = [s for s in wb.sheetnames if s != "汇总"]
        if data_sheets:
            parsed = p.parse_sheet(wb[data_sheets[0]], data_sheets[0])
            wb.close()
            rows = parsed.get("rows", [])
            if rows:
                # 找出第一行使用的 meas key
                meas_keys = [k for k in rows[0].get("values", {}) if k.startswith("meas_")]
                if meas_keys:
                    ak_lk = {k: SI(k, "range", min_v=0.5, max_v=0.7) for k in meas_keys}
                    r = _run_judgment_with_mock_spec(parsed, ak_lk, "F-RD09AK",
                                                    {meas_keys[0]: 0.1})
                    chk(f"  注入 {meas_keys[0]}=0.1 → has_ng=True", r["has_ng"])
                    m_j = r["row_0_judgments"].get(meas_keys[0], {}).get("judgment")
                    chk(f"  {meas_keys[0]}=0.1 → NG", m_j == "NG", f"got {m_j}")
                else:
                    print("  [SKIP] 無 meas_* key")
            else:
                print("  [SKIP] 無資料行")
        else:
            wb.close(); print("  [SKIP] 無 data sheet")
    else:
        print("  [SKIP] 檔案不存在")


# ═══════════════════════════════════════════════════════════
# E. 匯出標注：NG 紅色 / OK 無色 / 判定欄
# ═══════════════════════════════════════════════════════════
def test_E_export_annotation():
    section("E. 匯出標注 — NG 紅色儲存格 / 判定欄 OK/NG/—")
    from openpyxl import Workbook
    from services.export_service import _annotate_sheet, FILL_NG

    # 建立一個簡單的工作表：row 5 = data row
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "日期"
    ws["B1"] = "溫度"
    ws["C1"] = "壓力"
    ws["A5"] = "2026-01-01"
    ws["B5"] = 100       # 溫度：NG（超規）
    ws["C5"] = 135       # 壓力：OK
    ws["A6"] = "2026-01-02"
    ws["B6"] = 175       # 溫度：OK
    ws["C6"] = 200       # 壓力：NG
    ws["A7"] = "2026-01-03"
    ws["B7"] = None      # SKIP

    # Mock judged_data（與工作表對應）
    judged_data = {
        "has_spec": True,
        "judged_rows": [
            {
                "row_judgment": "NG",
                "values": {
                    "temp": {"raw": 100, "judgment": "NG", "spec": "165~185"},
                    "pressure": {"raw": 135, "judgment": "OK", "spec": "125~145"},
                }
            },
            {
                "row_judgment": "NG",
                "values": {
                    "temp": {"raw": 175, "judgment": "OK", "spec": "165~185"},
                    "pressure": {"raw": 200, "judgment": "NG", "spec": "125~145"},
                }
            },
            {
                "row_judgment": "SKIP",
                "values": {
                    "temp": {"raw": None, "judgment": "SKIP", "spec": ""},
                    "pressure": {"raw": 130, "judgment": "OK", "spec": "125~145"},
                }
            },
        ],
        "meta": {
            "judgment_col": None,  # 需自動建立
            "row_map": [
                {"row": 5, "cells": {"temp": [5, 2], "pressure": [5, 3]}},
                {"row": 6, "cells": {"temp": [6, 2], "pressure": [6, 3]}},
                {"row": 7, "cells": {"temp": [7, 2], "pressure": [7, 3]}},
            ],
        }
    }

    _annotate_sheet(ws, judged_data)

    # 驗證 NG 儲存格（B5, C6）是否有紅色 fill
    b5_fill = ws["B5"].fill
    c6_fill = ws["C6"].fill
    b6_fill = ws["B6"].fill  # OK → 不應紅色

    chk("B5(temp=100, NG) fill = FILL_NG",
        b5_fill.start_color.rgb == FILL_NG.start_color.rgb,
        f"got {b5_fill.start_color.rgb}")
    chk("C6(pressure=200, NG) fill = FILL_NG",
        c6_fill.start_color.rgb == FILL_NG.start_color.rgb,
        f"got {c6_fill.start_color.rgb}")
    chk("B6(temp=175, OK) fill ≠ FILL_NG",
        b6_fill.start_color.rgb != FILL_NG.start_color.rgb,
        f"got {b6_fill.start_color.rgb}")
    chk("C5(pressure=135, OK) fill ≠ FILL_NG",
        ws["C5"].fill.start_color.rgb != FILL_NG.start_color.rgb)

    # 驗證判定欄 — 找最後一欄（自動建立）
    # _annotate_sheet 會在最右欄+1 寫 判定
    max_col = ws.max_column
    # 判定欄應在右側
    row5_judgment = ws.cell(row=5, column=max_col).value
    row6_judgment = ws.cell(row=6, column=max_col).value
    row7_judgment = ws.cell(row=7, column=max_col).value

    chk(f"Row 5 判定欄 = 'NG'",   row5_judgment == "NG",   f"got {row5_judgment}")
    chk(f"Row 6 判定欄 = 'NG'",   row6_judgment == "NG",   f"got {row6_judgment}")
    chk(f"Row 7 判定欄 = '—'",    row7_judgment == "—",    f"got {row7_judgment}")

    # Row 5 判定欄也應有紅色
    r5_jcell = ws.cell(row=5, column=max_col)
    chk("Row 5 判定欄 fill = FILL_NG",
        r5_jcell.fill.start_color.rgb == FILL_NG.start_color.rgb,
        f"got {r5_jcell.fill.start_color.rgb}")

    wb.close()


# ═══════════════════════════════════════════════════════════
# F. parse_spec_string() 全格式
# ═══════════════════════════════════════════════════════════
def test_F_spec_string():
    section("F. parse_spec_string() — 所有格式")
    from utils.spec_parser import parse_spec_string

    cases = [
        ("165~185",    "range",     165,  185,  None, None,  None),
        ("165～185",   "range",     165,  185,  None, None,  None),  # 全形
        ("0~0.05",     "range",     0,    0.05, None, None,  None),
        ("≥3",         "threshold", None, None, None, 3,     ">="),
        (">=3",        "threshold", None, None, None, 3,     ">="),
        ("≤10",        "threshold", None, None, None, 10,    "<="),
        ("<=10",       "threshold", None, None, None, 10,    "<="),
        (">5",         "threshold", None, None, None, 5,     ">"),
        ("<5",         "threshold", None, None, None, 5,     "<"),
        ("√",          "check",     None, None, "√",  None,  None),
        ("✓",          "check",     None, None, "√",  None,  None),
        ("V",          "check",     None, None, "√",  None,  None),
        ("OK",         "text",      None, None, "OK", None,  None),
        ("/",          "skip",      None, None, None, None,  None),
        ("-",          "skip",      None, None, None, None,  None),
        ("",           "skip",      None, None, None, None,  None),
        (None,         "skip",      None, None, None, None,  None),
    ]
    for raw, exp_type, min_v, max_v, exp_text, th_v, th_op in cases:
        r = parse_spec_string(raw)
        ok = r["spec_type"] == exp_type
        if min_v is not None:
            ok = ok and abs(float(r.get("min_value", 0)) - min_v) < 0.001
        if max_v is not None:
            ok = ok and abs(float(r.get("max_value", 0)) - max_v) < 0.001
        if th_v is not None:
            ok = ok and abs(float(r.get("threshold_value", 0)) - th_v) < 0.001
        if th_op is not None:
            ok = ok and r.get("threshold_operator") == th_op
        if exp_text is not None:
            ok = ok and r.get("expected_text") == exp_text
        chk(f"  '{raw}' → {exp_type}", ok, str(r))


# ═══════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("全面 OK/NG 判定測試")

    test_A_judge_value()
    test_B_find_spec_for_key()
    test_C_rd09ab_fallback()
    test_D_pipeline()
    test_E_export_annotation()
    test_F_spec_string()

    print(f"\n{'=' * 60}")
    print(f"  最終結果：{PASS_CNT} PASS，{FAIL_CNT} FAIL")
    print('=' * 60)
    if FAIL_CNT > 0:
        sys.exit(1)
