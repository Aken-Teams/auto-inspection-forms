"""Test GenericParser judgment: create specs from actual headers, upload, verify OK/NG."""
import sys
import os
import json

sys.stdout.reconfigure(encoding='utf-8')

import httpx
from openpyxl import load_workbook

BASE = "http://localhost:12061"
AU_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "AU未建立规格点检表", "AU")

client = httpx.Client(base_url=BASE, timeout=30)


def read_file_headers(filepath):
    """Read headers from a file using the same logic as GenericParser."""
    wb = load_workbook(filepath, data_only=True)
    results = {}
    for sn in wb.sheetnames:
        if sn == "汇总":
            continue
        ws = wb[sn]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        # Find header row (first row with >= 3 non-empty cells)
        header_row = None
        for row_idx in range(1, min(20, max_row + 1)):
            non_empty = 0
            for col in range(1, min(max_col + 1, 50)):
                cell = ws.cell(row=row_idx, column=col)
                if cell.value is not None:
                    non_empty += 1
            if non_empty >= 3:
                header_row = row_idx
                break

        if not header_row:
            results[sn] = {"headers": [], "sample_row": {}}
            continue

        headers = []
        col_map = {}
        for col in range(1, min(max_col + 1, 50)):
            val = ws.cell(row=header_row, column=col).value
            if val:
                label = str(val).replace("\n", " ").strip()
                key = f"col_{col}"
                headers.append({"key": key, "label": label})
                col_map[col] = key

        # Read first data row to see sample values
        sample = {}
        if header_row + 1 <= max_row:
            for col, key in col_map.items():
                val = ws.cell(row=header_row + 1, column=col).value
                sample[key] = val

        # Count data rows
        data_rows = 0
        for row_idx in range(header_row + 1, max_row + 1):
            has_data = False
            for col in col_map:
                if ws.cell(row=row_idx, column=col).value is not None:
                    has_data = True
                    break
            if has_data:
                first_val = ws.cell(row=row_idx, column=1).value
                if first_val and ("备注" in str(first_val) or "REV" in str(first_val)):
                    break
                data_rows += 1

        results[sn] = {"headers": headers, "sample_row": sample, "data_rows": data_rows}

    wb.close()
    return results


def ensure_form_type(form_code, form_name):
    """Ensure form type exists, create if not."""
    resp = client.get("/api/specs/form-types")
    codes = [ft["form_code"] for ft in resp.json()]
    if form_code not in codes:
        import re
        client.post("/api/specs/form-types", json={
            "form_code": form_code,
            "form_name": form_name,
            "file_pattern": re.escape(form_code),
        })
        print(f"  Created form type: {form_code}")
    else:
        print(f"  Form type exists: {form_code}")


def create_specs_from_headers(form_code, sheet_name, headers, sample_row):
    """Create a spec group with items matching the file's actual headers."""
    # First create spec group (equipment_id = sheet_name)
    resp = client.post(f"/api/specs/form-types/{form_code}/specs", json={
        "equipment_id": sheet_name,
        "equipment_name": sheet_name,
    })
    if resp.status_code not in (200, 201):
        print(f"  Spec group create: {resp.status_code} {resp.text[:200]}")
        return False

    spec_data = resp.json()
    spec_id = spec_data.get("id")
    if not spec_id:
        print(f"  No spec_id in response: {spec_data}")
        return False

    # Build spec items based on headers + sample values
    items = []
    for h in headers:
        label = h["label"]
        sample_val = sample_row.get(h["key"])

        # Skip non-data columns (序号, 日期, 时间, 签名, 备注 etc.)
        skip_labels = ["序号", "日期", "时间", "签名", "确认", "备注", "编号", "NO", "No.", "no."]
        if any(s in label for s in skip_labels):
            continue

        # Determine spec type from sample value
        spec_type = "skip"
        min_val = None
        max_val = None
        expected_text = None

        if sample_val is not None:
            try:
                num = float(sample_val)
                # Create a range spec: ±20% of value (reasonable for testing)
                spec_type = "range"
                margin = abs(num * 0.3) if num != 0 else 10
                min_val = round(num - margin, 2)
                max_val = round(num + margin, 2)
            except (ValueError, TypeError):
                sv = str(sample_val).strip()
                if sv in ("√", "✓", "OK", "○", "〇", "正常", "良好", "V", "v"):
                    spec_type = "check"
                    expected_text = sv
                elif sv:
                    spec_type = "text"
                    expected_text = sv

        if spec_type == "skip":
            continue

        items.append({
            "item_name": label,
            "spec_type": spec_type,
            "min_value": min_val,
            "max_value": max_val,
            "expected_text": expected_text,
        })

    if not items:
        print(f"  No valid spec items could be created for {sheet_name}")
        return False

    # Update spec with items
    resp = client.put(f"/api/specs/specs/{spec_id}", json={"items": items})
    if resp.status_code == 200:
        print(f"  Created spec for {sheet_name}: {len(items)} items")
        for it in items[:5]:
            disp = f"{it['min_value']}~{it['max_value']}" if it['spec_type'] == 'range' else it.get('expected_text', '')
            print(f"    {it['item_name']}: {it['spec_type']} ({disp})")
        if len(items) > 5:
            print(f"    ... and {len(items) - 5} more")
        return True
    else:
        print(f"  Update spec items failed: {resp.status_code} {resp.text[:200]}")
        return False


def upload_and_check(filepath, form_code):
    """Upload file and check judgment results."""
    import shutil
    import tempfile

    fname = os.path.basename(filepath)
    tmp = os.path.join(tempfile.gettempdir(), fname)
    shutil.copy2(filepath, tmp)
    with open(tmp, "rb") as f:
        resp = client.post("/api/upload", files={"file": (fname, f, "application/octet-stream")})
    os.remove(tmp)

    if resp.status_code != 200:
        print(f"  Upload failed: {resp.status_code} {resp.text[:200]}")
        return None

    data = resp.json()
    upload_id = data.get("upload_id") or data.get("id")
    if not upload_id:
        print(f"  No upload_id: {json.dumps(data, ensure_ascii=False)[:300]}")
        return None

    # Get details
    detail = client.get(f"/api/results/{upload_id}").json()
    matched = detail.get("form_code")
    sheets = detail.get("sheets", [])

    print(f"\n  Upload results for {fname}:")
    print(f"    Matched form: {matched} (expected: {form_code})")
    print(f"    Total sheets: {len(sheets)}")

    ok = ng = no_spec = skip_all = err = 0
    for s in sheets:
        result = s["overall_result"]
        if result == "OK":
            ok += 1
        elif result == "NG":
            ng += 1
        elif result == "NO_SPEC":
            no_spec += 1
        elif result == "ERROR":
            err += 1

        # Show judgment detail
        jd = s.get("judged_data", {})
        summary = jd.get("summary", {})
        print(f"    {s['sheet_name']}: {result} (ok={summary.get('ok',0)}, ng={summary.get('ng',0)}, skip={summary.get('skip',0)})")

        # Show first few judged values
        judged_rows = jd.get("judged_rows", [])
        if judged_rows:
            first_row = judged_rows[0]
            vals = first_row.get("values", {})
            shown = 0
            for k, v in vals.items():
                if v["judgment"] != "SKIP" and shown < 3:
                    print(f"      {k}: raw={v['raw']} -> {v['judgment']} (spec: {v['spec']})")
                    shown += 1

    print(f"\n    Summary: OK={ok}, NG={ng}, NO_SPEC={no_spec}, ERROR={err}")

    success = ok + ng > 0
    return {"ok": ok, "ng": ng, "no_spec": no_spec, "success": success}


def test_file(filename, form_code):
    """Full test: read headers, create specs, upload, verify."""
    filepath = os.path.join(AU_DIR, filename)
    if not os.path.exists(filepath):
        print(f"[SKIP] File not found: {filename}")
        return

    print(f"\n{'='*80}")
    print(f"TESTING: {form_code} ({filename})")
    print(f"{'='*80}")

    # Step 1: Read file structure
    print("\n  Step 1: Reading file structure...")
    sheet_data = read_file_headers(filepath)
    for sn, info in sheet_data.items():
        print(f"    Sheet '{sn}': {len(info['headers'])} headers, {info.get('data_rows', '?')} data rows")
        for h in info["headers"][:8]:
            sample = info["sample_row"].get(h["key"], "")
            print(f"      {h['key']} = '{h['label']}' (sample: {sample})")

    # Step 2: Ensure form type
    print("\n  Step 2: Ensuring form type...")
    ensure_form_type(form_code, filename.replace(".xlsx", ""))

    # Step 3: Create specs from first sheet's headers
    print("\n  Step 3: Creating specs from headers...")
    for sn, info in sheet_data.items():
        if info["headers"] and info.get("data_rows", 0) > 0:
            create_specs_from_headers(form_code, sn, info["headers"], info["sample_row"])

    # Step 4: Upload and check
    print("\n  Step 4: Uploading and checking judgment...")
    result = upload_and_check(filepath, form_code)

    if result and result["success"]:
        print(f"\n  [PASS] {form_code}: Judgment working! OK={result['ok']}, NG={result['ng']}")
    elif result:
        print(f"\n  [FAIL] {form_code}: No OK/NG judgments produced. NO_SPEC={result['no_spec']}")
    else:
        print(f"\n  [FAIL] {form_code}: Upload/processing failed")

    return result


if __name__ == "__main__":
    import re

    print("GenericParser Judgment Test")
    print(f"AU directory: {AU_DIR}")

    # Pick 3 non-built-in files to test
    builtin = {"F-QA1021", "F-RD09AA", "F-RD09AB", "F-RD09AJ", "F-RD09AK"}
    form_code_re = re.compile(r"(F-[A-Z]{2}\d{2,4}[A-Z0-9]{0,3})(?=[_\-\s.\u4e00-\u9fff]|$)", re.IGNORECASE)

    test_files = []
    for fname in sorted(os.listdir(AU_DIR)):
        m = form_code_re.search(fname)
        if m:
            fc = m.group(1).upper()
            if fc not in builtin:
                test_files.append((fname, fc))

    # Test first 3
    results = {}
    for fname, fc in test_files[:3]:
        r = test_file(fname, fc)
        results[fc] = r

    print(f"\n{'='*80}")
    print("FINAL SUMMARY")
    print(f"{'='*80}")
    for fc, r in results.items():
        if r and r["success"]:
            print(f"  {fc}: PASS (OK={r['ok']}, NG={r['ng']})")
        elif r:
            print(f"  {fc}: FAIL (no OK/NG)")
        else:
            print(f"  {fc}: ERROR")
