"""Service for previewing spec imports - validation, parsing, and diff without committing."""
import logging
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models import FormType, FormSpec, SpecItem
from parsers.identifier import identify_form_type
from services.fingerprint_service import generate_fingerprint, compare_fingerprints
from services.ai_spec_parser import ai_parse_summary_sheet, validate_summary_structure
from services.spec_version_service import compute_diff
from services.spec_file_service import compute_file_hash, find_duplicate
from utils.spec_parser import parse_spec_string

logger = logging.getLogger(__name__)

# Built-in form codes that have hardcoded 汇总 parsers
BUILTIN_PARSERS = {"F-QA1021", "F-RD09AA", "F-RD09AB", "F-RD09AJ", "F-RD09AK"}


def preview_import(db: Session, filepath: str, file_content: bytes,
                   form_code: str, original_filename: str) -> dict:
    """Preview what an import would do WITHOUT committing changes.

    Returns a full preview with validation, parsed specs, and diffs.
    """
    file_hash = compute_file_hash(file_content)

    # Load workbook
    wb = load_workbook(filepath, data_only=True)

    try:
        result = {
            "form_code": form_code,
            "file_hash": file_hash,
            "original_filename": original_filename,
            "file_validation": _validate_file(db, wb, filepath, original_filename, form_code, file_hash),
            "structure_validation": {"valid": True, "warnings": []},
            "parsed_specs": [],
            "parse_method": None,
            "ai_confidence": None,
        }

        # Check if 汇总 sheet exists
        if "汇总" not in wb.sheetnames:
            # Fallback: extract specs from data sheet headers
            from services.header_spec_extractor import extract_specs_from_headers

            form_type = db.query(FormType).filter(
                FormType.form_code == form_code
            ).first()
            if not form_type:
                result["structure_validation"] = {
                    "valid": False,
                    "warnings": [f"Form type {form_code} not found"],
                }
                return result

            header_specs = extract_specs_from_headers(
                wb, form_code, form_type.form_name
            )
            # Determine which method succeeded and the parsed specs
            parsed_specs = None
            parse_method = None
            parse_warning = None

            if header_specs and any(eq.get("items") for eq in header_specs):
                parsed_specs = header_specs
                parse_method = "header"
                parse_warning = "此檔案無匯總 sheet，已從資料表頭自動提取檢查項目"
            else:
                # 3rd fallback: AI analysis of data sheet
                from services.ai_spec_parser import ai_parse_data_sheet
                data_sheets = [s for s in wb.sheetnames if s != "汇总"]
                if data_sheets:
                    ai_result = ai_parse_data_sheet(
                        wb[data_sheets[0]], form_code, form_type.form_name
                    )
                    if ai_result:
                        ai_specs = []
                        for eq in ai_result.get("equipment_specs", []):
                            items = []
                            for i, item in enumerate(eq.get("items", [])):
                                items.append({
                                    "item_name": item.get("item_name", f"item_{i}"),
                                    "spec_value": str(item.get("spec_value", "")),
                                    "parsed_spec": item.get("parsed_spec", {}),
                                    "group_name": item.get("group_name"),
                                    "sub_group": item.get("sub_group"),
                                    "display_order": item.get("display_order", i),
                                })
                            ai_specs.append({
                                "equipment_id": eq.get("equipment_id", "UNIVERSAL"),
                                "equipment_name": eq.get("equipment_name", "UNIVERSAL"),
                                "items": items,
                            })
                        if ai_specs and any(eq.get("items") for eq in ai_specs):
                            parsed_specs = ai_specs
                            parse_method = "ai"
                            result["ai_confidence"] = ai_result.get("confidence")
                            parse_warning = "此檔案無匯總 sheet，已透過 AI 分析資料表提取檢查項目"

            if parsed_specs:
                result["structure_validation"] = {
                    "valid": True,
                    "warnings": [parse_warning],
                }
                result["parse_method"] = parse_method
                result["parsed_specs"] = _compute_diffs(db, form_type, parsed_specs)

                from services.spec_file_service import check_specs_identical
                content_identical = check_specs_identical(
                    db, form_code, parsed_specs
                )
                result["content_identical"] = content_identical
                if content_identical:
                    result["file_validation"]["warnings"].append(
                        "此檔案的規格內容與現有資料完全相同，無需重複匯入"
                    )
                result["is_blocked"] = (
                    not result["structure_validation"]["valid"]
                    or len(result["parsed_specs"]) == 0
                    or result["file_validation"]["is_duplicate"]
                    or content_identical
                )
                return result
            else:
                result["structure_validation"] = {
                    "valid": False,
                    "warnings": [
                        "此檔案無匯總 sheet，規則提取與 AI 分析均無法提取檢查項目"
                    ],
                }
                return result

        ws = wb["汇总"]

        # Structure validation
        result["structure_validation"] = validate_summary_structure(ws)

        # Parse specs based on type
        form_type = db.query(FormType).filter(FormType.form_code == form_code).first()
        if not form_type:
            result["structure_validation"]["warnings"].append(f"Form type {form_code} not found")
            result["structure_validation"]["valid"] = False
            return result

        if form_code in BUILTIN_PARSERS:
            result["parse_method"] = "builtin"
            parsed_specs = _preview_builtin(db, ws, form_type)
        else:
            result["parse_method"] = "ai"
            parsed_specs, confidence = _preview_ai(ws, form_type)
            result["ai_confidence"] = confidence

        if parsed_specs is None:
            result["structure_validation"]["warnings"].append("Failed to parse specs from summary sheet")
            result["structure_validation"]["valid"] = False
            return result

        # Compute diffs for each equipment
        result["parsed_specs"] = _compute_diffs(db, form_type, parsed_specs)

        # Check if spec content is identical to existing DB specs
        from services.spec_file_service import check_specs_identical
        content_identical = check_specs_identical(db, form_code, parsed_specs)
        result["content_identical"] = content_identical
        if content_identical:
            result["file_validation"]["warnings"].append(
                "此檔案的規格內容與現有資料完全相同，無需重複匯入"
            )

        # Aggregate blocking conditions
        result["is_blocked"] = (
            not result["structure_validation"]["valid"]
            or len(result["parsed_specs"]) == 0
            or result["file_validation"]["is_duplicate"]
            or content_identical
        )

        return result
    finally:
        wb.close()


def _validate_file(db: Session, wb, filepath: str, original_filename: str,
                   target_form_code: str, file_hash: str) -> dict:
    """Validate the uploaded file against the target form type."""
    warnings = []

    # 1. Check if file matches the target form type via identification
    sheet_names = wb.sheetnames
    sheet_contents = {}
    for sn in sheet_names:
        if sn == "汇总":
            continue
        ws = wb[sn]
        text_parts = []
        for row in range(1, min(10, (ws.max_row or 0) + 1)):
            for col in range(1, min(20, (ws.max_column or 0) + 1)):
                val = ws.cell(row=row, column=col).value
                if val:
                    text_parts.append(str(val))
        sheet_contents[sn] = " ".join(text_parts)

    detected_code = identify_form_type(original_filename, sheet_names, sheet_contents, db=db)
    matches = detected_code == target_form_code if detected_code else None

    if detected_code and not matches:
        warnings.append(f"File appears to be '{detected_code}' but importing into '{target_form_code}'")

    # 2. Check for duplicate file
    duplicate = find_duplicate(db, target_form_code, file_hash)
    is_duplicate = duplicate is not None

    if is_duplicate:
        warnings.append(
            f"This exact file was previously imported "
            f"(equipment: {duplicate['equipment_id']}, time: {duplicate['created_at']})"
        )

    # 3. Check structural fingerprint
    fingerprint_similarity = None
    form_type = db.query(FormType).filter(FormType.form_code == target_form_code).first()
    if form_type and form_type.structural_fingerprint and "汇总" in wb.sheetnames:
        ws = wb["汇总"]
        new_fp = generate_fingerprint(ws)
        fp_result = compare_fingerprints(form_type.structural_fingerprint, new_fp)
        fingerprint_similarity = fp_result["similarity"]
        if not fp_result["match"]:
            warnings.append("File structure differs from previously imported files")
        warnings.extend(fp_result.get("warnings", []))

    return {
        "matches_form_type": matches,
        "detected_form_code": detected_code,
        "is_duplicate": is_duplicate,
        "duplicate_info": duplicate,
        "fingerprint_similarity": fingerprint_similarity,
        "warnings": warnings,
    }


def _preview_builtin(db: Session, ws, form_type) -> list[dict] | None:
    """Parse specs using built-in parser logic (read-only, no DB writes).

    Returns list of equipment spec dicts with items.
    """
    from services.spec_service import _cell_val
    import re

    form_code = form_type.form_code
    results = []

    if form_code == "F-QA1021":
        items = [
            {"item_name": "电源开关", "spec_value": "√", "group_name": "离子风扇"},
            {"item_name": "风扇飘带", "spec_value": "√", "group_name": "离子风扇"},
            {"item_name": "空气滤网", "spec_value": "√", "group_name": "离子风扇"},
            {"item_name": "风扇角度", "spec_value": "√", "group_name": "离子风扇"},
            {"item_name": "风扇风速", "spec_value": "√", "group_name": "离子风扇"},
        ]
        for i, item in enumerate(items):
            parsed = parse_spec_string(item["spec_value"])
            item["parsed_spec"] = parsed
            item["display_order"] = i
            item["sub_group"] = None
        results.append({
            "equipment_id": "RD-LZ-XX",
            "equipment_name": "通用离子消散设备",
            "items": items,
        })

    elif form_code == "F-RD09AA":
        row = 1
        while row <= (ws.max_row or 0):
            val = _cell_val(ws, row, 1)
            if val and "机台" in str(val):
                match = re.search(r"(WP\w+-\d+)", str(val))
                if match:
                    machine_id = match.group(1)
                    spec_row = row + 4
                    items = []
                    order = 0
                    params = [(4, "合模压力(ton)"), (5, "注塑压强(kgf/cm²)"),
                              (6, "固化时间(sec)"), (7, "注塑时间(sec)"), (8, "预热台温度(℃)")]
                    for col, name in params:
                        spec_str = _cell_val(ws, spec_row, col)
                        if spec_str:
                            items.append({
                                "item_name": name, "spec_value": str(spec_str),
                                "parsed_spec": parse_spec_string(str(spec_str)),
                                "group_name": "参数", "sub_group": None, "display_order": order,
                            })
                            order += 1
                    for offset, pos in enumerate(["上模", "下模"]):
                        r = spec_row + offset
                        set_val = _cell_val(ws, r, 11)
                        if set_val:
                            items.append({
                                "item_name": f"模温设定值({pos})", "spec_value": str(set_val),
                                "parsed_spec": parse_spec_string(str(set_val)),
                                "group_name": "模温", "sub_group": pos, "display_order": order,
                            })
                            order += 1
                        for i, col in enumerate([12, 13, 14, 15]):
                            disp_val = _cell_val(ws, r, col)
                            if disp_val:
                                items.append({
                                    "item_name": f"模温显示值{i+1}({pos})", "spec_value": str(disp_val),
                                    "parsed_spec": parse_spec_string(str(disp_val)),
                                    "group_name": "模温", "sub_group": pos, "display_order": order,
                                })
                                order += 1
                    results.append({
                        "equipment_id": machine_id,
                        "equipment_name": f"机台{machine_id}",
                        "items": items,
                    })
            row += 1

    elif form_code == "F-RD09AB":
        row = 1
        while row <= (ws.max_row or 0):
            val_a = _cell_val(ws, row, 1)
            val_b = _cell_val(ws, row, 2)
            if val_a and "机台" in str(val_a) and val_b:
                machine_id = str(val_b).strip()
                data_start = row + 4
                r = data_start
                while r <= (ws.max_row or 0):
                    reason = _cell_val(ws, r, 2)
                    if not reason:
                        break
                    method = _cell_val(ws, r, 3)
                    wash_key = f"{machine_id}_{reason}_{method}"
                    items = []
                    order = 0
                    for col, name in [(7, "合模压力(ton)"), (8, "注塑压强(kgf/cm²)")]:
                        spec_str = _cell_val(ws, r, col)
                        if spec_str:
                            items.append({"item_name": name, "spec_value": str(spec_str),
                                          "parsed_spec": parse_spec_string(str(spec_str)),
                                          "group_name": "参数", "sub_group": None, "display_order": order})
                            order += 1
                    for offset, pos in enumerate(["上模", "下模"]):
                        cur_r = r + offset
                        set_val = _cell_val(ws, cur_r, 10)
                        if set_val:
                            items.append({"item_name": f"模温设定值({pos})", "spec_value": str(set_val),
                                          "parsed_spec": parse_spec_string(str(set_val)),
                                          "group_name": "模温", "sub_group": pos, "display_order": order})
                            order += 1
                        for i, col in enumerate([11, 12, 13, 14]):
                            disp_val = _cell_val(ws, cur_r, col)
                            if disp_val:
                                items.append({"item_name": f"模温显示值{i+1}({pos})", "spec_value": str(disp_val),
                                              "parsed_spec": parse_spec_string(str(disp_val)),
                                              "group_name": "模温", "sub_group": pos, "display_order": order})
                                order += 1
                    for i, col in enumerate([15, 16, 17, 18, 19]):
                        app_val = _cell_val(ws, r, col)
                        if app_val:
                            items.append({"item_name": f"外观确认{i+1}", "spec_value": str(app_val),
                                          "parsed_spec": parse_spec_string(str(app_val)),
                                          "group_name": "外观", "sub_group": None, "display_order": order})
                            order += 1
                    for col, name in [(20, "模具状态"), (21, "定位针状态")]:
                        status_val = _cell_val(ws, r, col)
                        if status_val:
                            items.append({"item_name": name, "spec_value": str(status_val),
                                          "parsed_spec": parse_spec_string(str(status_val)),
                                          "group_name": "状态", "sub_group": None, "display_order": order})
                            order += 1
                    results.append({"equipment_id": wash_key, "equipment_name": f"{machine_id} {reason}-{method}", "items": items})
                    r += 2
                row = r
            else:
                row += 1

    elif form_code == "F-RD09AJ":
        row = 1
        while row <= (ws.max_row or 0):
            val = _cell_val(ws, row, 1)
            if val and "焊接炉编号" in str(val):
                furnace_id = str(_cell_val(ws, row, 2) or "").strip()
                if not furnace_id:
                    row += 1
                    continue
                spec_row = row + 4
                items = []
                order = 0
                for i, col in enumerate(range(2, 10)):
                    spec_str = _cell_val(ws, spec_row, col)
                    if spec_str:
                        items.append({"item_name": f"温区{i+1}设定SV(℃)", "spec_value": str(spec_str),
                                      "parsed_spec": parse_spec_string(str(spec_str)),
                                      "group_name": "温度设定", "sub_group": f"温区{i+1}", "display_order": order})
                        order += 1
                for i, col in enumerate(range(10, 18)):
                    spec_str = _cell_val(ws, spec_row, col)
                    if spec_str:
                        items.append({"item_name": f"温区{i+1}实际PV(℃)", "spec_value": str(spec_str),
                                      "parsed_spec": parse_spec_string(str(spec_str)),
                                      "group_name": "实际温度", "sub_group": f"温区{i+1}", "display_order": order})
                        order += 1
                gas_names_row = row + 2
                for col in range(18, 24):
                    gas_name = _cell_val(ws, gas_names_row, col)
                    spec_str = _cell_val(ws, spec_row, col)
                    if spec_str:
                        label = str(gas_name) if gas_name else f"氮气{col-17}"
                        items.append({"item_name": label, "spec_value": str(spec_str),
                                      "parsed_spec": parse_spec_string(str(spec_str)),
                                      "group_name": "氮气", "sub_group": None, "display_order": order})
                        order += 1
                cooling_spec = _cell_val(ws, spec_row, 24)
                if cooling_spec:
                    items.append({"item_name": "冷却水流量LPM", "spec_value": str(cooling_spec),
                                  "parsed_spec": parse_spec_string(str(cooling_spec)),
                                  "group_name": "冷却水", "sub_group": None, "display_order": order})
                    order += 1
                results.append({"equipment_id": furnace_id, "equipment_name": f"焊接炉{furnace_id}", "items": items})
                row = spec_row + 1
            else:
                row += 1

    elif form_code == "F-RD09AK":
        row = 1
        while row <= (ws.max_row or 0):
            val = _cell_val(ws, row, 1)
            if val and "Package" in str(val):
                package_match = re.search(r"Package[：:]\s*(\S+)", str(val))
                package = package_match.group(1) if package_match else ""
                machine_row = row + 1
                machine_val = _cell_val(ws, machine_row, 1)
                machine_match = re.search(r"(WTFB-\d+)", str(machine_val) if machine_val else "")
                machine_id = machine_match.group(1) if machine_match else ""
                if not machine_id:
                    row += 1
                    continue
                spec_key = f"{machine_id}_{package}"
                num_row = row + 3
                meas_count = 0
                for col in range(3, 30):
                    nv = _cell_val(ws, num_row, col)
                    if nv and str(nv).strip().isdigit():
                        meas_count = int(str(nv).strip())
                    else:
                        break
                items = []
                order = 0
                data_start = row + 4
                for part_row in range(data_start, data_start + 5):
                    if part_row > (ws.max_row or 0):
                        break
                    part_name = _cell_val(ws, part_row, 2)
                    if not part_name:
                        continue
                    spec_str = _cell_val(ws, part_row, 3)
                    if spec_str:
                        for meas_num in range(1, meas_count + 1):
                            items.append({"item_name": f"meas_{meas_num}", "spec_value": str(spec_str),
                                          "parsed_spec": parse_spec_string(str(spec_str)),
                                          "group_name": "测量", "sub_group": str(part_name), "display_order": order})
                            order += 1
                results.append({"equipment_id": spec_key, "equipment_name": f"{machine_id} {package}", "items": items})
                row = data_start + 5
            else:
                row += 1

    return results


def _preview_ai(ws, form_type) -> tuple[list[dict] | None, float | None]:
    """Parse specs using AI and return (specs, confidence)."""
    result = ai_parse_summary_sheet(ws, form_type.form_code, form_type.form_name)
    if not result:
        return None, None

    specs = []
    for eq in result.get("equipment_specs", []):
        items = []
        for i, item in enumerate(eq.get("items", [])):
            items.append({
                "item_name": item.get("item_name", f"item_{i}"),
                "spec_value": str(item.get("spec_value", "")),
                "parsed_spec": item.get("parsed_spec", {}),
                "group_name": item.get("group_name"),
                "sub_group": item.get("sub_group"),
                "display_order": item.get("display_order", i),
            })
        specs.append({
            "equipment_id": eq.get("equipment_id", "UNKNOWN"),
            "equipment_name": eq.get("equipment_name", eq.get("equipment_id", "UNKNOWN")),
            "items": items,
        })

    return specs, result.get("confidence")


def _compute_diffs(db: Session, form_type, parsed_specs: list[dict]) -> list[dict]:
    """For each equipment, compute diff between existing and new specs."""
    result = []
    for eq_spec in parsed_specs:
        eq_id = eq_spec["equipment_id"]

        # Find existing spec
        existing = db.query(FormSpec).filter(
            FormSpec.form_type_id == form_type.id,
            FormSpec.equipment_id == eq_id,
        ).first()

        new_items = []
        for item in eq_spec.get("items", []):
            parsed = item.get("parsed_spec", {})
            new_items.append({
                "item_name": item["item_name"],
                "spec_type": parsed.get("spec_type", "text"),
                "min_value": parsed.get("min_value"),
                "max_value": parsed.get("max_value"),
                "expected_text": parsed.get("expected_text"),
                "threshold_value": parsed.get("threshold_value"),
                "threshold_operator": parsed.get("threshold_operator"),
                "group_name": item.get("group_name"),
                "sub_group": item.get("sub_group"),
                "display_order": item.get("display_order", 0),
            })

        if existing:
            # Get existing items for diff
            existing_items = db.query(SpecItem).filter(
                SpecItem.form_spec_id == existing.id
            ).order_by(SpecItem.display_order).all()

            old_items = [
                {
                    "item_name": item.item_name,
                    "spec_type": item.spec_type,
                    "min_value": float(item.min_value) if item.min_value is not None else None,
                    "max_value": float(item.max_value) if item.max_value is not None else None,
                    "expected_text": item.expected_text,
                    "threshold_value": float(item.threshold_value) if item.threshold_value is not None else None,
                    "threshold_operator": item.threshold_operator,
                    "group_name": item.group_name,
                    "sub_group": item.sub_group,
                    "display_order": item.display_order,
                }
                for item in existing_items
            ]

            diff = compute_diff(old_items, new_items)
        else:
            diff = {
                "added": new_items,
                "removed": [],
                "modified": [],
                "unchanged": [],
                "summary": {"added": len(new_items), "removed": 0, "modified": 0, "unchanged": 0},
            }

        result.append({
            "equipment_id": eq_id,
            "equipment_name": eq_spec.get("equipment_name", eq_id),
            "is_new": existing is None,
            "existing_item_count": len(old_items) if existing else 0,
            "new_item_count": len(new_items),
            "diff": diff,
            "items": new_items,
        })

    return result
